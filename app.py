import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import xlrd, datetime, io, zipfile, base64, os, json
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.colors import white, black

# Server host (Streamlit Cloud...) chạy giờ UTC, khách sạn ở Việt Nam (UTC+7,
# không có giờ mùa hè) — mọi mốc giờ hiển thị/ghi log trong app PHẢI dùng
# now_vn()/today_vn() bên dưới, KHÔNG dùng thẳng datetime.datetime.now() hay
# datetime.date.today() (trả về giờ server, sai múi giờ thực tế của khách sạn).
VN_TZ = datetime.timezone(datetime.timedelta(hours=7))
def now_vn():
    return datetime.datetime.now(VN_TZ)
def today_vn():
    return now_vn().date()

# ── Chế độ giao diện (sáng/tối) — chọn tay hoặc tự động theo giờ Việt Nam.
# Mốc 6h-18h coi là ban ngày (giao diện sáng), ngoài khoảng đó là ban đêm
# (giao diện tối) — có thể chỉnh 2 số này nếu muốn đổi mốc giờ.
THEME_DAY_START_HOUR = 6
THEME_NIGHT_START_HOUR = 18

def _compute_effective_theme():
    mode = st.session_state.get('theme_mode', 'auto')
    if mode in ('light', 'dark'):
        return mode
    h = now_vn().hour
    return 'light' if THEME_DAY_START_HOUR <= h < THEME_NIGHT_START_HOUR else 'dark'

# ── Âm lịch Việt Nam ──────────────────────────────────────────────────────
# Dùng để biết hôm nay có phải Tết hay Trung Thu — 2 mốc này theo âm lịch nên
# KHÔNG thể suy ra từ ngày dương. Thuật toán thiên văn tiêu chuẩn (Hồ Ngọc Đức),
# múi giờ +7; tính thẳng nên không cần bảng tra ngày lễ phải cập nhật hằng năm.
def _jd_from_date(dd, mm, yy):
    a = (14 - mm) // 12
    y, m = yy + 4800 - a, mm + 12 * a - 3
    jd = dd + (153 * m + 2) // 5 + 365 * y + y // 4 - y // 100 + y // 400 - 32045
    if jd < 2299161:
        jd = dd + (153 * m + 2) // 5 + 365 * y + y // 4 - 32083
    return jd

def _new_moon_jd(k):
    import math
    T = k / 1236.85
    T2, dr = T * T, math.pi / 180
    T3 = T2 * T
    jd1 = 2415020.75933 + 29.53058868 * k + 0.0001178 * T2 - 0.000000155 * T3
    jd1 += 0.00033 * math.sin((166.56 + 132.87 * T - 0.009173 * T2) * dr)
    M = 359.2242 + 29.10535608 * k - 0.0000333 * T2 - 0.00000347 * T3
    Mpr = 306.0253 + 385.81691806 * k + 0.0107306 * T2 + 0.00001236 * T3
    F = 21.2964 + 390.67050646 * k - 0.0016528 * T2 - 0.00000239 * T3
    c1 = (0.1734 - 0.000393 * T) * math.sin(M * dr) + 0.0021 * math.sin(2 * dr * M)
    c1 -= 0.4068 * math.sin(Mpr * dr) + 0.0161 * math.sin(dr * 2 * Mpr)
    c1 -= 0.0004 * math.sin(dr * 3 * Mpr)
    c1 += 0.0104 * math.sin(dr * 2 * F) - 0.0051 * math.sin(dr * (M + Mpr))
    c1 -= 0.0074 * math.sin(dr * (M - Mpr)) + 0.0004 * math.sin(dr * (2 * F + M))
    c1 -= 0.0004 * math.sin(dr * (2 * F - M)) - 0.0006 * math.sin(dr * (2 * F + Mpr))
    c1 += 0.0010 * math.sin(dr * (2 * F - Mpr)) + 0.0005 * math.sin(dr * (2 * Mpr + M))
    if T < -11:
        dt = (0.001 + 0.000839 * T + 0.0002261 * T2
              - 0.00000845 * T3 - 0.000000081 * T * T3)
    else:
        dt = -0.000278 + 0.000265 * T + 0.000262 * T2
    return jd1 + c1 - dt

def _new_moon_day(k, tz=7):
    return int(_new_moon_jd(k) + 0.5 + tz / 24)

def _sun_longitude_deg6(jdn, tz=7):
    import math
    t = (jdn - 0.5 - tz / 24 - 2451545.0) / 36525
    t2, dr = t * t, math.pi / 180
    M = 357.52910 + 35999.05030 * t - 0.0001559 * t2 - 0.00000048 * t * t2
    L0 = 280.46645 + 36000.76983 * t + 0.0003032 * t2
    dl = (1.914600 - 0.004817 * t - 0.000014 * t2) * math.sin(dr * M)
    dl += (0.019993 - 0.000101 * t) * math.sin(dr * 2 * M) + 0.000290 * math.sin(dr * 3 * M)
    L = (L0 + dl) * dr
    L -= math.pi * 2 * int(L / (math.pi * 2))
    return int(L / math.pi * 6)

def _lunar_month11(yy, tz=7):
    off = _jd_from_date(31, 12, yy) - 2415021
    k = int(off / 29.530588853)
    nm = _new_moon_day(k, tz)
    return _new_moon_day(k - 1, tz) if _sun_longitude_deg6(nm, tz) >= 9 else nm

def _leap_month_offset(a11, tz=7):
    k = int((a11 - 2415021.076998695) / 29.530588853 + 0.5)
    i = 1
    arc = _sun_longitude_deg6(_new_moon_day(k + i, tz), tz)
    while True:
        last, i = arc, i + 1
        arc = _sun_longitude_deg6(_new_moon_day(k + i, tz), tz)
        if arc == last or i >= 14:
            break
    return i - 1

def solar_to_lunar(d, tz=7):
    """Ngày dương (datetime.date) -> (ngày, tháng, năm, có_nhuận) âm lịch."""
    day_number = _jd_from_date(d.day, d.month, d.year)
    k = int((day_number - 2415021.076998695) / 29.530588853)
    month_start = _new_moon_day(k + 1, tz)
    if month_start > day_number:
        month_start = _new_moon_day(k, tz)
    a11 = _lunar_month11(d.year, tz)
    b11 = a11
    if a11 >= month_start:
        lunar_year = d.year
        a11 = _lunar_month11(d.year - 1, tz)
    else:
        lunar_year = d.year + 1
        b11 = _lunar_month11(d.year + 1, tz)
    lunar_day = day_number - month_start + 1
    diff = int((month_start - a11) / 29)
    lunar_leap, lunar_month = 0, diff + 11
    if b11 - a11 > 365:
        leap_off = _leap_month_offset(a11, tz)
        if diff >= leap_off:
            lunar_month = diff + 10
            if diff == leap_off:
                lunar_leap = 1
    if lunar_month > 12:
        lunar_month -= 12
    if lunar_month >= 11 and diff < 4:
        lunar_year -= 1
    return lunar_day, lunar_month, lunar_year, lunar_leap

# ── Mùa / dịp lễ của màn chào ─────────────────────────────────────────────
# Ngày lễ (theo âm lịch hoặc ngày cố định) được ưu tiên ĐÈ LÊN mùa.
SEASON_LABEL = {
    'spring':   '🌸 TIẾT XUÂN',   'summer': '🌊 TIẾT HẠ',
    'autumn':   '🍂 TIẾT THU',    'winter': '🌧️ TIẾT ĐÔNG',
    'tet':      '🎋 TẾT NGUYÊN ĐÁN',
    'trungthu': '🏮 TRUNG THU',   'noel':   '🎄 GIÁNG SINH',
}

def _season_key(d=None):
    d = d or today_vn()
    if d.month == 12 and 20 <= d.day <= 26:
        return 'noel'
    try:
        ld, lm, _, leap = solar_to_lunar(d)
        if not leap:
            if lm == 1 and ld <= 7:          # mùng 1 -> mùng 7 Tết
                return 'tet'
            if lm == 12 and ld >= 28:        # 28, 29, 30 Tết
                return 'tet'
            if lm == 8 and 14 <= ld <= 16:   # rằm tháng 8
                return 'trungthu'
    except Exception:
        pass                                  # lỗi lịch không được làm hỏng màn chào
    return ('spring' if d.month <= 3 else 'summer' if d.month <= 6
            else 'autumn' if d.month <= 9 else 'winter')

def _hour_key(h=None):
    h = now_vn().hour if h is None else h
    return ('dawn' if 5 <= h < 8 else 'day' if 8 <= h < 16
            else 'dusk' if 16 <= h < 19 else 'night')

# ── Tiến độ ca làm việc — lưu trên đĩa server để SỐNG SÓT qua việc tải lại
# trang (F5) trong ngày. LƯU Ý: file này KHÔNG bền vững qua các lần deploy lại
# app (Streamlit Cloud xóa filesystem mỗi lần deploy) — chỉ chống việc mất dữ
# liệu do reload trang trong 1 ngày làm việc, không thay thế backup lâu dài.
# Chỉ lưu tiến độ + số liệu TỔNG HỢP (không tên/hộ chiếu khách) cho các công cụ
# xử lý dữ liệu khách; RIÊNG Sổ giao ca lưu đầy đủ nội dung vì đó chính là mục
# đích của sổ giao ca (thông tin cần truyền lại nguyên vẹn cho ca sau).
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

def _progress_path(date_iso=None):
    os.makedirs(DATA_DIR, exist_ok=True)
    return os.path.join(DATA_DIR, f'progress_{date_iso or today_vn().isoformat()}.json')

def _default_progress():
    return {'date': today_vn().isoformat(), 'tasks': {}, 'handover_entries': []}

def _load_progress():
    p = _progress_path()
    if os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if data.get('date') == today_vn().isoformat():
                return data
        except Exception:
            pass
    return _default_progress()

def _yesterday_total():
    """Tổng khách HÔM QUA đọc từ file tiến độ ngày hôm trước — dùng để so sánh
    trên thẻ tổng quan. Không có file (hoặc hôm qua chưa chạy công cụ) thì trả
    None và thẻ sẽ không hiện dòng so sánh, KHÔNG suy đoán số."""
    try:
        p = _progress_path((today_vn() - datetime.timedelta(days=1)).isoformat())
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return ((data.get('tasks', {}) or {}).get('daily', {}) or {}).get('summary', {}).get('total')
    except Exception:
        pass
    return None

def _atomic_write_json(path, data):
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def _progress_update(mutate_fn):
    """Đọc file tiến độ MỚI NHẤT từ đĩa (không phải bản trong session_state) rồi
    mới sửa và ghi lại — giảm rủi ro 2 tab/phiên trong ngày ghi đè mất dữ liệu
    của nhau. Đồng bộ luôn vào session_state để hiển thị ngay trong lượt chạy
    hiện tại."""
    state = _load_progress()
    mutate_fn(state)
    state['last_updated'] = now_vn().strftime('%H:%M:%S')
    try:
        _atomic_write_json(_progress_path(), state)
    except Exception:
        pass  # đĩa lỗi/không ghi được không được làm crash app — tính năng chỉ là tiện ích
    st.session_state.progress = state
    return state

# ── Sổ giao ca — lưu trữ đám mây (Supabase/Postgres), BỀN VỮNG qua mọi lần
# deploy lại (khác với data/progress_*.json ở trên chỉ sống qua 1 ngày). Cần
# cấu hình st.secrets["connections"]["supabase_db"]["url"] — xem hướng dẫn
# trong secrets.toml.example. NẾU CHƯA CẤU HÌNH: mọi hàm db_* trả về None/rỗng
# một cách an toàn, màn Sổ giao ca tự động dùng lại lưu tạm trên đĩa (không
# bền vững qua deploy) — app KHÔNG bao giờ crash vì thiếu Supabase.
try:
    from sqlalchemy import text as _sql_text
except Exception:
    _sql_text = None

def _redact_db_error(e):
    """Ẩn mật khẩu (nếu lỡ lọt vào chuỗi kết nối trong thông báo lỗi của
    SQLAlchemy) trước khi hiển thị cho người dùng — tránh lộ secret lên UI."""
    import re as _re_local
    return _re_local.sub(r'://([^:/@\s]+):([^@\s]+)@', r'://\1:***@', str(e))

def _get_db():
    if _sql_text is None:
        st.session_state['_db_last_error'] = ("Thiếu thư viện SQLAlchemy/psycopg2-binary trong môi trường chạy — "
                                               "vào Manage app → Reboot app để cài lại requirements.txt.")
        return None
    try:
        return st.connection("supabase_db", type="sql")
    except Exception as e:
        st.session_state['_db_last_error'] = _redact_db_error(e)
        return None

@st.cache_resource(show_spinner=False)
def _db_schema_ready():
    """Tạo bảng shift_handover nếu chưa có — chỉ chạy 1 lần mỗi phiên server
    (cache_resource), không phải mỗi lần rerun."""
    conn = _get_db()
    if conn is None:
        return False
    try:
        with conn.session as s:
            s.execute(_sql_text("""
                CREATE TABLE IF NOT EXISTS shift_handover (
                    id BIGSERIAL PRIMARY KEY,
                    shift_date DATE NOT NULL,
                    entry_time TEXT NOT NULL,
                    category TEXT NOT NULL,
                    room TEXT,
                    note TEXT NOT NULL,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT now()
                )
            """))
            s.execute(_sql_text(
                "CREATE INDEX IF NOT EXISTS idx_shift_handover_date ON shift_handover(shift_date)"))
            s.commit()
        return True
    except Exception as e:
        st.session_state['_db_last_error'] = _redact_db_error(e)
        return False

def db_available():
    return _get_db() is not None and _db_schema_ready()

def db_add_entry(shift_date, entry_time, category, room, note):
    conn = _get_db()
    with conn.session as s:
        s.execute(_sql_text("""INSERT INTO shift_handover (shift_date, entry_time, category, room, note)
                               VALUES (:d, :t, :c, :r, :n)"""),
                  {'d': shift_date, 't': entry_time, 'c': category, 'r': room, 'n': note})
        s.commit()

def db_delete_entry(entry_id):
    conn = _get_db()
    with conn.session as s:
        s.execute(_sql_text("DELETE FROM shift_handover WHERE id = :id"), {'id': int(entry_id)})
        s.commit()

def db_load_entries(shift_date):
    """Trả về DataFrame [id, entry_time, category, room, note] cho 1 ngày, mới nhất trước."""
    conn = _get_db()
    return conn.query(
        "SELECT id, entry_time, category, room, note FROM shift_handover "
        "WHERE shift_date = :d ORDER BY entry_time DESC, id DESC",
        params={'d': shift_date}, ttl=0)

def db_load_dates(limit=180):
    """Danh sách các ngày đã có ghi chú (mới nhất trước) — phục vụ ô chọn ngày xem lại lịch sử."""
    conn = _get_db()
    df = conn.query(
        "SELECT DISTINCT shift_date FROM shift_handover ORDER BY shift_date DESC LIMIT :lim",
        params={'lim': limit}, ttl=0)
    return list(df['shift_date']) if not df.empty else []

def compute_day_summary(df_entries):
    """Tổng hợp tự động: đếm ghi chú theo phân loại + danh sách phòng được nhắc tới."""
    if df_entries is None or df_entries.empty:
        return {'total': 0, 'by_category': {}, 'rooms': []}
    by_cat = df_entries['category'].value_counts().to_dict()
    rooms = sorted(set(str(r).strip() for r in df_entries['room'].dropna() if str(r).strip()))
    return {'total': len(df_entries), 'by_category': by_cat, 'rooms': rooms}

# Load app icon (favicon) từ icon.b64
@st.cache_resource
def _load_app_icon():
    try:
        from PIL import Image
        p = os.path.join(os.path.dirname(__file__), 'icon.b64')
        with open(p, 'r') as f:
            raw = base64.b64decode(f.read())
        img = Image.open(io.BytesIO(raw))
        img.load()  # ép tải đầy đủ ảnh ngay tại đây — tránh lỗi lazy-load khi
                    # st.cache_resource dùng lại ảnh này ở ngữ cảnh/luồng khác
        return img
    except Exception:
        return "🌸"

st.set_page_config(page_title="Tân Hotel", page_icon=_load_app_icon(), layout="wide")

# ── Load embedded templates ──────────────────────────────────────────────
@st.cache_resource
def load_template(name):
    path = os.path.join(os.path.dirname(__file__), f'tmpl_{name}.b64')
    with open(path, 'r') as f:
        return base64.b64decode(f.read())

@st.cache_resource
def _dark_bg_data_uri():
    """Ảnh nền chế độ tối (mèo con ngủ) — nhúng thẳng base64 vào CSS, cùng
    kiểu với load_template() ở trên, không cần hosting/URL ngoài."""
    path = os.path.join(os.path.dirname(__file__), 'bg_dark.b64')
    with open(path, 'r') as f:
        return 'data:image/jpeg;base64,' + f.read().strip()

@st.cache_resource
def _light_bg_data_uri():
    """Ảnh nền chế độ sáng (mèo con chui trong túi giấy) — cùng cơ chế với
    _dark_bg_data_uri() ở trên."""
    path = os.path.join(os.path.dirname(__file__), 'bg_light.b64')
    with open(path, 'r') as f:
        return 'data:image/jpeg;base64,' + f.read().strip()

@st.cache_resource
def _season_bg_data_uri(key, theme):
    """Ảnh nền màn chào theo mùa/dịp lễ (bg_<key>.b64). Mùa nào CHƯA có ảnh
    riêng thì dùng lại ảnh nền sáng/tối sẵn có — không bao giờ để trống."""
    path = os.path.join(os.path.dirname(__file__), f'bg_{key}.b64')
    if os.path.exists(path):
        try:
            with open(path, 'r') as f:
                return 'data:image/jpeg;base64,' + f.read().strip()
        except Exception:
            pass
    return _dark_bg_data_uri() if theme == 'dark' else _light_bg_data_uri()

# ── Lookup tables ─────────────────────────────────────────────────────────
# Full nationality mapping (normalized keys → "CODE - Name") — 350 entries
import unicodedata as _ud, re as _re
def _norm_nat(s):
    s = str(s).lower().strip()
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return _re.sub(r'[^a-z0-9]', '', s)

NAT_NORM = {
    "achentina": "ARG - Argentina",
    "acmenia": "ARM - Armenia",
    "adecbaigian": "AZE - Azerbaijan",
    "aicap": "EGY - Egypt",
    "airolen": "IRL - Ireland",
    "aixolen": "ISL - Iceland",
    "albania": "ALB - Albania",
    "anbani": "ALB - Albania",
    "andorra": "AND - Andorra",
    "anggola": "AGO - Angola",
    "angola": "AGO - Angola",
    "anh": "GBR - United Kingdom",
    "anmach": "DNK - Denmark",
    "ano": "IND - India",
    "ao": "AUT - Austria",
    "aosip": "CYP - Cyprus",
    "arapthongnhat": "ARE - United Arab Emirates",
    "arapxaui": "SAU - Saudi Arabia",
    "argentina": "ARG - Argentina",
    "armenia": "ARM - Armenia",
    "australia": "AUS - Australia",
    "austria": "AUT - Austria",
    "azerbaijan": "AZE - Azerbaijan",
    "bacbaot": "BRB - Barbados",
    "bahama": "BHS - Bahamas",
    "bahamas": "BHS - Bahamas",
    "bahrain": "BHR - Bahrain",
    "balan": "POL - Poland",
    "bangladesh": "BGD - Bangladesh",
    "banglaet": "BGD - Bangladesh",
    "barain": "BHR - Bahrain",
    "barbados": "BRB - Barbados",
    "becmua": "BMU - Bermuda",
    "belarus": "BLR - Belarus",
    "belarut": "BLR - Belarus",
    "belgium": "BEL - Belgium",
    "belixe": "BLZ - Belize",
    "belize": "BLZ - Belize",
    "benanh": "BEN - Benin",
    "benin": "BEN - Benin",
    "bermuda": "BMU - Bermuda",
    "bhutan": "BTN - Bhutan",
    "bi": "BEL - Belgium",
    "boaonha": "PRT - Portugal",
    "bolivia": "BOL - Bolivia",
    "bosniaandherzegovina": "BIH - Bosnia and Herzegovina",
    "botswana": "BWA - Botswana",
    "botxoana": "BWA - Botswana",
    "boxniahecdegovina": "BIH - Bosnia and Herzegovina",
    "bradin": "BRA - Brazil",
    "brazil": "BRA - Brazil",
    "britishindiaoceanterritory": "IOT - British India Ocean Territory",
    "bulgaria": "BGR - Bulgaria",
    "bungari": "BGR - Bulgaria",
    "buockinaphaxo": "BFA - Burkina Faso",
    "burkinafaso": "BFA - Burkina Faso",
    "burundi": "BDI - Burundi",
    "buruni": "BDI - Burundi",
    "butan": "BTN - Bhutan",
    "cameroon": "CMR - Cameroon",
    "camorun": "CMR - Cameroon",
    "canada": "CAN - Canada",
    "capeverde": "CPV - Cape Verde",
    "capve": "CPV - Cape Verde",
    "chad": "TCD - Chad",
    "charapxyri": "SYR - Syrian Arab Republic",
    "chdcndtrieutien": "PRK - Korea Democratic Peoples Republic of",
    "chhanquoc": "KOR - Korea (South)",
    "chhoigiaoiran": "IRN - Iran Ilasmic Republic of",
    "chile": "CHL - Chile",
    "china": "CHN - China",
    "chinataiwan": "CHN - China",
    "chlienbanguc": "D - Germany",
    "chmaxeonia": "MKD - Macedonia",
    "chominicana": "DMA - Dominica",
    "colombia": "COL - Colombia",
    "como": "COM - Comoros",
    "comoros": "COM - Comoros",
    "conggo": "COG - Congo",
    "conghoasec": "CZE - Czech Republic",
    "congo": "COG - Congo",
    "congquocanora": "AND - Andorra",
    "congquoclichtenxten": "LIE - Liechtenstein",
    "congquocmonaco": "MCO - Monaco",
    "cooet": "KWT - Kuwait",
    "costarica": "CRI - Costa Rica",
    "cotedivoire": "CIV - Cote d' Ivoire",
    "cotivoa": "CIV - Cote d' Ivoire",
    "coxtarica": "CRI - Costa Rica",
    "croatia": "HRV - Croatia",
    "cuba": "CUB - Cuba",
    "cyprus": "CYP - Cyprus",
    "czechrepublic": "CZE - Czech Republic",
    "dambia": "ZMB - Zambia",
    "denmark": "DNK - Denmark",
    "dimbabue": "ZWE - Zimbabwe",
    "djibouti": "DJI - Djibouti",
    "dominica": "DMA - Dominica",
    "dominicana": "DMA - Dominica",
    "ecuador": "ECU - Ecuador",
    "ecuao": "ECU - Ecuador",
    "egypt": "EGY - Egypt",
    "elsalvador": "SLV - El Salvado",
    "enxanvao": "SLV - El Salvado",
    "equatorialguinea": "GNQ - Equatorial Guinea",
    "eritoria": "ERI - Eritrea",
    "eritrea": "ERI - Eritrea",
    "estonia": "EST - Estonia",
    "ethiopia": "ETH - Ethiopia",
    "etiopia": "ETH - Ethiopia",
    "extonia": "EST - Estonia",
    "fiji": "FJI - Fiji",
    "finland": "FIN - Finland",
    "france": "FRA - France",
    "francemetropolitan": "FRA - France",
    "gabon": "GAB - Gabon",
    "gabong": "GAB - Gabon",
    "gambia": "GMB - Gambia",
    "gana": "GHA - Ghana",
    "georgia": "GEO - Georgia",
    "germany": "D - Germany",
    "ghana": "GHA - Ghana",
    "ghine": "GIN - Guinea",
    "ghinebitxao": "GNB - Guinea-Bissau",
    "ghinexichao": "GNQ - Equatorial Guinea",
    "giamahiriiaaraplibinhandan": "LBY - Libyan Arab Jamahiriya",
    "gibraltar": "GIB - Gibraltar",
    "gibranta": "GIB - Gibraltar",
    "goatemala": "GTM - Guatemala",
    "greece": "GRC - Greece",
    "greenland": "GRL - Greenland",
    "grenaa": "GRD - Grenada",
    "grenada": "GRD - Grenada",
    "grinlon": "GRL - Greenland",
    "grudia": "GEO - Georgia",
    "guatemala": "GTM - Guatemala",
    "guina": "GUY - Guyana",
    "guinea": "GIN - Guinea",
    "guineabissau": "GNB - Guinea-Bissau",
    "guyana": "GUY - Guyana",
    "haiti": "HTI - Haiti",
    "halan": "NLD - Netherland",
    "hanquoc": "KOR - Korea (South)",
    "honduras": "HND - Honduras",
    "hondurat": "HND - Honduras",
    "hylap": "GRC - Greece",
    "iaphanthuoclienhiepanh": "GBD - United Kingdom British Territories Citizen",
    "ibouti": "DJI - Djibouti",
    "iceland": "ISL - Iceland",
    "india": "IND - India",
    "indonesia": "IDN - Indonesia",
    "inonexia": "IDN - Indonesia",
    "irac": "IRQ - Iraq",
    "iran": "IRN - Iran Ilasmic Republic of",
    "iraq": "IRQ - Iraq",
    "ireland": "IRL - Ireland",
    "israel": "ISR - Israel",
    "italia": "ITA - Italy",
    "italy": "ITA - Italy",
    "ixraen": "ISR - Israel",
    "jamaica": "JAM - Jamaica",
    "japan": "JPN - Japan",
    "jocan": "JOR - Jordan",
    "jordan": "JOR - Jordan",
    "kadacxtan": "KAZ - Kazakhstan",
    "kazakhstan": "KAZ - Kazakhstan",
    "kenia": "KEN - Kenya",
    "kenya": "KEN - Kenya",
    "kiecghidia": "KGZ - Kyrgyzstan",
    "kiribati": "KIR - Kiribati",
    "koreademocraticpeoplesrepublic": "PRK - Korea Democratic Peoples Republic of",
    "koreasouth": "KOR - Korea (South)",
    "kosovo": "RKS - Kosovo",
    "kuwait": "KWT - Kuwait",
    "kyrgyzstan": "KGZ - Kyrgyzstan",
    "latvia": "LVA - Latvia",
    "lebanon": "LBN - Lebanon",
    "lesotho": "LSO - Lesotho",
    "lexotho": "LSO - Lesotho",
    "liban": "LBN - Lebanon",
    "liberia": "LBR - Liberia",
    "libya": "LBY - Libyan Arab Jamahiriya",
    "liechtenstein": "LIE - Liechtenstein",
    "lienbangnga": "RUS - Russia",
    "lithuania": "LTU - Lithuania",
    "luxembourg": "LUX - Luxembourg",
    "luychxembua": "LUX - Luxembourg",
    "maagaxca": "MDG - Madagascar",
    "macedonia": "MKD - Macedonia",
    "madagascar": "MDG - Madagascar",
    "malaixia": "MYS - Malaysia",
    "malauy": "MWI - Malawi",
    "malawi": "MWI - Malawi",
    "malaysia": "MYS - Malaysia",
    "maldives": "MDV - Maldives",
    "mali": "MLI - Mali",
    "malta": "MLT - Malta",
    "manivo": "MDV - Maldives",
    "manta": "MLT - Malta",
    "maroc": "MAR - Morocco",
    "marshallislands": "MHL - Marshall Islands",
    "mauritania": "MRT - Mauritania",
    "mauritius": "MUS - Mauritius",
    "mexico": "MEX - Mexico",
    "mianma": "MMR - Myanmar",
    "micronesia": "FSM - Micronesia",
    "modambich": "MOZ - Mozambique",
    "moldova": "MDA - Moldova",
    "monaco": "MCO - Monaco",
    "mongco": "MNG - Mongolia",
    "mongolia": "MNG - Mongolia",
    "monova": "MDA - Moldova",
    "montenegro": "MNE - Montenegro",
    "montserrat": "MSR - Montserrat",
    "monxerat": "MSR - Montserrat",
    "moratani": "MRT - Mauritania",
    "morixo": "MUS - Mauritius",
    "morocco": "MAR - Morocco",
    "mozambique": "MOZ - Mozambique",
    "my": "USA - United States of America",
    "myanmarburma": "MMR - Myanmar",
    "namibia": "NAM - Namibia",
    "nauru": "NRU - Nauru",
    "nauy": "NOR - Norway",
    "nepal": "NPL - Nepal",
    "nepan": "NPL - Nepal",
    "netherland": "NLD - Netherland",
    "netherlandantilles": "NLD - Netherland",
    "newzealand": "NZL - New Zealand",
    "nhatban": "JPN - Japan",
    "nicaragoa": "NIC - Nicaragua",
    "nicaragua": "NIC - Nicaragua",
    "niger": "NER - Niger",
    "nigeria": "NGA - Nigeria",
    "nigie": "NER - Niger",
    "nigieria": "NGA - Nigeria",
    "niudilan": "NZL - New Zealand",
    "norway": "NOR - Norway",
    "oman": "OMN - Oman",
    "ominica": "DMA - Dominica",
    "ongtimo": "TLS - Timor Leste",
    "oxtraylia": "AUS - Australia",
    "pakistan": "PAK - Pakistan",
    "pakixtan": "PAK - Pakistan",
    "palau": "PLW - Palau",
    "palestine": "PSE - Palestine",
    "palextin": "PSE - Palestine",
    "panama": "PAN - Panama",
    "papuanewguinea": "PNG - Papua New Guinea",
    "papuaniughine": "PNG - Papua New Guinea",
    "paragoay": "PRY - Paraguay",
    "paraguay": "PRY - Paraguay",
    "peru": "PER - Peru",
    "phanlan": "FIN - Finland",
    "phap": "FRA - France",
    "philippin": "PHL - Philippines",
    "philippine": "PHL - Philippines",
    "poland": "POL - Poland",
    "portugal": "PRT - Portugal",
    "qatar": "QAT - Qatar",
    "quanaoantithuochalan": "NLD - Netherland",
    "quanaomacsan": "MHL - Marshall Islands",
    "quanaonamgrudiavanamsanuych": "GEO - Georgia",
    "quanaoxaysen": "SYC - Seychelles",
    "quata": "QAT - Qatar",
    "romania": "ROU - Romania",
    "ruana": "RWA - Rwanda",
    "rumani": "ROU - Romania",
    "russia": "RUS - Russia",
    "rwanda": "RWA - Rwanda",
    "saintlucia": "LCA - Saint Lucia",
    "sanmarino": "SMR - San Marino",
    "sat": "TCD - Chad",
    "saudiarabia": "SAU - Saudi Arabia",
    "scotland": "SC- - Scotland",
    "senegal": "SEN - Senegal",
    "serbia": "SRB - Serbia",
    "seychelles": "SYC - Seychelles",
    "singapore": "SGP - Singapore",
    "slovakia": "SVK - Slovakia",
    "slovenia": "SVN - Slovenia",
    "somalia": "SOM - Somalia",
    "southgeorgiaandthesouths": "GEO - Georgia",
    "spain": "ESP - Spain",
    "srilanka": "LKA - Sri Lanka",
    "sudan": "SDN - Sudan",
    "suriname": "SUR - Suriname",
    "swaziland": "SWZ - Swaziland",
    "sweden": "SWE - Sweden",
    "switzerland": "CHE - Switzerland",
    "syria": "SYR - Syrian Arab Republic",
    "tagikixtan": "TJK - Tajikistan",
    "tajikistan": "TJK - Tajikistan",
    "taybannha": "ESP - Spain",
    "thailan": "THA - Thailand",
    "thailand": "THA - Thailand",
    "thonhiky": "TUR - Turkey",
    "thuyien": "SWE - Sweden",
    "thuysi": "CHE - Switzerland",
    "timorleste": "TLS - Timor Leste",
    "tochucdantocthongnhat": "UNO - United Nations Organization",
    "togo": "TGO - Togo",
    "tonga": "TON - Tonga",
    "trungquoc": "CHN - China",
    "trungquocailoan": "CHN - China",
    "tunidi": "TUN - Tunisia",
    "tunisia": "TUN - Tunisia",
    "tuocmenixtan": "TKM - Turkmenistan",
    "turkey": "TUR - Turkey",
    "turkmenistan": "TKM - Turkmenistan",
    "tuvalu": "TUV - Tuvalu",
    "uc": "D - Germany",
    "ucraina": "UKR - Ukraine",
    "udobekixtan": "UZB - Uzbekistan",
    "uganda": "UGA - Uganda",
    "ukraine": "UKR - Ukraine",
    "unitedarabemirates": "ARE - United Arab Emirates",
    "unitedkingdom": "GBD - United Kingdom British Territories Citizen",
    "unitednationsorganization": "UNO - United Nations Organization",
    "unitedstates": "USA - United States of America",
    "urugoay": "URY - Uruguay",
    "uruguay": "URY - Uruguay",
    "uzbekistan": "UZB - Uzbekistan",
    "vanuatu": "VUT - Vanuatu",
    "vaticancity": "VAT - Holy See (Vatican City State )",
    "vaticang": "VAT - Holy See (Vatican City State )",
    "veneduela": "VEN - Venezuela",
    "venezuela": "VEN - Venezuela",
    "vietnam": "VNM - Viet Nam",
    "vungatthuocanhoanoduong": "IOT - British India Ocean Territory",
    "vungthuophap": "FRA - France",
    "vuongquocnauy": "NOR - Norway",
    "westernsamoa": "WSM - Western Samoa",
    "xamoa": "WSM - Western Samoa",
    "xanhluxia": "LCA - Saint Lucia",
    "xanmarino": "SMR - San Marino",
    "xcolent": "SC- - Scotland",
    "xecbia": "SRB - Serbia",
    "xenegan": "SEN - Senegal",
    "xingapo": "SGP - Singapore",
    "xlovakia": "SVK - Slovakia",
    "xoadilen": "SWZ - Swaziland",
    "xomali": "SOM - Somalia",
    "xrilanca": "LKA - Sri Lanka",
    "xuang": "SDN - Sudan",
    "xurinam": "SUR - Suriname",
    "y": "ITA - Italy",
    "yemen": "YEM - Yemen",
    "zambia": "ZMB - Zambia",
    "zimbabwe": "ZWE - Zimbabwe"
}

def lookup_nat_kbtt(raw):
    """Khớp thông minh: chuẩn hóa dấu/khoảng trắng để tìm mã quốc tịch."""
    if not raw: return ''
    raw = str(raw).strip()
    key = _norm_nat(raw)
    if key in NAT_NORM:
        return NAT_NORM[key]
    # already in CODE - Name form?
    if _re.match(r'^[A-Z]{2,3} - ', raw):
        return raw
    return raw  # unknown -> keep original (sẽ hiện cảnh báo)

NAT_DK14 = {
    'AFG':'Afganistan  ( Ap-ga-ni-xtan )','ZAF':'Africa (South)  ( Nam Phi )',
    'ALB':'Albania  ( An-ba-ni )','DZA':'Algieria  ( An-giê-ri )',
    'ASM':'American Samoa  ( Đông Sa-moa )','AND':'Andorra  ( Công quốc An-đơ-ra )',
    'ATA':'Antarctica  ( Nam Cực )','AGO':'Angola  ( Ăng-gô-la )',
    'AIA':'Anguilla  ( Ăng-gui-la )','ARG':'Argentina  ( Ac-hen-ti-na )',
    'ARM':'Armenia  ( Ac-mê-ni-a )','ABW':'Aruba  ( A-ru-ba )',
    'AUS':'Australia  ( Ô-xtrây-li-a )','AUT':'Austria  ( áo )',
    'AZE':'Azerbaijan  ( A-déc-bai-gian )','BHS':'Bahamas  ( Ba-ha-ma )',
    'BHR':'Bahrain  ( Ba-ra-in )','BGD':'Bangladesh  ( Băng-la-đét )',
    'BRB':'Barbados  ( Bác-ba-đốt )','BLR':'Belarus  ( Bê-la-rút )',
    'BEL':'Belgium  ( Bỉ )','BLZ':'Belize  ( Bê-li-xê )',
    'BEN':'Benin  ( Bê-nanh )','BMU':'Bermuda  ( Béc-mu-đa )',
    'BTN':'Bhutan  ( Bu-tan )','BOL':'Bolivia  ( Bô-li-vi-a )',
    'BIH':'Bosnia and Herzegovina  ( Bô-xni-a Héc-dê-gô-vi-na )','BWA':'Botswana  ( Bốt-xoa-na )',
    'BRA':'Brazil  ( Bra-din )','BRN':'Brunei Darussalam  ( Đa-ru-xa-lem thuộc Brunei )',
    'BGR':'Bulgaria  ( Bun-ga-ri )','BFA':'Burkina Faso  ( Buốc-ki-na Pha-xô )',
    'BDI':'Burundi  ( Bu-run-đi )','CMR':'Cameroon  ( Ca-mơ-run )',
    'CAN':'Canada  ( Ca-na-da )','COL':'Colombia  ( Cô-lôm-bi-a )',
    'COM':'Comoros  ( Cô-mo )','COG':'Congo  ( Công-gô )',
    'COK':'Cook Islands  ( Quần đảo Cúc )','CRI':'Costa Rica  ( Cô-xta Ri-ca )',
    'HRV':'Croatia  ( Crô-a-ti-a )','CUB':'Cuba  ( Cu Ba )',
    'CYP':'Cyprus  ( Đảo Síp )','CZE':'Czech Republic  ( Cộng hoà Séc )',
    'TCD':'Chad  ( Sát )','CHL':'Chile  ( Chi-lê )',
    'CHN':'China  ( Trung Quốc )','TWN':'China (Taiwan)  ( Trung Quốc (Đài Loan) )',
    'DNK':'Denmark  ( Đan Mạch )','DJI':'Djibouti  ( Đi-bô-u-ti )',
    'DMA':'Dominica  ( Đô-mi-ni-ca )','ECU':'Ecuador  ( Ê-cu-a-đo )',
    'EGY':'Egypt  ( Ai Cập )','SLV':'El Salvador  ( En Xan-va-đo )',
    'GNQ':'Equatorial Guinea  ( Ghi-nê Xích đạo )','ERI':'Eritrea  ( Ê-ri-tơ-ri-a )',
    'EST':'Estonia  ( Ê-xtô-ni-a )','ETH':'Ethiopia  ( Ê-ti-ô-pi-a )',
    'FJI':'Fiji  ( Fi-ji )','FIN':'Finland  ( Phần Lan )',
    'FRA':'France  ( Pháp )','GAB':'Gabon  ( Ga-bông )',
    'GMB':'Gambia  ( Găm-bi-a )','GEO':'Georgia  ( Gru-di-a )',
    'DEU':'Germany  ( CH Liên bang Đức )','GHA':'Ghana  ( Ga-na )',
    'GRC':'Greece  ( Hy Lạp )','GRL':'Greenland  ( Grin-lơn )',
    'GTM':'Guatemala  ( Goa-tê-ma-la )','GIN':'Guinea  ( Ghi-nê )',
    'GNB':'Guinea-Bissau  ( Ghi-nê Bít-xao )','GUY':'Guyana  ( Gui-na )',
    'HTI':'Haiti  ( Ha-i-ti )','HND':'Honduras  ( Hon-du-rat )',
    'HKG':'HongKong  ( Hồng-Kông )','HUN':'Hungari  ( Hung-ga-ri )',
    'ISL':'Iceland  ( Ai-xơ-len )','IND':'India  ( Ân Độ )',
    'IDN':'Indonesia  ( In-đô-nê-xi-a )','IRN':'Iran  ( CH Hồi giáo I-ran )',
    'IRQ':'Iraq  ( I-rắc )','IRL':'Ireland  ( Ai-rơ-len )',
    'ISR':'Israel  ( I-xra-en )','ITA':'Italy  ( I-ta-li-a )',
    'JAM':'Jamaica  ( Ja-mai-ca )','JPN':'Japan  ( Nhật Bản )',
    'JOR':'Jordan  ( Joc-đan )','CAP':'Kampuchea  ( Căm-pu-chia )',
    'KAZ':'Kazakhstan  ( Ka-dắc-xtan )','KEN':'Kenya  ( Kê-ni-a )',
    'KOR':'Korea (South)  ( CH Hàn Quốc )','PRK':'Korea Democratic Peoples Republic  ( CHDCND Triều Tiên )',
    'KWT':'Kuwait  ( Cô-oét )','KGZ':'Kyrgyzstan  ( Kiếc-ghi-di-a )',
    'LAO':'Laos  ( CHDCND Lào )','LVA':'Latvia  ( Lát-vi-a )',
    'LBN':'Lebanon  ( Li-ban )','LSO':'Lesotho  ( Lê-xô-thô )',
    'LBR':'Liberia  ( Li-bê-ri-a )','LBY':'Libya  ( Gia-ma-hi-ri-i-a A-rập Li-bi Nhân dân )',
    'LIE':'Liechtenstein  ( Công quốc Lích-ten-xtên )','LTU':'Lithuania  ( Lit-hua-ni-a )',
    'LUX':'Luxembourg  ( Luých-xem-bua )','MAC':'Macau  ( Ma Cao )',
    'MDG':'Madagascar  ( Ma-đa-ga-xca )','MWI':'Malawi  ( Ma-la-uy )',
    'MYS':'Malaysia  ( Ma-lai-xi-a )','MDV':'Maldives  ( Man-đi-vơ )',
    'MLI':'Mali  ( Ma-li )','MLT':'Malta  ( Man-ta )',
    'MHL':'Marshall Islands  ( Quần đảo Mác-san )','MRT':'Mauritania  ( Mô-ra-ta-ni )',
    'MUS':'Mauritius  ( Mô-ri-xơ )','MEX':'Mexico  ( Mê-xi-cô )',
    'MDA':'Moldova  ( Môn-đô-va )','MCO':'Monaco  ( Công quốc Mô-na-cô )',
    'MNE':'Montenegro  ( Môn-tê-nê-grô )','MNG':'Mongolia  ( Mông Cổ )',
    'MAR':'Morocco  ( Ma-rốc )','MOZ':'Mozambique  ( Mô-dăm-bích )',
    'MMR':'Myanmar (Burma)  ( Mi-an-ma )','NAM':'Namibia  ( Na-mi-bi-a )',
    'NPL':'Nepal  ( Nê-pan )','NLD':'Netherland  ( Hà Lan )',
    'NZL':'New Zealand  ( Niu Di-lân )','NIC':'Nicaragua  ( Ni-ca-ra-goa )',
    'NER':'Niger  ( Ni-giê )','NGA':'Nigeria  ( Ni-giê-ri-a )',
    'NOR':'Norway  ( Vương quốc Na-uy )','OMN':'Oman  ( Ô-man )',
    'PAK':'Pakistan  ( Pa-ki-xtan )','PLW':'Palau  ( Pa-lau )',
    'PSE':'Palestine  ( Pa-le-xtin )','PAN':'Panama  ( Pa-na-ma )',
    'PNG':'Papua New Guinea  ( Pa-pua Niu Ghi-nê )','PRY':'Paraguay  ( Pa-ra-goay )',
    'PER':'Peru  ( Pê-ru )','POL':'Poland  ( Ba Lan )',
    'PRT':'Portugal  ( Bồ Đào Nha )','PRI':'Puerto Rico  ( Pu-éc-tô Ri-cô )',
    'PHL':'Philippine  ( Phi-líp-pin )','QAT':'Qatar  ( Qua-ta )',
    'ROU':'Romania  ( Ru-ma-ni )','RUS':'Russia  (Liên bang Nga)',
    'RWA':'Rwanda  ( Ru-an-đa )','LCA':'Saint Lucia  ( Xanh Lu-xi-a )',
    'SMR':'San Marino  ( Xan Ma-ri-nô )','SAU':'Saudi Arabia  ( A-rập Xau-đi )',
    'GBR':'United Kingdom  ( Liên hiệp Vương quốc Anh và Bắc Ailen )',
    # Một số hệ thống PMS ghi Vương quốc Anh là GBS thay vì GBR — nhận cả hai
    'GBS':'United Kingdom  ( Liên hiệp Vương quốc Anh và Bắc Ailen )',
    'SEN':'Senegal  ( Xe-ne-gan )',
    'SRB':'Serbia  ( Xéc-bi-a )','SYC':'Seychelles  ( Quần đảo Xây-sen )',
    'SGP':'Singapore  ( Xin-ga-po )','SVK':'Slovakia  ( Xlô-va-ki-a )',
    'SVN':'Slovenia  ( Slo-vê-ni-a )','SOM':'Somalia  ( Xô-ma-li )',
    'ESP':'Spain  ( Tây Ban Nha )','LKA':'Srilanka  ( Xri-Lan-ca )',
    'SDN':'Sudan  ( Xu-đăng )','SWE':'Sweden  ( Thuỵ Điển )',
    'CHE':'Switzerland  ( Thuỵ Sĩ )','TJK':'Tajikistan  ( Ta-gi-ki-xtan )',
    'TGO':'Togo  ( Tô-gô )','TON':'Tonga  ( Tôn-ga )',
    'TUN':'Tunisia  ( Tu-ni-di )','TUR':'Turkey  ( Thổ Nhĩ Kỳ )',
    'TKM':'Turkmenistan  ( Tuốc-mê-ni-xtan )','TUV':'Tuvalu  ( Tu-va-lu )',
    'THA':'Thailand  ( Thái Lan )','UGA':'Uganda  ( U-gan-da )',
    'UKR':'Ukraine  ( U-crai-na )','ARE':'United Arab Emirates  ( A-rập thống nhất )',
    'TZA':'United Republic of Tanzania  ( CH thống nhất Tan-da-ni-a )',
    'USA':'United States  ( Mỹ )','URY':'Uruguay  ( U-ru-goay )',
    'UZB':'Uzbekistan  ( U-dơ-bê-ki-xtan )','VUT':'Vanuatu  ( Va-nu-a-tu )',
    'VEN':'Venezuela  ( Vê-nê-du-ê-la )','VNM':'Vietnam  ( Việt Nam )',
    'YEM':'Yemen  ( Y-ê-men )','ZMB':'Zambia  ( Dăm-bi-a )',
    'ZWE':'Zimbabwe  ( Dim-ba-bu-ê )',
}
LOAI_GIAY = {
    'Căn cước công dân':'8 - Thẻ Căn Cước','Hộ chiếu':'4 - Hộ chiếu',
    'Chứng minh nhân dân':'2 - Thẻ CMND','Căn cước':'1 - Thẻ CCCD',
    'Giấy khai sinh':'5 - Giấy khai sinh',
}
TINH = {
    'DAK LAK':'605 - Đắk Lắk','DAK NONG':'607 - Đắk Nông','PHU YEN':'509 - Phú Yên',
    'HO CHI MINH':'701 - TP. Hồ Chí Minh','THP HO CHI MINH':'701 - TP. Hồ Chí Minh',
    'HCM':'701 - TP. Hồ Chí Minh','TP HCM':'701 - TP. Hồ Chí Minh',
    'GIA LAI':'603 - Gia Lai','QUANG NGAI':'505 - Quảng Ngãi','LAM DONG':'703 - Lâm Đồng',
    'LAM DONG.':'703 - Lâm Đồng','KHANH HOA':'511 - Khánh Hòa','BEN TRE':'817 - Bến Tre',
    'VINH LONG':'809 - Vĩnh Long','NINH THUAN':'513 - Ninh Thuận','DONG NAI':'713 - Đồng Nai',
    'TIEN GIANG':'807 - Tiền Giang','LONG AN':'801 - Long An','BINH DUONG':'707 - Bình Dương',
    'BINH THUAN':'705 - Bình Thuận','CAN THO':'815 - TP. Cần Thơ','DA NANG':'501 - TP. Đà Nẵng',
    'HA NOI':'101 - TP. Hà Nội','BA RIA VUNG TAU':'711 - Bà Rịa - Vũng Tàu',
    'DONG THAP':'803 - Đồng Tháp','AN GIANG':'805 - An Giang','KIEN GIANG':'819 - Kiên Giang',
    'HA TINH':'407 - Hà Tĩnh','BINH DINH':'507 - Bình Định','VIET NAM':'',
}

# ── Helpers ───────────────────────────────────────────────────────────────
def fmt(v):
    if v is None or str(v) in ('NaT','nan',''): return ''
    if hasattr(v,'strftime'): return v.strftime('%d/%m/%Y')
    return str(v).strip()[:10]

def make_code(prefix, ns):
    if not ns: return prefix
    p = ns.replace('-','/').split('/')
    return f"{prefix}{p[0].zfill(2)}{p[1].zfill(2)}{p[2][-2:]}" if len(p)==3 else prefix

def cp(src, dst):
    for a in ('font','fill','border','alignment'):
        v = getattr(src,a)
        if v: setattr(dst,a,copy(v))
    dst.number_format = src.number_format

def serial2date(s):
    if not s: return None
    try: return datetime.datetime(1899,12,30)+datetime.timedelta(days=int(s))
    except: return None

def wb_to_bytes(wb):
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── Tiện ích nghiệp vụ lễ tân (tỷ giá, kiểm tra dữ liệu, visa, báo cáo, giao ca) ──
@st.cache_data(ttl=600, show_spinner=False)
def fetch_vcb_rates():
    """Lấy tỷ giá CHUYỂN KHOẢN USD/EUR → VND từ Vietcombank (cache 10 phút).
    Chỉ là tiện ích — lễ tân vẫn nhập tay được nếu mạng/VCB lỗi."""
    import urllib.request
    import xml.etree.ElementTree as ET
    url = 'https://portal.vietcombank.com.vn/Usercontrols/TVPortal.TyGia/pXML.aspx'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=10) as resp:
        root = ET.fromstring(resp.read())
    rates = {}
    for ex in root.iter('Exrate'):
        code = (ex.get('CurrencyCode') or '').upper()
        if code in ('USD', 'EUR'):
            try:
                rates[code] = float((ex.get('Transfer') or '').replace(',', ''))
            except ValueError:
                pass
    if 'USD' not in rates:
        raise ValueError('không đọc được tỷ giá USD trong dữ liệu VCB')
    return rates, now_vn().strftime('%H:%M %d/%m/%Y')

def _gv(row, *names):
    """Lấy giá trị đầu tiên khác rỗng theo danh sách tên cột (chịu biến thể tên cột)."""
    for n in names:
        if n in row.index:
            v = row[n]
            if pd.notna(v) and str(v).strip() != '':
                return v
    return None

def _fmt_room(v):
    """Số phòng/số điện thoại đọc từ Excel đôi khi ra dạng số thực (103.0)
    → trả về '103'. Dùng `v or ''` sẽ SAI với NaN (NaN truthy trong Python)
    → lọt chuỗi "nan" ra file; phải chặn None/NaN riêng trước."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ''
    s = str(v).strip()
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s

def _to_date(v):
    """Đọc ngày linh hoạt (datetime / chuỗi dd/mm/yyyy) → datetime.date hoặc None."""
    if v is None:
        return None
    if hasattr(v, 'year') and not isinstance(v, str):
        try:
            return datetime.date(v.year, v.month, v.day)
        except Exception:
            return None
    t = pd.to_datetime(str(v).strip(), dayfirst=True, errors='coerce')
    return None if pd.isna(t) else t.date()

def validate_guests(df):
    """Kiểm tra chất lượng dữ liệu khách TRƯỚC khi nộp hồ sơ KBTT/VNM/ĐK14.
    🔴 = lỗi dễ khiến công an trả hồ sơ · 🟡 = nên kiểm tra lại trước khi nộp."""
    issues = []
    for idx, row in df.iterrows():
        line = idx + 2  # số dòng trên Excel gốc (dòng 1 là header)
        ht = str(_gv(row, 'HỌ TÊN ', 'HỌ TÊN') or '').strip()
        sp = _fmt_room(_gv(row, 'SỐ PHÒNG'))
        intl = str(_gv(row, 'LOẠI KHÁCH') or '').strip() == 'Quốc tế'
        def add(sev, msg):
            issues.append({'Mức độ': sev, 'Dòng': line, 'Họ tên': ht or '(trống)',
                           'Phòng': sp, 'Vấn đề': msg})
        if not ht:
            add('🔴', 'Thiếu họ tên')
        sg = str(_gv(row, 'SỐ GIẤY TỜ') or '').strip()
        if not sg:
            add('🔴' if intl else '🟡',
                'Thiếu số giấy tờ' + (' (hộ chiếu bắt buộc cho KBTT)' if intl else ''))
        elif intl and not _re.fullmatch(r'[A-Za-z0-9]{4,15}', sg.replace(' ', '')):
            add('🟡', f'Số hộ chiếu có ký tự lạ: "{sg}"')
        if _gv(row, 'NGÀY SINH') is None:
            add('🟡', 'Thiếu ngày sinh')
        if str(_gv(row, 'GIỚI TÍNH') or '').strip() not in ('Nam', 'Nữ'):
            add('🟡', 'Giới tính trống/không chuẩn (cần "Nam" hoặc "Nữ")')
        if not sp:
            add('🟡', 'Thiếu số phòng')
        nd = _to_date(_gv(row, 'NGÀY ĐẾN'))
        ni = _to_date(_gv(row, 'NGÀY ÐI', 'NGÀY ĐI'))
        if nd and ni and ni < nd:
            add('🔴', f'Ngày đi {ni.strftime("%d/%m/%Y")} TRƯỚC ngày đến {nd.strftime("%d/%m/%Y")}')
        if intl:
            qt = str(_gv(row, 'QUỐC TỊCH') or '').strip()
            if not qt:
                add('🔴', 'Thiếu quốc tịch')
            elif not _re.match(r'^[A-Z]{2,3} - ', str(lookup_nat_kbtt(qt))):
                add('🟡', f'Quốc tịch chưa có mã: "{qt}"')
    return pd.DataFrame(issues, columns=['Mức độ', 'Dòng', 'Họ tên', 'Phòng', 'Vấn đề'])

def check_visa_expiry(df_intl, visa_map=None):
    """Gom hạn tạm trú/visa từng khách quốc tế — CHỈ tin dữ liệu từ file Visa
    rời do lễ tân chủ động upload (khớp theo hộ chiếu trước, tên là dự phòng).
    KHÔNG tự đọc cột 'TẠM TRÚ'/'TẠM TRÚ ĐẾN NGÀY' có sẵn trong file dữ liệu
    khách (customer.xls) nữa — cột đó thường do PMS tự điền mặc định, không
    phải visa đã xác nhận thật, tin nhầm có thể điền sai lên hồ sơ khai báo
    công an. → list dict ngày dạng ISO, lưu được vào session để lọc lại theo
    số ngày cảnh báo mà không cần xử lý lại."""
    visa_map = visa_map or {}
    by_pp = visa_map.get('by_pp', {}) if isinstance(visa_map, dict) else {}
    by_name = visa_map.get('by_name', {}) if isinstance(visa_map, dict) else {}
    out = []
    for _, row in df_intl.iterrows():
        ht = str(_gv(row, 'HỌ TÊN ', 'HỌ TÊN') or '').strip()
        sg = str(_gv(row, 'SỐ GIẤY TỜ') or '').strip()
        vd_raw = None
        if by_pp or by_name:
            vd_raw = by_pp.get(_norm_pp(sg)) or by_name.get(_norm_name(ht))
        vd = _to_date(vd_raw)
        if vd is None:
            continue
        ni = _to_date(_gv(row, 'NGÀY ÐI', 'NGÀY ĐI'))
        out.append({'name': ht, 'room': _fmt_room(_gv(row, 'SỐ PHÒNG')),
                    'nat': str(_gv(row, 'QUỐC TỊCH') or '').strip(),
                    'visa': vd.isoformat(), 'dep': ni.isoformat() if ni else None})
    return out

def build_handover_xlsx(info, entries):
    """Xuất sổ giao ca thành file Excel in được, có cột 'Đã xử lý' để ca sau tick tay."""
    wb = Workbook(); ws = wb.active; ws.title = 'Giao ca'
    thin = Side(style='thin'); bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, w in zip('ABCDEF', [5, 8, 26, 9, 58, 10]):
        ws.column_dimensions[col].width = w
    ws.merge_cells('A1:F1')
    c = ws.cell(1, 1)
    c.value = (f"SỔ GIAO CA — {info.get('date', '')} — {info.get('shift', '')}"
               f" — Lễ tân: {info.get('staff', '') or '…'}")
    c.font = Font(name='Times New Roman', size=14, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(['STT', 'Giờ ghi', 'Phân loại', 'Phòng', 'Nội dung bàn giao', 'Đã xử lý'], 1):
        cell = ws.cell(2, ci); cell.value = h
        cell.font = Font(name='Times New Roman', size=11, bold=True)
        cell.fill = PatternFill('solid', fgColor='DDEBF7'); cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, e in enumerate(entries, 1):
        ws.row_dimensions[i + 2].height = 32
        vals = [i, e.get('time', ''), e.get('cat', ''), e.get('room', ''), e.get('note', ''), '☐']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, ci); cell.value = v
            cell.font = Font(name='Times New Roman', size=11); cell.border = bdr
            cell.alignment = Alignment(horizontal='left' if ci == 5 else 'center',
                                       vertical='center', wrap_text=True)
    return wb

def build_daily_report(date_str, daily, arr_stats, recon, reconr):
    """Báo cáo ngày 1 trang (Excel) cho quản lý — tổng hợp mọi số liệu các công cụ
    đã chạy trong phiên; phần nào chưa chạy thì tự bỏ qua."""
    wb = Workbook(); ws = wb.active; ws.title = 'Bao cao ngay'
    thin = Side(style='thin'); bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, w in zip('AB', [40, 22]):
        ws.column_dimensions[col].width = w
    bold = Font(name='Times New Roman', size=11, bold=True)
    norm = Font(name='Times New Roman', size=11)
    fill_h = PatternFill('solid', fgColor='DDEBF7')
    ws.merge_cells('A1:B1')
    t = ws.cell(1, 1); t.value = f"BÁO CÁO NGÀY {date_str} — TÂN HOTEL (FRONT OFFICE)"
    t.font = Font(name='Times New Roman', size=14, bold=True)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    r = 3
    def section(title):
        nonlocal r
        ws.merge_cells(f'A{r}:B{r}')
        c = ws.cell(r, 1); c.value = title; c.font = bold; c.fill = fill_h; c.border = bdr
        ws.cell(r, 2).border = bdr
        r += 1
    def kv(label, value):
        nonlocal r
        a = ws.cell(r, 1); a.value = label; a.font = norm; a.border = bdr
        b = ws.cell(r, 2); b.value = value; b.font = bold; b.border = bdr
        b.alignment = Alignment(horizontal='center')
        r += 1
    if daily:
        section('1. KHÁCH LƯU TRÚ (file dữ liệu khách)')
        kv('Tổng khách', daily.get('total'))
        kv('Khách quốc tế', daily.get('intl'))
        kv('Khách Việt Nam', daily.get('vn'))
        if daily.get('rooms_cnt'):
            kv('Số phòng có khách', daily.get('rooms_cnt'))
        if daily.get('avg_nights'):
            kv('Số đêm lưu trú bình quân', daily.get('avg_nights'))
        kv('Trẻ em (GKS) + Giấy bảo lãnh (GBL)', f"{daily.get('gks', 0)} + {daily.get('gbl', 0)}")
        r += 1
        if daily.get('nat_top'):
            section('2. TOP QUỐC TỊCH')
            for nat, cnt in daily['nat_top']:
                kv(nat, cnt)
            r += 1
    if arr_stats:
        section('3. BOOKING ĐẾN & THANH TOÁN (file ARR)')
        kv('Số booking arrival', arr_stats.get('bookings'))
        kv('Số phòng arrival', arr_stats.get('rooms'))
        if arr_stats.get('ota') is not None:
            kv('Booking qua OTA', arr_stats.get('ota'))
        kv('Cần cà thẻ (CÀ THẺ)', arr_stats.get('ca_the'))
        kv('Cần thu tiền (THU TIỀN)', arr_stats.get('thu_tien'))
        kv('Xem lại BU', arr_stats.get('xem_lai_bu'))
        kv('FOC Late C/O', arr_stats.get('foc_lco'))
        r += 1
    if recon:
        section('4. ĐỐI CHIẾU LƯU TRÚ NGƯỜI NƯỚC NGOÀI')
        kv('Khách chưa đăng ký lưu trú', len(recon.get('chua_dk', [])))
        kv('Có trên lưu trú, thiếu trên Smile', len(recon.get('thua', [])))
        kv('Đăng ký trùng', len(recon.get('dup', [])))
        r += 1
    if reconr:
        section('5. ĐỐI CHIẾU HỆ THỐNG PHÒNG')
        kv('Phòng chưa đăng ký', len(reconr.get('room_chua', [])))
        kv('Phòng thừa trong file', len(reconr.get('room_thua', [])))
        kv('Phòng trùng trong file', len(reconr.get('sys_dup', [])))
        r += 1
    r += 1
    f = ws.cell(r, 1)
    f.value = f"Xuất lúc {now_vn().strftime('%H:%M %d/%m/%Y')} — Hotel Tool Pro"
    f.font = Font(name='Times New Roman', size=9, italic=True, color='FF888888')
    return wb

# ── Cặp phòng connecting — tự động chia đều ĐƠN GIÁ khi 1 phòng bị bỏ trống ──
# Danh sách cố định theo sơ đồ tầng khách sạn (nguồn: C_p_CNT.xlsx, tầng 5→12A,
# mỗi tầng 6 cặp). Đây là kết cấu vật lý của toà nhà, gần như không đổi — để
# hằng số thay vì nạp file mỗi lần, dễ đối chiếu/sửa khi khách sạn cải tạo.
# CHỈ áp dụng cho các cặp có TRONG danh sách này — cặp phòng nào phát sinh
# ngoài danh sách (VD: tầng chưa được liệt kê) thì GIỮ NGUYÊN giá, không đoán.
_CNT_FLOOR_ROWS = [
    (538, 638, 738, 838, 934, 1034, 1134, 1230, '12A30'),
    (540, 640, 740, 840, 936, 1036, 1136, 1232, '12A32'),
    (542, 642, 742, 842, 938, 1038, 1138, 1234, '12A34'),
    (544, 644, 744, 844, 940, 1040, 1140, 1236, '12A36'),
    (546, 646, 746, 846, 942, 1042, 1142, 1238, '12A38'),
    (548, 648, 748, 848, 944, 1044, 1144, 1240, '12A40'),
    (541, 641, 741, 841, 937, 1037, 1137, 1233, '12A33'),
    (543, 643, 743, 843, 939, 1039, 1139, 1235, '12A35'),
    (545, 645, 745, 845, 941, 1041, 1141, 1237, '12A37'),
    (547, 647, 747, 847, 943, 1043, 1143, 1239, '12A39'),
    (549, 649, 749, 849, 945, 1045, 1145, 1241, '12A41'),
    (550, 650, 750, 850, 946, 1046, 1146, 1242, '12A42'),
]
CONNECTING_ROOM_PAIRS = []
for _fi in range(9):  # 9 cột tầng: 5,6,7,8,9,10,11,12,12A
    for _pi in range(0, 12, 2):  # mỗi 2 dòng liên tiếp là 1 cặp, 6 cặp/tầng
        CONNECTING_ROOM_PAIRS.append(
            (_fmt_room(_CNT_FLOOR_ROWS[_pi][_fi]), _fmt_room(_CNT_FLOOR_ROWS[_pi + 1][_fi])))

def split_connecting_room_prices(xlsx_bytes):
    """Tự động chia đều ĐƠN GIÁ cho cặp phòng connecting khi CHUNG MÃ CHECKIN
    (cùng 1 lượt đặt) mà chỉ 1 phòng được ghi giá, phòng kia bằng 0 — PMS
    thường gộp giá cả 2 phòng connecting vào 1 phòng khi đặt chung.

    Quy tắc CHỈ áp dụng khi cả 2 điều kiện đúng:
    1. Cặp (roomA, roomB) có TRONG CONNECTING_ROOM_PAIRS (không đoán ngoài
       danh sách — cặp phát sinh ngoài ý muốn thì GIỮ NGUYÊN giá).
    2. Cả 2 phòng CHUNG 1 MÃ CHECKIN, và đúng 1 phòng có tổng giá > 0, phòng
       còn lại tổng giá = 0.

    Tổng lẻ (không chia hết 2) thì phòng SỐ NHỎ HƠN nhận phần làm tròn XUỐNG,
    phòng SỐ LỚN HƠN nhận phần làm tròn LÊN — cộng lại đúng bằng tổng ban đầu
    (đối chiếu đúng theo file QLLT mẫu đã tách tay: VD 8.313.043 → 4.156.521 +
    4.156.522). Trả về (bytes file đã sửa, danh sách bản ghi đã chia để hiển
    thị minh bạch cho người dùng kiểm tra lại)."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    headers_norm = [_norm_nat(h) if h else None for h in headers]

    def _col(name):
        target = _norm_nat(name)
        return next((i + 1 for i, hn in enumerate(headers_norm) if hn == target), None)

    ci_checkin, ci_room, ci_gia = _col('MÃ CHECKIN'), _col('SỐ PHÒNG'), _col('ĐƠN GIÁ')
    if not (ci_checkin and ci_room and ci_gia):
        return xlsx_bytes, []  # thiếu cột cần thiết — không đụng vào file

    # Gom danh sách dòng theo (mã checkin, số phòng) để tính tổng giá mỗi phòng
    # và biết CHÍNH XÁC dòng nào đang giữ giá trị > 0 cần sửa.
    groups = {}  # (checkin, room) -> list các dòng (row index)
    for r in range(2, ws.max_row + 1):
        checkin = ws.cell(r, ci_checkin).value
        room = _fmt_room(ws.cell(r, ci_room).value)
        if checkin is None or str(checkin).strip() == '' or not room:
            continue
        checkin = str(checkin).strip()
        groups.setdefault((checkin, room), []).append(r)

    rooms_by_checkin = {}
    for (checkin, room) in groups:
        rooms_by_checkin.setdefault(checkin, set()).add(room)

    report = []
    for checkin, rooms in rooms_by_checkin.items():
        for room_a, room_b in CONNECTING_ROOM_PAIRS:
            if room_a not in rooms or room_b not in rooms:
                continue
            rows_a = groups[(checkin, room_a)]
            rows_b = groups[(checkin, room_b)]

            def _total(rows):
                s = 0
                for r in rows:
                    v = ws.cell(r, ci_gia).value
                    if isinstance(v, (int, float)) and not pd.isna(v):
                        s += v
                return s

            total_a, total_b = _total(rows_a), _total(rows_b)
            if not ((total_a > 0) != (total_b > 0)):
                continue  # cả 2 đều có giá, hoặc cả 2 đều 0 — không đụng vào
            nonzero_room, nonzero_rows, nonzero_total = (
                (room_a, rows_a, total_a) if total_a > 0 else (room_b, rows_b, total_b))
            zero_room, zero_rows = (room_b, rows_b) if nonzero_room == room_a else (room_a, rows_a)

            lower, higher = sorted([room_a, room_b])
            half_floor = int(nonzero_total) // 2
            half_ceil = int(nonzero_total) - half_floor
            new_val = {lower: half_floor, higher: half_ceil}

            # Dòng đang giữ giá trị > 0 → sửa thành phần chia của chính phòng đó.
            # Dòng ĐẦU TIÊN của phòng đang 0 → điền phần chia còn lại. Các dòng
            # khác (khách đi cùng phòng) giữ nguyên 0 — không đụng.
            for r in nonzero_rows:
                v = ws.cell(r, ci_gia).value
                if isinstance(v, (int, float)) and v == nonzero_total:
                    ws.cell(r, ci_gia).value = new_val[nonzero_room]
                    break
            ws.cell(zero_rows[0], ci_gia).value = new_val[zero_room]

            report.append({'checkin': checkin, 'room_a': room_a, 'room_b': room_b,
                           'total': int(nonzero_total),
                           'split_a': new_val[room_a], 'split_b': new_val[room_b]})

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report

# ── Processing ────────────────────────────────────────────────────────────
def process_xlsx(xlsx_bytes, rate):
    """Điền dữ liệu file đầu vào lên FILE MẪU customer (QLLT) — giữ nguyên 100%
    template: sheet 'customer' + sheet 'Danh-muc', định dạng header, style ô dữ liệu.
    Map cột theo tên (không phân biệt hoa thường / dấu / khoảng trắng), quy đổi
    ĐƠN GIÁ (USD → VND, tô vàng ô đã đổi), đánh lại STT."""
    src_wb = load_workbook(io.BytesIO(xlsx_bytes))
    src_ws = src_wb.active

    # Map header nguồn (chuẩn hóa) → chỉ số cột nguồn
    src_map = {}
    for c in src_ws[1]:
        if c.value is not None and str(c.value).strip():
            src_map.setdefault(_norm_nat(c.value), c.column)

    # Nạp template QLLT (customer + Danh-muc), dòng 2 là mẫu định dạng ô dữ liệu
    wb = load_workbook(io.BytesIO(load_template('customer')))
    ws = wb['customer']
    n_cols = ws.max_column
    ref = [ws.cell(2, ci) for ci in range(1, n_cols + 1)]
    ref_styles = [(copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment), c.number_format) for c in ref]

    # Với mỗi cột template, tìm cột nguồn tương ứng theo tên
    headers = [ws.cell(1, ci).value for ci in range(1, n_cols + 1)]
    headers_norm = [_norm_nat(h) if h else None for h in headers]
    col_src = [src_map.get(hn) if hn else None for hn in headers_norm]
    # Khớp tên tuyệt đối thất bại với 1 số cột nếu file nguồn đặt tên rút gọn
    # hơn mẫu QLLT (vd nguồn "CỬA KHẨU" ↔ mẫu "CỬA KHẨU NHẬP CẢNH", nguồn
    # "TẠM TRÚ" ↔ mẫu "TẠM TRÚ ĐẾN NGÀY") → mất trắng dữ liệu cột đó dù có
    # trong file nguồn. Với cột còn thiếu, thử khớp NỚI LỎNG: 1 trong 2 tên là
    # phần đầu của tên còn lại (đủ dài để tránh khớp nhầm với tên ngắn/mơ hồ),
    # mỗi cột nguồn chỉ được dùng khớp 1 lần để tránh nhầm giữa 2 cột đích.
    _used_src_cols = {c for c in col_src if c}
    for i, hn in enumerate(headers_norm):
        if col_src[i] or not hn:
            continue
        for src_key, src_col in src_map.items():
            if src_col in _used_src_cols:
                continue
            if len(src_key) >= 4 and (hn.startswith(src_key) or src_key.startswith(hn)):
                col_src[i] = src_col
                _used_src_cols.add(src_col)
                break
    don_gia_idx = next((i + 1 for i, h in enumerate(headers) if h and _norm_nat(h) == _norm_nat('ĐƠN GIÁ')), None)

    ws.delete_rows(2)  # bỏ dòng mẫu

    conv = 0
    er = 1
    for row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row):
        if all(c.value is None or str(c.value).strip() == '' for c in row):
            continue
        er += 1
        for ci in range(1, n_cols + 1):
            cell = ws.cell(er, ci)
            f, fl, b, a, nf = ref_styles[ci - 1]
            cell.font = copy(f); cell.fill = copy(fl); cell.border = copy(b)
            cell.alignment = copy(a); cell.number_format = nf
            sc = col_src[ci - 1]
            if sc is not None and row[sc - 1].value is not None:
                cell.value = row[sc - 1].value
        ws.cell(er, 1).value = er - 1  # STT đánh lại
        if don_gia_idx:
            dg = ws.cell(er, don_gia_idx)
            if dg.value and isinstance(dg.value, (int, float)) and 0 < dg.value < 1000:
                dg.value = round(dg.value * rate)
                conv += 1
    return wb, conv

def split_wb(wb, loai):
    wb2 = load_workbook(io.BytesIO(wb_to_bytes(wb)))
    ws2 = wb2.active
    lc = next(c.column for c in ws2[1] if c.value=='LOẠI KHÁCH')
    dels = [row[0].row for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row)
            if row[lc-1].value != loai]
    for r in reversed(dels): ws2.delete_rows(r)
    for i, row in enumerate(ws2.iter_rows(min_row=2,max_row=ws2.max_row),1): row[0].value=i
    return wb2

def _norm_name(s):
    """Chuẩn hóa tên để khớp giữa 2 file: bỏ dấu, hoa thường, gộp khoảng trắng."""
    s = str(s).lower().strip()
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return _re.sub(r'\s+', ' ', _re.sub(r'[^a-z0-9 ]', '', s)).strip()

def parse_visa_file(visa_bytes):
    """Đọc file thô visa/quản lý người nước ngoài → dict {"by_pp": {...}, "by_name": {...}}.
    Hỗ trợ 2 định dạng cột:
    1) File "Trang quản lý người nước ngoài" thật: 'HỌ TÊN' (tên đầy đủ 1 cột),
       'SỐ HỘ CHIẾU', và 'THỜI HẠN ĐƯỢC PHÉP TẠM TRÚ TẠI VIỆT NAM'.
    2) File cũ dạng Last Name / First Name / Visa date.
    Khớp theo SỐ HỘ CHIẾU trước (chính xác nhất, không sợ trùng tên/sai thứ tự),
    tên đã chuẩn hóa dùng làm dự phòng khi không có/không khớp số hộ chiếu.
    Tự động BỎ QUA khách Việt Nam (nếu file có cột quốc tịch) — công dân VN
    không có "thời hạn tạm trú tại Việt Nam" nên không cần đưa vào visa_map."""
    df = pd.read_excel(io.BytesIO(visa_bytes))
    def _find(*names):
        for n in names:
            for c in df.columns:
                if _norm_nat(c) == _norm_nat(n):
                    return c
        return None
    c_full  = _find('HỌ TÊN', 'HO TEN', 'Họ và tên', 'Full Name', 'FullName')
    c_last  = _find('Last Name', 'LastName', 'Họ')
    c_first = _find('First Name', 'FirstName', 'Tên')
    c_pp    = _find('SỐ HỘ CHIẾU', 'So Ho Chieu', 'Passport', 'Passport Number', 'PassportNo')
    c_nat   = _find('QUỐC TỊCH', 'MÃ QUỐC TỊCH', 'Nationality', 'LOẠI KHÁCH')
    c_date  = _find('Visa date', 'Visadate', 'Thời hạn tạm trú',
                     'THỜI HẠN ĐƯỢC PHÉP TẠM TRÚ TẠI VIỆT NAM',
                     'Thoi han duoc phep tam tru tai Viet Nam', 'Tam tru den')
    if c_date is None:
        raise ValueError("File visa không có cột ngày visa/thời hạn tạm trú. Vui lòng kiểm tra lại file.")

    def _is_vn(val):
        """Nhận diện khách Việt Nam qua cột quốc tịch/loại khách (nếu có)."""
        s = _norm_nat(val)
        return s in ('vnm', 'viet nam', 'vietnam') or s.startswith('vnm') or s == 'viet nam'

    by_pp = {}
    by_name = {}
    n_skipped_vn = 0
    for _, r in df.iterrows():
        # Bỏ qua khách Việt Nam nếu nhận diện được qua cột quốc tịch/loại khách
        if c_nat and pd.notna(r[c_nat]) and _is_vn(r[c_nat]):
            n_skipped_vn += 1
            continue
        d = r[c_date]
        if pd.isna(d):
            continue
        # Cột ngày đọc THẲNG (month=tháng thật, day=ngày thật) — KHÔNG đảo như
        # cột Departure. Đã kiểm chứng: có ngày 25/26/31 nên day chính là ngày thật.
        if hasattr(d, 'year') and not isinstance(d, str):
            dstr = f"{d.day:02d}/{d.month:02d}/{d.year}"
        else:
            s = str(d).strip()
            p = s.split('/')
            if len(p) == 3:
                dd, mm, yy = p
                if len(yy) == 2: yy = '20' + yy
                try:
                    dstr = f"{int(dd):02d}/{int(mm):02d}/{yy}"
                except Exception:
                    continue
            else:
                try:
                    t = pd.to_datetime(s, dayfirst=True)
                    dstr = f"{t.day:02d}/{t.month:02d}/{t.year}"
                except Exception:
                    continue

        # Khớp theo số hộ chiếu — ưu tiên, chính xác nhất
        if c_pp and pd.notna(r[c_pp]):
            pp_key = _norm_pp(r[c_pp])
            if pp_key:
                by_pp.setdefault(pp_key, dstr)

        # Khớp theo tên — dự phòng khi không có/không khớp số hộ chiếu
        if c_full and pd.notna(r[c_full]):
            key = _norm_name(str(r[c_full]))
            if key:
                by_name.setdefault(key, dstr)
        else:
            ln = str(r[c_last]).strip() if c_last and pd.notna(r[c_last]) else ''
            fn = str(r[c_first]).strip() if c_first and pd.notna(r[c_first]) else ''
            # Lưu cả 2 thứ tự để khớp linh hoạt dù file NNN đảo Họ/Tên
            for combo in ((ln + ' ' + fn), (fn + ' ' + ln)):
                key = _norm_name(combo)
                if key:
                    by_name.setdefault(key, dstr)
    return {"by_pp": by_pp, "by_name": by_name, "skipped_vn": n_skipped_vn}

def build_kbtt(df_intl, visa_map=None):
    """Điền mẫu KBTT. Dòng 3 là dòng "[TEST] SAMPLE" BẮT BUỘC giữ nguyên (không
    bị ghi đè) — dữ liệu khách thật được điền bắt đầu từ dòng 4 trở xuống.
    Cột L 'THỜI HẠN ĐƯỢC PHÉP TẠM TRÚ TẠI VIỆT NAM' CHỈ điền khi có visa_map
    từ file Visa rời do lễ tân chủ động upload (khớp theo SỐ HỘ CHIẾU trước,
    tên là dự phòng). KHÔNG tự đọc cột 'TẠM TRÚ'/'TẠM TRÚ ĐẾN NGÀY' có sẵn
    trong file dữ liệu khách nữa — cột đó thường do PMS tự điền mặc định
    (vd trùng ngày đi), không phải visa đã xác nhận thật; tin nhầm có thể
    khai sai thời hạn tạm trú lên hồ sơ chính thức nộp công an.
    Trả về (wb, danh_sách_tên_không_khớp, nguồn_visa, invalid_ids)."""
    visa_map = visa_map or {}
    # Tương thích ngược: nếu visa_map là dict phẳng {tên: ngày} kiểu cũ, coi như by_name
    if isinstance(visa_map, dict) and ("by_pp" in visa_map or "by_name" in visa_map):
        by_pp = visa_map.get("by_pp", {})
        by_name = visa_map.get("by_name", {})
    else:
        by_pp = {}
        by_name = visa_map
    unmatched = []
    invalid_ids = []  # khách có SỐ GIẤY TỜ chỉ là mã tạm nội bộ (GKS/GBL...), đã bị để trống
    wb = load_workbook(io.BytesIO(load_template('kbtt')))
    ws = wb['KBTT']
    # Cấu trúc mẫu: dòng 1 = ô merge A1:L1 (tiêu đề + chú ý đỏ), dòng 2 = header,
    # dòng 3 = "[TEST] SAMPLE" BẮT BUỘC giữ nguyên, dữ liệu khách thật từ dòng 4.
    ref = [ws.cell(3,c) for c in range(1,13)]   # dùng style dòng 3 làm mẫu định dạng cho các dòng khách
    n = len(df_intl)
    # Xóa dòng dữ liệu thừa (nếu có), luôn giữ tối thiểu tới dòng 3 (dòng TEST)
    last_data_row = 3 + n
    if ws.max_row > last_data_row:
        ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)
    for i,(_,row) in enumerate(df_intl.iterrows(),1):
        er=i+3   # dữ liệu khách thật bắt đầu dòng 4 (dòng 3 là TEST, giữ nguyên)
        ht=str(row.get('HỌ TÊN ',row.get('HỌ TÊN',''))).strip()
        ns=fmt(row['NGÀY SINH']); nd=fmt(row['NGÀY ĐẾN']); ni=fmt(row.get('NGÀY ÐI',row.get('NGÀY ĐI','')))
        ni=_fix_departure_swap(ni, nd)
        gt='M - Nam' if str(row.get('GIỚI TÍNH','')).strip()=='Nam' else 'F - Nữ'
        qt=lookup_nat_kbtt(row.get('QUỐC TỊCH',''))
        sh=str(row.get('SỐ GIẤY TỜ','')).strip(); sp=str(row.get('SỐ PHÒNG','')).strip()
        # SỐ GIẤY TỜ chỉ là mã tạm nội bộ (GKS/GBL/GKA/GBS = trẻ em dùng giấy
        # khai sinh/giấy bảo lãnh, chưa có hộ chiếu thật) — KHÔNG được nộp lên
        # hồ sơ KBTT như số hộ chiếu thật, để trống + cảnh báo thay vì điền sai.
        if sh.upper() in ('GKS','GBL','GKA','GBS'):
            invalid_ids.append(ht); sh=''
        # Cột L (12) — THỜI HẠN ĐƯỢC PHÉP TẠM TRÚ TẠI VIỆT NAM — CHỈ điền từ
        # file Visa rời (upload thủ công), khớp theo SỐ HỘ CHIẾU trước, tên là dự phòng
        vd = ''
        if by_pp or by_name:
            vd = by_pp.get(_norm_pp(sh), '') or by_name.get(_norm_name(ht), '')
            if not vd:
                unmatched.append(ht)
        vals=[i,ht,ns,'D - Ngày',gt,qt,sh,sp,nd,ni,ni,vd]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(er,ci); cell.value=val if isinstance(val,int) else str(val)
            cp(ref[ci-1],cell)
    # Bảo toàn ô merge tiêu đề + chiều cao dòng 1 (phòng khi delete_rows làm xê dịch)
    if 'A1:L1' not in [str(m) for m in ws.merged_cells.ranges]:
        try: ws.merge_cells('A1:L1')
        except Exception: pass
    ws.row_dimensions[1].height = 41.1
    # Khôi phục rich text ô A1: tiêu đề (đen) + dòng chú ý (ĐỎ) — openpyxl làm mất khi save
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles.colors import Color
        from openpyxl.styles import Alignment
        a1 = ws.cell(1,1)
        a1.value = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=16, b=True),
                      'DANH SÁCH HỒ SƠ KBTT\r\n'),
            TextBlock(InlineFont(rFont='Times New Roman', sz=16, b=True, color=Color(rgb='FFFF0000')),
                      '(*Lưu ý: Người khai báo chịu trách nhiệm trước pháp luật về các nội dung thông tin cung cấp)')
        )
        a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except Exception:
        pass
    # Cập nhật vùng Table1 cho khớp số dòng thực (header + dòng TEST + dữ liệu khách)
    if 'Table1' in ws.tables:
        ws.tables['Table1'].ref = f"A2:L{3 + n}"
    return wb, unmatched, ('file' if (by_pp or by_name) else None), invalid_ids

def build_vnm(df_vn):
    wb = load_workbook(io.BytesIO(load_template('vnm')))
    wsn = next((s for s in wb.sheetnames if 'KHACH' in s or 'DS' in s), wb.sheetnames[0])
    ws = wb[wsn]
    ref = [ws.cell(5,c) for c in range(1,ws.max_column+1)]
    for r in range(ws.max_row,4,-1): ws.delete_rows(r)
    gks_cnt=0; gbl_cnt=0
    ward_unmatched=[]  # (tên khách, phường/xã gốc) không tự tra được mã — giữ raw, cần lễ tân kiểm tra
    _CUTRU_MAP={'thuongtru':'1 - Thường trú','tamtru':'2 - Tạm trú'}
    for i,(_,row) in enumerate(df_vn.iterrows(),1):
        er=i+4
        ht=str(row.get('HỌ TÊN ',row.get('HỌ TÊN',''))).strip()
        ns=fmt(row['NGÀY SINH']); nd=fmt(row['NGÀY ĐẾN']); ni=fmt(row.get('NGÀY ÐI',row.get('NGÀY ĐI','')))
        ni=_fix_departure_swap(ni, nd)
        gt='F - Nữ' if str(row.get('GIỚI TÍNH','')).strip()=='Nữ' else 'M - Nam'
        sg_raw=str(row.get('SỐ GIẤY TỜ','')).strip()
        lg_raw=str(row.get('LOẠI GIẤY TỜ','')).strip()
        is_gks='GKS' in sg_raw.upper(); is_gbl='GBL' in sg_raw.upper()
        ten_giay=''
        if is_gks:
            # Mã tạm nội bộ (chưa có giấy khai sinh thật cấp số) — theo đúng
            # danh mục DANH_MUC của mẫu, KHÔNG dùng "5 - Giấy khai sinh" (mã
            # đó dành cho giấy khai sinh CÓ số thật), ghi "9 - Giấy Tờ Khác"
            # + nêu rõ loại trong TÊN GIẤY TỜ để tránh khai sai giấy tờ.
            sg=make_code('GKS',ns); lg='9 - Giấy Tờ Khác'; ten_giay='giấy khai sinh'; gks_cnt+=1
        elif is_gbl:
            sg=make_code('GBL',ns); lg='9 - Giấy Tờ Khác'; ten_giay='giấy bảo lãnh'; gbl_cnt+=1
        elif sg_raw and sg_raw[0].isalpha():
            # Số giấy tờ bắt đầu bằng chữ cái (vd: P02628567) → là hộ chiếu
            sg=sg_raw; lg='4 - Hộ chiếu'
        else:
            sg=sg_raw; lg=LOAI_GIAY.get(lg_raw,lg_raw)
        # Dùng _fmt_room (không phải str(x).strip() thẳng) vì ô trống trong
        # Excel đọc qua pandas ra NaN (float) — str(nan)='nan' vẫn là chuỗi
        # KHÔNG rỗng nên lọt qua điều kiện "if tinh_raw" và bị tra cứu/in ra
        # chữ "nan" thẳng vào file VNM, dù thực chất khách đó không có dữ liệu.
        tinh_raw=_fmt_room(row.get('TP/TỈNH',''))
        tinh=lookup_province_vnm(tinh_raw) if tinh_raw else ''
        phuong_raw=_fmt_room(row.get('PHƯỜNG/XÃ',''))
        phuong, ward_ok, inferred_prov = lookup_ward_vnm(phuong_raw, tinh or None)
        if phuong_raw and not ward_ok:
            ward_unmatched.append((ht, phuong_raw))
        if inferred_prov and not tinh:
            tinh = inferred_prov
        dc=str(row.get('ÐỊA CHỈ',row.get('ĐỊA CHỈ',''))).strip()
        sp=str(row.get('SỐ PHÒNG','')).strip()
        dt=_fmt_room(row.get('SỐ ĐIỆN THOẠI',''))  # phòng vệ thêm nếu cột vẫn lọt qua dạng số (mất số 0 đầu)
        cutru=_CUTRU_MAP.get(_norm_nat(row.get('THƯỜNG TRÚ / TẠM TRÚ','')),'1 - Thường trú')
        vals=[i,ht,ns,gt,'VNM - Viet Nam',lg,ten_giay,sg,dt,cutru,tinh,phuong,dc,nd,ni,sp,'1 - Du lịch','','']
        for ci,val in enumerate(vals,1):
            cell=ws.cell(er,ci); cell.value=val if isinstance(val,int) else str(val)
            if ci<=len(ref): cp(ref[ci-1],cell)
    return wb, gks_cnt, gbl_cnt, ward_unmatched

def _dk_map_gender(g):
    """Chuẩn hóa giới tính về 'Nam'/'Nữ' — chấp nhận cả chữ cái đơn (M/F, kể
    cả không hoa) lẫn tiếng Việt đầy đủ (Nam/Nữ), khác bản cũ chỉ nhận đúng
    ký tự 'M'/'F' (sai hoàn toàn nếu nguồn ghi 'Nam'/'Nữ' như mọi cột GIỚI
    TÍNH khác trong app — khi đó ngày sinh sẽ không vào được cột nào)."""
    s = str(g or '').strip().lower()
    if s in ('nam', 'm', 'male'):
        return 'Nam'
    if s in ('nữ', 'nu', 'f', 'female'):
        return 'Nữ'
    return str(g or '').strip()

def _dk_is_dummy(name, room):
    """Nhận diện dòng dummy/test — không phải khách lưu trú thật, không đưa
    vào sổ ĐK14 chính thức nộp công an."""
    n = str(name or '').strip().lower()
    if not n:
        return True, 'Tên trống'
    if 'dummy' in n:
        return True, 'Tên là dummy'
    if 'pending' in n:
        return True, 'Tên là pending'
    if 'water sport' in n:
        return True, 'Không phải khách lưu trú'
    if _re.fullmatch(r'[A-Za-z\s]+\d{4,}', str(name or '').strip()):
        return True, 'Tên chứa mã đặt phòng'
    digits = _re.sub(r'\D', '', str(room or ''))
    if digits and int(digits) >= 9000:
        return True, f'Phòng {room} ≥ 9000 (phòng ảo/posting master)'
    return False, ''

# ── ĐK14: chuẩn hoá giá trị đọc từ file nguồn ─────────────────────────────
# xlrd/openpyxl trả ô ngày về SỐ SERIAL Excel (không phải datetime) khi đọc
# thô — cố tình đọc thô vì mọi thư viện đọc sẵn kiểu ngày đều có nguy cơ lệch
# múi giờ; công thức (serial - 25569) * 86400 giây quy đổi thẳng nên không lệch.
def _dk_cell_str(v):
    """Ô số nguyên (số giấy tờ, số phòng...) bị đọc thành float — bỏ đuôi '.0'
    để không ghi '123456789.0' vào sổ nộp công an."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def _dk_to_date(v):
    """Mọi kiểu ô ngày (serial Excel, datetime, chuỗi dd/mm/yyyy) → date."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (datetime.datetime(1970, 1, 1)
                    + datetime.timedelta(seconds=round((float(v) - 25569) * 86400))).date()
        except Exception:
            return None
    m = _re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', str(v).strip())
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None

def _dk_fmt_date(v):
    d = _dk_to_date(v)
    return d.strftime('%d/%m/%Y') if d else ''

def _dk_is_invalid_id(v):
    """Mã giấy tờ CHƯA CẤP (không phải số thật) — xoá khỏi sổ. GKS/GBL được xử
    lý riêng ở nơi gọi vì đó là dữ liệu hợp lệ (giấy khai sinh / giấy bảo lãnh)."""
    s = str(v or '').strip().upper()
    return len(s) <= 4 and bool(_re.fullmatch(r'[A-Z]{2,4}', s)) and s in ('GKS', 'GBL', 'GKA', 'GBS')

def _dk_detect_columns(rows):
    """Tự dò dòng tiêu đề (trong 5 dòng đầu) và vị trí từng cột theo TÊN tiêu
    đề, thay vì bám chỉ số cột cứng — file IH đổi thứ tự/thêm cột vẫn chạy đúng."""
    header_row, headers = -1, None
    for i in range(min(5, len(rows))):
        r = rows[i] or []
        has_name = any(c is not None and _re.match(r'^(họ và tên|name|full name)', str(c).strip(), _re.I) for c in r)
        has_dob = any(c is not None and _re.search(r'(ngày sinh|date of birth|ngày tháng năm sinh|dob)', str(c), _re.I) for c in r)
        if has_name or has_dob:
            header_row, headers = i, r
            break
    if header_row == -1:
        header_row, headers = 0, (rows[0] if rows else [])
    cols = {'headerRow': header_row, 'dataStart': header_row + 1}
    for idx, cell in enumerate(headers or []):
        if cell is None:
            continue
        s = str(cell).lower().strip()
        if not s:
            continue
        if s in ('stt', 'no', 'no.', 'số tt') or s.startswith('stt '):
            cols['stt'] = idx
        elif _re.search(r'họ và tên (khách|kh)', s) or s in ('họ và tên', 'name', 'full name', 'guest name'):
            cols['name'] = idx
        elif _re.search(r'(ngày sinh|date of birth|ngày tháng năm sinh|dob)', s) and 'thông báo' not in s:
            cols['dob'] = idx
        elif _re.fullmatch(r'(giới tính|gender|sex)', s) or s.startswith('giới tính') or s.startswith('gender'):
            cols['gender'] = idx
        elif 'quốc tịch' in s:
            cols['nationality'] = idx
        elif 'quốc gia' in s and 'nationality' not in cols:
            cols['country'] = idx
        elif _re.search(r'country.*residence|residence.*country|nationality', s):
            cols['nationality'] = idx
        elif 'country' in s and 'nationality' not in cols:
            cols['country'] = idx
        elif _re.match(r'^(số giấy tờ|passport|cccd|cmnd|id card)', s) or _re.search(r'(passport|id card)', s):
            cols['id'] = idx
        elif s.startswith('loại giấy tờ'):
            cols['docType'] = idx
        elif s.startswith('tên giấy tờ'):
            cols['docName'] = idx
        elif _re.search(r'(số điện thoại|phone|tel|mobile)', s):
            cols['phone'] = idx
        elif 'loại cư trú' in s:
            cols['cuTru'] = idx
        elif _re.match(r'^(tỉnh|tp|province|city)', s) or s.startswith('tỉnh/'):
            cols['tinh'] = idx
        elif _re.match(r'^(quận|huyện|district)', s) or s.startswith('quận/'):
            cols['quan'] = idx
        elif _re.match(r'^(phường|xã|ward|commune)', s) or s.startswith('phường/'):
            cols['phuong'] = idx
        elif _re.search(r'(địa chỉ chi tiết|address line|^address$)', s):
            cols['addressDetail'] = idx
        elif s.startswith('địa chỉ') and 'addressDetail' not in cols:
            cols['addressDetail'] = idx
        elif _re.search(r'(arrival|ngày đến|thời gian.*đến|check.in)', s) or s.startswith('đến'):
            cols['arrival'] = idx
        elif _re.search(r'(depature|departure|ngày đi|thời gian.*đi|check.out)', s) or s.startswith('đi'):
            cols['departure'] = idx
        elif _re.search(r'(số buồng|số phòng|^phòng|room\s*#|^room$|room number)', s) or 'phòng/khoa' in s:
            cols['room'] = idx
        elif _re.search(r'(người thông báo|notifier|reporter)', s):
            cols['notifier'] = idx
    return cols, (headers or [])

def _dk_addr_from_parts(cu_tru, tinh, quan, phuong, detail, iso):
    """File IH không có cột địa chỉ gộp thì ghép từ các cột rời. CHỈ ghép cho
    khách Việt Nam — khách nước ngoài để trống theo đúng mẫu ĐK14."""
    if iso != 'VNM':
        return '   '
    cu_tru = str(cu_tru or '').strip()
    segs = [str(p or '').strip() for p in (detail, phuong, quan, tinh)]
    segs = [p for p in segs if p and p not in ('null', 'undefined', 'nan')]
    addr = ', '.join(segs)
    if not addr and not cu_tru:
        return '   '
    if cu_tru:
        return f'{cu_tru}: {addr}' if addr else cu_tru
    return addr

def _dk_read_rows(file_bytes):
    """Đọc file nguồn IH ở dạng THÔ (ô ngày giữ nguyên số serial). Nhận cả .xls
    (xlrd) lẫn .xlsx (openpyxl) — nhận diện bằng chữ ký file, không theo đuôi."""
    if file_bytes[:2] == b'PK':
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

def _dk_transform(rows, default_notifier, default_checkin):
    """IH → các dòng dữ liệu ĐK14, kèm nhật ký từng dòng đã sửa/bỏ qua."""
    cols, headers = _dk_detect_columns(rows)
    issues, skipped, data = [], [], []
    missing = [k for k in ('name', 'dob') if k not in cols]
    if missing:
        issues.append(('warn', cols['headerRow'] + 1, '(tiêu đề)',
                       'Không tìm thấy cột: ' + ', '.join(missing)))

    def get(row, key):
        i = cols.get(key)
        return row[i] if i is not None and i < len(row) else None

    seen_ids, stt = set(), 0
    for i in range(cols['dataStart'], len(rows)):
        row = rows[i] or []
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        raw_name = get(row, 'name')
        name_str = str(raw_name or '')
        room = _dk_cell_str(get(row, 'room'))
        is_dummy, reason = _dk_is_dummy(name_str, room)
        if is_dummy:
            skipped.append((name_str.strip() or '(trống)', reason))
            issues.append(('skip', i + 1, name_str.strip() or '(trống)', f'Bỏ qua — {reason}'))
            continue
        stt += 1
        note = []
        name = name_str.strip()
        if name_str != name:
            note.append('Trim khoảng trắng tên')
        gender = _dk_map_gender(get(row, 'gender'))
        dob = _dk_fmt_date(get(row, 'dob'))
        if not dob:
            note.append('Thiếu ngày sinh')
        if gender not in ('Nam', 'Nữ'):
            note.append('Thiếu/không xác định giới tính')
        iso = _dk_cell_str(get(row, 'nationality') or get(row, 'country')).upper()
        nat = NAT_DK14.get(iso, iso)
        if iso and iso not in NAT_DK14:
            note.append(f'Mã quốc tịch "{iso}" không có trong danh mục')

        id_num = _dk_cell_str(get(row, 'id'))
        id_up = id_num.upper()
        if id_up in ('GKS', 'GBL'):
            # Giấy khai sinh (GKS) / giấy bảo lãnh (GBL) — dữ liệu HỢP LỆ (khách
            # dùng thay CCCD/hộ chiếu). Giữ nguyên, không xoá, không tính trùng
            # vì nhiều khách có thể cùng dùng chung 1 mã.
            pass
        elif _dk_is_invalid_id(id_num):
            note.append(f'Số giấy tờ "{id_num}" chưa cấp → xóa')
            id_num = ''
        elif not id_num:
            note.append('Thiếu số giấy tờ')
        elif id_num in seen_ids:
            note.append(f'Số giấy tờ "{id_num}" trùng dòng trước')
        else:
            seen_ids.add(id_num)

        if 'addressDetail' in cols:
            addr = _dk_cell_str(get(row, 'addressDetail')).strip() or '   '
        else:
            addr = _dk_addr_from_parts(get(row, 'cuTru'), get(row, 'tinh'), get(row, 'quan'),
                                       get(row, 'phuong'), get(row, 'addressDetail'), iso)

        arr = _dk_to_date(get(row, 'arrival'))
        if arr is None:
            arr = default_checkin
            if 'arrival' in cols:
                note.append('Thiếu ngày đến → dùng ngày mặc định')
        dep = _dk_to_date(get(row, 'departure'))
        if dep is None:
            note.append('Thiếu ngày đi')

        notifier = _dk_cell_str(get(row, 'notifier')).strip() or default_notifier

        if note:
            kind = 'fix' if any('→' in x or 'Trim' in x for x in note) else 'warn'
            issues.append((kind, i + 1, name, ' · '.join(note)))
        data.append([stt, name,
                     dob if gender == 'Nam' else '', dob if gender == 'Nữ' else '',
                     nat, id_num, addr, arr, dep, room, notifier, '', ''])
    return data, issues, skipped, cols, headers

# ── ĐK14: ghi file bằng cách CHÈN THẲNG XML vào mẫu ──────────────────────
# Chỉ thay phần <sheetData> từ dòng 18 trở đi, giữ nguyên byte mọi thứ còn lại
# của file mẫu (kiểu ô, viền, font, gộp ô, khổ giấy, hình vẽ, cấu hình in...).
# Cách này giữ được nhiều hơn so với dựng lại workbook bằng openpyxl — openpyxl
# không giữ hình vẽ và cấu hình máy in.
_DK_COLS = 'ABCDEFGHIJKLM'
_DK_DEFAULT_STYLES = {'A': '8', 'B': '8', 'C': '19', 'D': '19', 'E': '8', 'F': '13',
                      'G': '8', 'H': '8', 'I': '8', 'J': '8', 'K': '8', 'L': '8', 'M': '8'}
_DK_DEFAULT_ROW_ATTRS = 's="4" customFormat="1" ht="14.25" x14ac:dyDescent="0.15"'

def _dk_esc(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))

def _dk_template_styles(xml):
    """Lấy kiểu ô/thuộc tính dòng từ chính các dòng dữ liệu mẫu (18–25) để dòng
    sinh ra trông y hệt mẫu; thiếu thì dùng bộ mặc định của mẫu TT30/2026."""
    col_styles, row_attrs = {}, ''
    for test_row in range(18, 26):
        m = _re.search(r'<row([^>]*\br="%d"[^>]*)>(.*?)</row>' % test_row, xml, _re.S)
        if not m:
            continue
        if not row_attrs:
            attrs = [f'{k}="{v}"' for k, v in _re.findall(r'(\w+(?::\w+)?)="([^"]*)"', m.group(1))
                     if k not in ('r', 'spans')]
            row_attrs = ' '.join(attrs)
        for col, cattr in _re.findall(r'<c\s+r="([A-Z]+)%d"([^>/]*)' % test_row, m.group(2)):
            sm = _re.search(r'\ss="(\d+)"', cattr)
            if sm and col not in col_styles:
                col_styles[col] = sm.group(1)
    for col, style in _DK_DEFAULT_STYLES.items():
        col_styles.setdefault(col, style)
    return col_styles, (row_attrs or _DK_DEFAULT_ROW_ATTRS)

def _dk_row_xml(row_num, values, col_styles, row_attrs):
    out = [f'<row r="{row_num}" spans="1:13" {row_attrs}>']
    for ci, col in enumerate(_DK_COLS):
        val = values[ci] if ci < len(values) else ''
        ref = f'{col}{row_num}'
        s_attr = f' s="{col_styles[col]}"' if col_styles.get(col) else ''
        if val is None or val == '':
            out.append(f'<c r="{ref}"{s_attr}/>')
        elif isinstance(val, (datetime.date, datetime.datetime)):
            # Ghi ngày dạng CHỮ "dd/mm/yyyy" thay vì số serial: số serial chỉ hiện
            # đúng khi ô có định dạng ngày, mà mẫu không đảm bảo điều đó.
            out.append(f'<c r="{ref}"{s_attr} t="inlineStr"><is><t>'
                       f'{val.strftime("%d/%m/%Y")}</t></is></c>')
        elif isinstance(val, int) and not isinstance(val, bool):
            out.append(f'<c r="{ref}"{s_attr}><v>{val}</v></c>')
        else:
            sv = str(val)
            t_attr = ' xml:space="preserve"' if sv != sv.strip() else ''
            out.append(f'<c r="{ref}"{s_attr} t="inlineStr"><is><t{t_attr}>'
                       f'{_dk_esc(sv)}</t></is></c>')
    out.append('</row>')
    return ''.join(out)

def _dk_set_cell_text(xml, ref, text):
    """Ghi đè 1 ô của mẫu bằng chuỗi (kể cả ô đang trỏ tới sharedStrings)."""
    esc = _dk_esc(text)
    t_attr = ' xml:space="preserve"' if text != text.strip() else ''
    def _sub(m):
        attrs = _re.sub(r'\s+t="[^"]*"', '', m.group(1))
        return f'<c r="{ref}"{attrs} t="inlineStr"><is><t{t_attr}>{esc}</t></is></c>'
    return _re.sub(r'<c\s+r="%s"([^>/]*?)(?:/>|>.*?</c>)' % ref, _sub, xml, flags=_re.S)

def build_dk14(xls_bytes, default_notifier='', default_checkin=None):
    """IH → file ĐK14 .xlsx. Trả (bytes file, số khách, danh sách bỏ qua,
    nhật ký dòng, thông tin cột đã dò được)."""
    rows = _dk_read_rows(xls_bytes)
    data, issues, skipped, cols, headers = _dk_transform(rows, default_notifier, default_checkin)

    src = io.BytesIO(load_template('dk14'))
    out = io.BytesIO()
    sheet_path = 'xl/worksheets/sheet1.xml'
    with zipfile.ZipFile(src) as zin:
        if sheet_path not in zin.namelist():
            raise ValueError('Mẫu ĐK14 không có xl/worksheets/sheet1.xml')
        xml = zin.read(sheet_path).decode('utf-8')

        col_styles, row_attrs = _dk_template_styles(xml)
        deps = [r[8] for r in data if isinstance(r[8], datetime.date)]
        xml = _dk_set_cell_text(xml, 'J11',
                                'Từ ngày: ' + (default_checkin.strftime('%d/%m/%Y') if default_checkin else ''))
        xml = _dk_set_cell_text(xml, 'J12',
                                'Đến ngày: ' + (max(deps).strftime('%d/%m/%Y') if deps else ''))

        new_rows = ''.join(_dk_row_xml(18 + i, r, col_styles, row_attrs) for i, r in enumerate(data))
        end = xml.index('</sheetData>')
        start = -1
        for m in _re.finditer(r'<row\s+r="(\d+)"[^>]*>', xml):
            if int(m.group(1)) >= 18:
                start = m.start()
                break
        xml = (xml[:end] if start == -1 else xml[:start]) + new_rows + xml[end:]

        last_row = max(17 + len(data), 18)
        xml = _re.sub(r'<dimension\s+ref="[^"]*"\s*/>', f'<dimension ref="A1:M{last_row}"/>', xml)

        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, xml.encode('utf-8') if item.filename == sheet_path
                              else zin.read(item.filename))
    return out.getvalue(), len(data), skipped, issues, (cols, headers)


# ── Regcard PDF builder ───────────────────────────────────────────────────
@st.cache_resource
def load_regcard_template():
    path = os.path.join(os.path.dirname(__file__), 'tmpl_regcard.b64')
    with open(path, 'r') as f:
        return base64.b64decode(f.read())

def load_group_template():
    path = os.path.join(os.path.dirname(__file__), 'tmpl_group.b64')
    with open(path, 'r') as f:
        return base64.b64decode(f.read())

def _grp_date(d):
    """Ngày cho regcard group: dd/mm/yyyy (như mẫu 24/07/2026)."""
    if pd.isna(d): return ''
    if hasattr(d, 'strftime'):
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    s = str(d).strip()
    if '/' in s:
        p = s.split('/')
        if len(p) == 3:
            dd, mm, yy = p
            if len(yy) == 2: yy = '20' + yy
            return f"{int(dd):02d}/{int(mm):02d}/{yy}"
        return s
    # Chuỗi ISO yyyy-mm-dd
    try:
        t = pd.to_datetime(s)
        return f"{t.day:02d}/{t.month:02d}/{t.year}"
    except Exception:
        return s

def _rc_clean_name(n):
    if pd.isna(n): return ''
    return str(n).strip().rstrip(',').strip()

def _rc_conf(c):
    if pd.isna(c): return ''
    return str(int(c)) if isinstance(c,(int,float)) else str(c)

def _rc_date(d):
    if pd.isna(d): return ''
    if hasattr(d,'strftime') and not isinstance(d, str):
        return f"{d.day:02d}/{d.month:02d}/{d.year}"   # dd/mm/yyyy
    s=str(d).strip()
    if '/' in s:
        p=s.split('/')
        if len(p)==3:
            dd,mm,yy=p
            if len(yy)==2: yy='20'+yy
            return f"{int(dd):02d}/{int(mm):02d}/{yy}"   # pad 7/8 → 07/08
        return s
    try:
        t=pd.to_datetime(s)                            # ISO yyyy-mm-dd
        return f"{t.day:02d}/{t.month:02d}/{t.year}"
    except Exception:
        return s

def _rc_nights(arr, dep):
    """Số đêm = Departure - Arrival. Xử lý cả datetime lẫn chuỗi dd/mm/yyyy."""
    def _to_ts(v):
        if pd.isna(v): return None
        if hasattr(v, 'year') and not isinstance(v, str):
            return pd.Timestamp(v)          # datetime/Timestamp — giữ nguyên
        s = str(v).strip()
        if '/' in s:                        # chuỗi dd/mm/yyyy hoặc dd/mm/yy
            p = s.split('/')
            if len(p) == 3:
                dd, mm, yy = p
                if len(yy) == 2: yy = '20' + yy
                try:
                    return pd.Timestamp(year=int(yy), month=int(mm), day=int(dd))
                except Exception:
                    return None
        try:
            return pd.to_datetime(s)        # ISO yyyy-mm-dd hoặc dạng khác
        except Exception:
            return None
    a, d = _to_ts(arr), _to_ts(dep)
    if a is None or d is None:
        return ''
    n = (d - a).days
    return str(n) if n >= 0 else ''

def build_group_regcard(grp_df, tmpl_bytes):
    """Vẽ 1 Registration Card for Group từ các dòng cùng 1 mã Group.
    Trả về trang PDF đã merge. Bảng Kind of rooms để trống (điền tay)."""
    H = 841.0
    FONT = "Times-Roman"; SIZE = 11
    first = grp_df.iloc[0]

    # Group Code = giá trị cột Group
    gcode = first.get('Group')
    if pd.notna(gcode):
        gcode = str(int(gcode)) if isinstance(gcode, (int, float)) else str(gcode)
    else:
        gcode = ''

    # Số phòng: đếm phòng unique, loại phòng ảo 9xxx
    rooms = set()
    for r in grp_df['Rm'].dropna():
        s = str(r).strip()
        if s.endswith('.0'): s = s[:-2]
        if s and not _re.fullmatch(r'9\d{3}', s):
            rooms.add(s.upper())
    n_rooms = len(rooms)

    # Số pax = tổng Adt + Chl + Enf
    def _sum(col):
        return int(grp_df[col].fillna(0).sum()) if col in grp_df.columns else 0
    n_pax = _sum('Adt') + _sum('Chl') + _sum('Enf')

    data = {
        'arrival':   (155.6, 181.2, _grp_date(first.get('Arrival'))),
        'departure': (436.9, 181.2, _grp_date(first.get('Departure'))),
        'gcode':     (48.2,  261.6, gcode),
        'gname':     (165.7, 263.9, _rc_clean_name(first.get('Name'))),
        'agent':     (364.4, 263.9, str(first.get('Company')) if pd.notna(first.get('Company')) else ''),
        'nrooms':    (50.9,  330.6, str(n_rooms)),
        'npax':      (187.2, 326.2, str(n_pax)),
    }

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(595, 841))
    c.setFillColor(black)
    c.setFont(FONT, SIZE)
    for key, (x, bottom, val) in data.items():
        if not val:
            continue
        # thu nhỏ nếu tên đoàn / hãng quá dài
        maxw = {'gname': 185, 'agent': 175}.get(key)
        fs = SIZE
        if maxw:
            while fs > 7 and c.stringWidth(val, FONT, fs) > maxw:
                fs -= 0.3
        c.setFont(FONT, fs)
        c.drawString(x, H - bottom, val)
        c.setFont(FONT, SIZE)
    c.save(); buf.seek(0)

    base = PdfReader(io.BytesIO(tmpl_bytes))
    overlay = PdfReader(buf)
    page = base.pages[0]
    page.merge_page(overlay.pages[0])
    return page


def _fix_date(v):
    """Chuẩn hóa ngày về pd.Timestamp đúng nghĩa dd/mm.

    File Smile export bị lỗi: ngày dạng dd/mm với ngày ≤ 12 (vd '7/8/2026' = 7 tháng 8)
    bị Excel hiểu nhầm kiểu Mỹ mm/dd → lưu thành datetime(month=7, day=8) kèm format
    mm-dd-yy. Khi đọc lại, cần HOÁN month↔day để khôi phục: datetime(y,7,8) → 7 tháng 8.
    Ngày ≥ 13 thì Excel không nhầm được nên giữ dạng text dd/mm bình thường.
    """
    if v is None or (not isinstance(v, str) and pd.isna(v)):
        return None
    # datetime từ Excel → đã bị đảo month/day, khôi phục bằng cách hoán lại
    if hasattr(v, 'year') and not isinstance(v, str):
        try:
            return pd.Timestamp(year=v.year, month=v.day, day=v.month)
        except Exception:
            return pd.Timestamp(v)   # day>12: không đảo được, giữ nguyên
    s = str(v).strip()
    if '/' in s:
        p = s.split('/')
        if len(p) == 3:
            dd, mm, yy = p
            if len(yy) == 2: yy = '20' + yy
            try:
                return pd.Timestamp(year=int(yy), month=int(mm), day=int(dd))
            except Exception:
                return None
    try:
        return pd.to_datetime(s, dayfirst=True)
    except Exception:
        return None

def _fix_departure_swap(ni_str, nd_str):
    """Nếu NGÀY ĐI (dd/mm/yyyy) đọc ra TRƯỚC NGÀY ĐẾN — thường do lỗi đảo
    dd/mm khi cả 2 số ≤12 — thử hoán dd↔mm; dùng bản đã hoán nếu nó không còn
    trước ngày đến nữa. Không đụng các trường hợp khác (giữ nguyên nếu không
    chắc chắn). Cùng heuristic đã kiểm chứng ở ARR/XNC Converter."""
    if not ni_str or not nd_str:
        return ni_str
    try:
        dep = datetime.datetime.strptime(ni_str, '%d/%m/%Y')
        arr = datetime.datetime.strptime(nd_str, '%d/%m/%Y')
    except Exception:
        return ni_str
    if dep >= arr:
        return ni_str
    m = _re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', ni_str)
    if not m:
        return ni_str
    dd, mm, yy = int(m.group(1)), int(m.group(2)), m.group(3)
    if dd > 12 or mm > 12 or dd == mm:
        return ni_str
    swapped = f"{mm:02d}/{dd:02d}/{yy}"
    try:
        sw = datetime.datetime.strptime(swapped, '%d/%m/%Y')
    except Exception:
        return ni_str
    return swapped if sw >= arr else ni_str

def _norm_addr(s):
    """Chuẩn hóa tên tỉnh/phường để so khớp: bỏ dấu, hoa/thường, bỏ tiền tố
    Xã/Phường/Thị trấn/Đặc khu/Tỉnh/TP..., bỏ ký tự không phải chữ/số."""
    s = str(s or '').strip()
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    s = s.lower()
    s = _re.sub(r'^(xa|phuong|thi tran|dac khu|tinh|thanh pho|tp\.?)\s+', '', s)
    s = _re.sub(r'[^a-z0-9 ]', ' ', s)
    s = _re.sub(r'\s+', ' ', s).strip()
    return s

@st.cache_resource(show_spinner=False)
def load_vn_admin_lookup():
    """Đọc bảng TINH_THANH + PHUONG_XA nhúng sẵn trong chính mẫu VNM (đúng
    danh mục dropdown DANH_MUC đang dùng, không phải bảng ngoài) → dict tra
    cứu theo tên đã chuẩn hóa, dùng để tự điền mã Tỉnh/Thành + Phường/Xã
    thay vì bỏ trống như trước."""
    wb = load_workbook(io.BytesIO(load_template('vnm')))
    prov_by_norm = {}
    prov_by_code = {}
    for row in wb['TINH_THANH'].iter_rows(min_row=2, values_only=True):
        matt, tentt, display = row[0], row[1], row[2]
        if not matt or not display:
            continue
        prov_by_code[str(matt)] = display
        prov_by_norm.setdefault(_norm_addr(tentt), display)
    ward_by_norm = {}
    for row in wb['PHUONG_XA'].iter_rows(min_row=2, values_only=True):
        ma, ten, matt, display = row[0], row[1], row[2], row[3]
        if not ma or not display:
            continue
        ward_by_norm.setdefault(_norm_addr(ten), []).append((str(matt), display))
    return prov_by_norm, prov_by_code, ward_by_norm

def lookup_province_vnm(raw):
    """Mã Tỉnh/Thành cho cột 'TỈNH/ THÀNH PHỐ' của mẫu VNM — khớp đúng tên
    trong chính mẫu (TINH_THANH) trước, dự phòng bảng TINH cũ (alias/viết
    tắt như 'TP HCM'), giữ nguyên raw nếu không khớp được (không đoán bừa)."""
    raw = str(raw or '').strip()
    if not raw:
        return ''
    prov_by_norm, _, _ = load_vn_admin_lookup()
    hit = prov_by_norm.get(_norm_addr(raw))
    if hit:
        return hit
    legacy = TINH.get(raw.upper())
    if legacy:
        return legacy
    return raw

def lookup_ward_vnm(raw, prov_display=None):
    """Mã Phường/Xã cho cột 'PHƯỜNG/ XÃ/ ĐẶC KHU' của mẫu VNM. Tên phường có
    thể trùng giữa nhiều tỉnh, nên khi đã biết tỉnh (prov_display dạng
    '101 - TP. Hà Nội') sẽ ưu tiên khớp đúng phường thuộc tỉnh đó. Không tự
    suy đoán khi mơ hồ — trả lại raw để lễ tân tự kiểm tra thay vì điền sai.
    Trả về (giá_trị_để_điền, đã_khớp: bool, tỉnh_suy_ra_được: str|None)."""
    raw = str(raw or '').strip()
    if not raw:
        return '', False, None
    _, prov_by_code, ward_by_norm = load_vn_admin_lookup()
    cands = ward_by_norm.get(_norm_addr(raw)) or []
    if not cands:
        return raw, False, None
    prov_code = None
    if prov_display:
        m = _re.match(r'^(\d+)\s*-', str(prov_display))
        if m:
            prov_code = m.group(1)
    if prov_code:
        scoped = [d for c, d in cands if c == prov_code]
        if scoped:
            return scoped[0], True, None
    if len(cands) == 1:
        c, d = cands[0]
        return d, True, (None if prov_display else prov_by_code.get(c))
    return raw, False, None

def build_regcards(xlsx_bytes, only_main=True):
    """Tạo PDF regcard hàng loạt, gộp theo Conf# (đoàn nhiều phòng → 1 regcard,
    các số phòng gộp chung vào ô Room No)."""
    df = pd.read_excel(io.BytesIO(xlsx_bytes))
    # Chuẩn hóa ngày (sửa lỗi Excel đảo dd/mm↔mm/dd với ngày ≤12 từ Smile export)
    for _dc in ('Arrival', 'Departure'):
        if _dc in df.columns:
            df[_dc] = df[_dc].map(_fix_date)
    H = 841.0
    FONT = "Times-Roman"; SIZE = 9.8
    # Baseline chính xác (bottom) đo từ dữ liệu mẫu gốc — chữ trùng khít 100%
    POS = {
        'name':(125.50,109.60),'conf':(526.75,108.52),'arrival':(119.90,144.22),
        'departure':(329.90,144.22),'nights':(500.30,144.22),'type':(113.60,179.92),
        'rm':(360.60,179.92),'company':(221.40,216.32),
        'special':(137.40,288.40),
    }
    # Ô che dữ liệu cũ — vừa khít vùng chữ, không lấn đường kẻ bảng
    BLANK = [
        (125,99,250,110.5),(526,98,568,109.5),(119,133.5,168,145.2),
        (329,133.5,378,145.2),(500,133.5,512,145.2),(113,169,145,180.9),
        (360,169,410,180.9),(221,205.5,320,217.3),
        (135,277,575,289.4),   # che dòng "AI Lunch ( EUR )..." in sẵn trên template
    ]

    # ── Gộp theo Conf# ──
    # Điền mã Conf# xuống các dòng trống (khách đi cùng booking),
    # rồi gộp TẤT CẢ dòng cùng 1 mã Conf# thành 1 regcard, gộp mọi số phòng.
    def _clean_room(r):
        s = str(r).strip()
        if s.endswith('.0'): s = s[:-2]
        return s

    df = df.copy()
    df['_conf_ff'] = df['Conf#'].ffill()

    # ── Tách các booking ĐOÀN (có mã Group) để dùng mẫu Registration Card for Group ──
    # KHÔNG ffill cột Group — chỉ dòng có sẵn mã Group mới thuộc đoàn.
    # Các dòng cùng Conf# trong đoàn: điền Group xuống theo từng Conf# nếu dòng đầu có Group.
    if 'Group' in df.columns:
        # Điền mã Group xuống các dòng cùng Conf# (đoàn nhiều phòng, chỉ dòng đầu có Group)
        df['_group_ff'] = df.groupby('_conf_ff')['Group'].transform(
            lambda s: s.ffill().bfill() if s.notna().any() else s)
    else:
        df['_group_ff'] = pd.NA
    has_group = df['_group_ff'].notna()
    df_group = df[has_group].copy()

    writer = PdfWriter()
    count = 0

    tmpl_bytes = load_regcard_template()
    grp_tmpl = None  # nạp lười khi gặp đoàn đầu tiên
    _rendered_groups = set()  # tránh vẽ trùng 1 đoàn khi trải nhiều Conf#

    # Duyệt theo ĐÚNG THỨ TỰ xuất hiện trong file (gộp theo Conf#).
    # Nhóm nào có mã Group → vẽ 1 trang Registration Card for Group (gộp cả đoàn);
    # nhóm thường → regcard thường như cũ.
    for conf_val, grp in df.groupby('_conf_ff', sort=False):
        main_rows = grp[grp['Conf#'].notna()]
        if len(main_rows) == 0:
            continue
        main = main_rows.iloc[0]

        # ── Nếu booking này thuộc ĐOÀN → dùng mẫu group ──
        gid = grp['_group_ff'].dropna().iloc[0] if grp['_group_ff'].notna().any() else None
        if gid is not None:
            if gid in _rendered_groups:
                continue  # đoàn đã vẽ ở lần gặp trước
            _rendered_groups.add(gid)
            if grp_tmpl is None:
                grp_tmpl = load_group_template()
            gdf = df_group[df_group['_group_ff'] == gid]
            page = build_group_regcard(gdf, grp_tmpl)
            writer.add_page(page)
            count += 1
            continue

        # ── Booking thường → regcard thường ──
        rooms = []
        for r in grp['Rm']:
            if pd.notna(r):
                rs = _clean_room(r)
                # Bỏ phòng ảo đầu 9 dạng 9000-9999 (9002, 9005, 9010, 9040... — posting master)
                if rs and rs not in rooms and not _re.fullmatch(r'9\d{3}', rs):
                    rooms.append(rs)
        # Gom mã Specials của cả nhóm (để bắt CN/EB dù nằm ở dòng nào)
        spec_codes = set()
        if 'Specials' in grp.columns:
            for sv in grp['Specials'].dropna():
                for code in str(sv).split(','):
                    code = code.strip().upper()
                    if code:
                        spec_codes.add(code)
        g = {'main': main, 'rooms': rooms, 'specials': spec_codes}
        row = g['main']
        name = _rc_clean_name(row.get('Name'))
        if not name:
            continue
        company = str(row.get('Company')) if pd.notna(row.get('Company')) else ''
        # ── Ô SPECIAL REQUEST ──
        # Mặc định: AI Lunch, AI Dinner, Minibar set up.
        # CELERIS → thêm ( EUR ) sau Lunch & Dinner.
        # Specials có CN → thêm Connecting Room; có EB → thêm Extra Bed.
        if _norm_nat(company).startswith('celeris'):
            sr = 'AI Lunch ( EUR ), AI Dinner ( EUR ), Minibar set up'
        else:
            sr = 'AI Lunch, AI Dinner, Minibar set up'
        _spec = g.get('specials', set())
        if 'CN' in _spec:
            sr += ', Connecting Room'
        if 'EB' in _spec:
            sr += ', Extra Bed'

        data = {
            'name': name,
            'conf': _rc_conf(row.get('Conf#')),
            'arrival': _rc_date(row.get('Arrival')),
            'departure': _rc_date(row.get('Departure')),
            'nights': _rc_nights(row.get('Arrival'), row.get('Departure')),
            'type': str(row.get('Type')) if pd.notna(row.get('Type')) else '',
            'rm': ', '.join(g['rooms']),   # gộp các số phòng
            'company': company,
            'special': sr,
        }
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=(595,841))
        c.setFillColor(white)
        for x0,top,x1,bot in BLANK:
            c.rect(x0, H-bot, (x1-x0), (bot-top), fill=1, stroke=0)
        c.setFillColor(black)
        c.setFont(FONT, SIZE)
        MAXW = {'name': 300, 'company': 200}  # ô 1 dòng: thu nhỏ nếu tràn
        for key,(x,bottom) in POS.items():
            val = data[key]
            if not val:
                continue
            if key == 'rm':
                # Ô số phòng: 1 phòng → vị trí chuẩn như mẫu gốc.
                # Nhiều phòng → căn giữa CHÍNH XÁC giữa 2 nhãn (đo từ 2 ảnh crop độc lập,
                # sai lệch <1pt): nhãn "ROOM NO./(Số phòng)" kết thúc ~261pt,
                # nhãn "ROOM RATE/(Giá phòng)" bắt đầu ~441pt.
                # → Vùng vẽ 266–436pt, TÂM = 351pt, rộng 170pt
                #   (đủ ~8 phòng/dòng ở cỡ chữ nguyên vẹn 9.8pt → 24 phòng/3 dòng).
                rooms = [s.strip() for s in val.split(',') if s.strip()]
                RM_X0, RM_X1 = 266.0, 436.0
                RM_CX = (RM_X0 + RM_X1) / 2        # = 351.0
                RM_MAXW = RM_X1 - RM_X0            # = 170pt mỗi dòng
                def _wrap(items, fs):
                    lines=[]; cur=''
                    for it in items:
                        test = (cur + ', ' + it) if cur else it
                        if c.stringWidth(test, FONT, fs) <= RM_MAXW:
                            cur = test
                        else:
                            if cur: lines.append(cur)
                            cur = it
                    if cur: lines.append(cur)
                    return lines
                if len(rooms) <= 1:
                    c.drawString(x, H - bottom, val)
                else:
                    fs = SIZE
                    lines = _wrap(rooms, fs)
                    if len(lines) <= 3:
                        gap = fs + 1.6             # tới ~24 phòng: cỡ chữ 9.8pt nguyên vẹn
                    else:
                        gap = fs + 0.6             # 4 dòng sát nhau, vẫn nguyên cỡ chữ (~32 phòng)
                        while len(lines) > 4 and fs > 7:
                            fs -= 0.3              # chống tràn tuyệt đối cho case phi thực tế
                            lines = _wrap(rooms, fs)
                    c.setFont(FONT, fs)
                    for i, ln in enumerate(lines):
                        c.drawCentredString(RM_CX, H - bottom - i*gap, ln)
                    c.setFont(FONT, SIZE)
            else:
                maxw = MAXW.get(key)
                if key == 'special':
                    # Ô special: vẽ từ x=137.4, giới hạn mép phải bảng ~573 → rộng ~436pt.
                    # Thu nhỏ nhẹ nếu quá dài (hiếm), giữ nguyên baseline dòng in sẵn.
                    SP_MAXW = 573 - x
                    fs = SIZE
                    while fs > 7 and c.stringWidth(val, FONT, fs) > SP_MAXW:
                        fs -= 0.2
                    c.setFont(FONT, fs)
                    c.drawString(x, H - bottom, val)
                    c.setFont(FONT, SIZE)
                elif maxw:
                    fs = SIZE
                    while fs > 6 and c.stringWidth(val, FONT, fs) > maxw:
                        fs -= 0.3
                    c.setFont(FONT, fs)
                    c.drawString(x, H-bottom, val)
                    c.setFont(FONT, SIZE)
                else:
                    c.drawString(x, H-bottom, val)
        c.save(); buf.seek(0)
        base = PdfReader(io.BytesIO(tmpl_bytes))
        overlay = PdfReader(buf)
        page = base.pages[0]
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
        count += 1
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue(), count

# ── Đối chiếu Smile vs Trang lưu trú người nước ngoài ─────────────────────
def _norm_pp(p):
    """Chuẩn hóa số hộ chiếu."""
    if pd.isna(p): return ''
    s = str(p).strip().upper().replace(' ','')
    if s.endswith('.0'): s = s[:-2]
    return s

def _norm_room(r):
    if pd.isna(r): return ''
    s = str(r).strip()
    if s.endswith('.0'): s = s[:-2]
    return s.upper()  # 12a05 và 12A05 là một phòng

def reconcile(smile_bytes, luutru_bytes, today):
    """Đối chiếu file Smile (inhouse) với file trang quản lý lưu trú.
    today: pd.Timestamp ngày xuất file (hôm nay)."""
    # ── Đọc Smile ──
    df1 = pd.read_excel(io.BytesIO(smile_bytes), header=0)
    smile = df1[['Passport #','NAT','Rm#','Arrival','Last Name','First Name']].copy()
    smile = smile.dropna(subset=['Passport #'])
    smile['pp'] = smile['Passport #'].apply(_norm_pp)
    smile['Arrival'] = pd.to_datetime(smile['Arrival'], errors='coerce')
    smile['name'] = (smile['Last Name'].astype(str).str.strip() + ' ' +
                     smile['First Name'].astype(str).str.strip())
    smile['room'] = smile['Rm#'].apply(_norm_room)
    # Cột Departure (dò tên linh hoạt phòng khi khác tên)
    _dep_col = next((c for c in df1.columns if 'depart' in str(c).lower()), None)
    smile['Departure'] = pd.to_datetime(df1[_dep_col], errors='coerce') if _dep_col else pd.NaT
    # Lọc: bỏ VNM + bỏ arrival = hôm nay + bỏ departure = hôm nay (khách trả phòng)
    smile_f = smile[(smile['NAT'] != 'VNM') &
                    (smile['Arrival'].dt.date != today.date()) &
                    (smile['Departure'].dt.date != today.date())].copy()

    # ── Đọc Lưu trú ──
    df2 = pd.read_excel(io.BytesIO(luutru_bytes), header=9)
    df2 = df2.dropna(subset=['Họ tên'])
    df2['pp'] = df2['Số hộ chiếu'].apply(_norm_pp)
    # Cột ngày đi dự kiến: file mới đổi tên thành "Thời gian dự kiến tạm trú tại CSLT".
    # Dò linh hoạt để chạy được cả file mẫu mới lẫn cũ.
    _ddk_col = next((c for c in df2.columns
                     if 'dự kiến' in str(c).lower() and 'tạm trú' in str(c).lower()), None)
    if _ddk_col is None:
        _ddk_col = next((c for c in df2.columns if 'đi dự kiến' in str(c).lower()), None)
    df2['ddk'] = pd.to_datetime(df2[_ddk_col], format='%d/%m/%Y', errors='coerce') if _ddk_col else pd.NaT
    df2['room'] = df2['Số phòng'].apply(_norm_room)
    # Lọc: bỏ ngày đi dự kiến = hôm nay
    luutru_f = df2[df2['ddk'].dt.date != today.date()].copy()

    smile_pp = set(smile_f['pp'])
    luutru_pp = set(luutru_f['pp'])

    # Đối chiếu người
    chua_dk = smile_f[~smile_f['pp'].isin(luutru_pp)][['name','pp','NAT','room','Arrival']].copy()
    chua_dk.columns = ['Họ tên','Số hộ chiếu','Quốc tịch','Số phòng','Ngày đến']
    chua_dk['Ngày đến'] = chua_dk['Ngày đến'].dt.strftime('%d/%m/%Y')

    thua = luutru_f[~luutru_f['pp'].isin(smile_pp)].copy()
    # Giai đoạn lưu trú: Ngày đến → Ngày đi dự kiến
    _dd = pd.to_datetime(thua['Ngày đến '], format='%d/%m/%Y', errors='coerce') if 'Ngày đến ' in thua else pd.Series([pd.NaT]*len(thua))
    thua['_luutru'] = (_dd.dt.strftime('%d/%m/%Y').fillna('?') + ' → ' +
                       thua['ddk'].dt.strftime('%d/%m/%Y').fillna('?'))
    _qt = thua['QT'] if 'QT' in thua else ''
    thua = pd.DataFrame({
        'Họ tên': thua['Họ tên'].values,
        'Số hộ chiếu': thua['pp'].values,
        'Quốc tịch': _qt.values if hasattr(_qt,'values') else _qt,
        'Số phòng': thua['room'].values,
        'Giai đoạn lưu trú': thua['_luutru'].values,
    })

    dup = luutru_f[luutru_f['pp'].duplicated(keep=False) & (luutru_f['pp']!='')]
    dup = dup[['Họ tên','pp','room']].sort_values('pp').copy()
    dup.columns = ['Họ tên','Số hộ chiếu','Số phòng']

    # Đối chiếu phòng
    smile_rooms = set(r for r in smile_f['room'] if r)
    luutru_rooms = set(r for r in luutru_f['room'] if r)
    def _sortkey(x): return (len(x), x)
    room_chua = sorted(smile_rooms - luutru_rooms, key=_sortkey)
    room_thua = sorted(luutru_rooms - smile_rooms, key=_sortkey)

    return {
        'smile_total': len(smile), 'smile_filtered': len(smile_f),
        'luutru_total': len(df2), 'luutru_filtered': len(luutru_f),
        'chua_dk': chua_dk, 'thua': thua, 'dup': dup,
        'room_chua': room_chua, 'room_thua': room_thua,
        'room_match': len(smile_rooms & luutru_rooms),
        'smile_rooms': len(smile_rooms), 'luutru_rooms': len(luutru_rooms),
    }



# ── UI ────────────────────────────────────────────────────────────────────

# Giao diện Neumorphism/soft UI — đơn sắc sáng, đổ bóng nổi-chìm
if not st.session_state.get("_app_scripts_injected"):
    st.session_state["_app_scripts_injected"] = True
    # Toàn bộ lớp giao diện Neumorphism (CSS + hoa anh đào nền + hiệu ứng chữ
    # + màn khởi động) gộp trong 1 lệnh st.iframe duy nhất, tiêm đúng 1 lần
    # mỗi phiên — mọi animation chỉ dùng transform/opacity (chạy trên GPU
    # compositor), không backdrop-filter, không chạy lại khi rerun.
    _boot_script = """
<script>
(function(){
    var doc = window.parent.document;
    // Đặt theme NGAY từ khối này để màn chào không bị chớp sáng khi đang ở
    // chế độ tối (khối gán data-theme mỗi lần rerun chạy sau khối này)
    doc.documentElement.setAttribute('data-theme', '__GREET_THEME__');
    if (doc.getElementById('main-app-style')) return;
    var css = doc.createElement('style');
    css.id = 'main-app-style';
    css.textContent = `
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header[data-testid="stHeader"] {background: transparent;}

    /* ── Hệ màu (light mặc định) — khai báo ở html để lớp popover/dropdown
       render NGOÀI .stApp cũng kế thừa được, và khai báo lại ở .stApp vì
       khai báo trực tiếp trên phần tử luôn thắng giá trị kế thừa. ── */
    html, .stApp {
        --ease: cubic-bezier(0.4, 0, 0.2, 1);
        --bg: #f6f7fa; --surf: #ffffff; --surf2: #fafbfc;
        --line: #e7e9ef; --line2: #f0f2f6; --chip: #f2f4f8;
        --tx: #0f172a; --tx2: #5b6478; --tx3: #98a1b3;
        --acc: #4f46e5; --acc2: #7c3aed;
        --ok: #0d9668; --warn: #d97706; --err: #dc2626;
        --sh: 0 1px 2px rgba(15,23,42,.05), 0 8px 24px rgba(15,23,42,.05);
        --sh-lg: 0 2px 6px rgba(15,23,42,.06), 0 14px 36px rgba(15,23,42,.09);
        --r-lg: 16px; --r-md: 11px; --r-sm: 9px; --r-pill: 999px;
    }
    html[data-theme="dark"], html[data-theme="dark"] .stApp {
        --bg: #12151d; --surf: #1a1f2b; --surf2: #161b25;
        --line: #262d3b; --line2: #212734; --chip: #242b39;
        --tx: #eef2fa; --tx2: #a3adc2; --tx3: #727d95;
        --acc: #8b93f8; --acc2: #a78bfa;
        --ok: #34d399; --warn: #fbbf24; --err: #f87171;
        --sh: 0 1px 2px rgba(0,0,0,.3), 0 10px 30px rgba(0,0,0,.32);
        --sh-lg: 0 2px 8px rgba(0,0,0,.35), 0 18px 44px rgba(0,0,0,.42);
    }
    .stApp {
        background-color: var(--bg);
        color: var(--tx);
        background-image: linear-gradient(rgba(246,247,250,.80), rgba(246,247,250,.93)),
                           url("__LIGHT_BG_DATA_URI__");
        background-size: cover; background-position: center center;
        background-repeat: no-repeat; background-attachment: fixed;
    }
    html[data-theme="dark"] .stApp {
        background-image: linear-gradient(rgba(18,21,29,.78), rgba(18,21,29,.92)),
                           url("__DARK_BG_DATA_URI__");
    }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {
        background: var(--surf); border-right: 1px solid var(--line); box-shadow: none;
    }
    section[data-testid="stSidebar"] .stButton button {
        background: transparent; border: 1px solid transparent;
        border-radius: var(--r-sm);
        color: var(--tx2); font-weight: 560; text-align: left;
        justify-content: flex-start; padding: 0.5rem 0.6rem;
        box-shadow: none; transition: background 0.12s var(--ease), color 0.12s var(--ease);
    }
    section[data-testid="stSidebar"] .stButton button:hover {
        background: var(--chip); color: var(--tx);
    }
    section[data-testid="stSidebar"] .stButton button[kind="primary"] {
        background: linear-gradient(135deg, rgba(99,91,255,.12), rgba(167,139,250,.10)) !important;
        color: var(--tx) !important; font-weight: 700;
        border-color: rgba(120,110,250,.28) !important; box-shadow: none !important;
    }
    /* Icon mỗi mục nav nằm trong ô chip; mục đang mở thì chip đổi sang gradient */
    section[data-testid="stSidebar"] .stButton button [data-testid="stIconMaterial"] {
        background: var(--chip); border-radius: 7px; padding: 4px;
        width: 23px; height: 23px; display: inline-flex; align-items: center;
        justify-content: center; font-size: 15px !important; margin-right: 2px;
    }
    section[data-testid="stSidebar"] .stButton button[kind="primary"] [data-testid="stIconMaterial"] {
        background: linear-gradient(135deg, var(--acc), var(--acc2)); color: #fff;
    }
    /* Badge số bên phải mục nav (số điền động ở khối style tiêm mỗi lần rerun) */
    section[data-testid="stSidebar"] .stButton button::after {
        margin-left: auto; font-size: 0.62rem; font-weight: 800; color: var(--tx2);
        background: var(--chip); padding: 1px 7px; border-radius: var(--r-pill);
    }
    /* Khung nền cho cụm nút gạt giao diện (hàng cột duy nhất trong sidebar) */
    section[data-testid="stSidebar"] div[data-testid="stHorizontalBlock"] {
        background: var(--chip); border-radius: 10px; padding: 3px; gap: 3px !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stHorizontalBlock"] .stButton button {
        justify-content: center; text-align: center; padding: 0.36rem 0.2rem;
        font-size: 0.72rem; font-weight: 680; border: none;
    }
    section[data-testid="stSidebar"] div[data-testid="stHorizontalBlock"] .stButton button[kind="primary"] {
        background: var(--surf) !important; color: var(--tx) !important;
        border-color: transparent !important; box-shadow: 0 1px 3px rgba(15,23,42,.16) !important;
    }
    .sb-brand {padding: 0.1rem 0.3rem 0.9rem; border-bottom: 1px solid var(--line2); margin-bottom: 0.5rem;}
    .sb-mascot {background: linear-gradient(140deg,#ff8fb1,#c084fc); border-radius: 9px;
        padding: 4px; box-shadow: 0 4px 12px rgba(200,120,220,.35);}
    .sb-brand-title {color: var(--tx); font-weight: 730; font-size: 0.92rem; letter-spacing: -.015em;}
    .sb-mascot {display:inline-block; width:22px; height:22px; vertical-align:-5px; margin-right:5px;
        animation: flowerSway 3.4s ease-in-out infinite;}
    .sb-mascot svg {width:100%; height:100%;}
    .sb-brand-sub {color: var(--tx3); font-size: 0.7rem; margin-top: 1px;}
    .sb-section {color: var(--tx3); font-size: 0.6rem; font-weight: 800;
        text-transform: uppercase; letter-spacing: 0.1em; padding: 0.8rem 0.55rem 0.3rem;}
    .sb-user {display:flex; align-items:center; gap:9px; padding-top:0.8rem;
        margin-top:0.5rem; border-top: 1px solid var(--line2);}
    .sb-av {width:28px; height:28px; border-radius:50%; flex-shrink:0; color:#fff;
        background: linear-gradient(135deg,#818cf8,#c084fc);
        display:flex; align-items:center; justify-content:center; font-size:0.72rem; font-weight:760;}
    .sb-un {font-size:0.76rem; font-weight:680; color: var(--tx);}
    .sb-ur {font-size:0.66rem; color: var(--tx3);}
    .sb-status {display:flex; align-items:center; gap:7px; color: var(--tx3);
        font-size:0.68rem; padding-top:0.55rem;}
    .sb-dot {position:relative; width:7px; height:7px; border-radius:50%; background:#34d399; flex-shrink:0;}
    .sb-dot::after {content:""; position:absolute; inset:0; border-radius:50%; background:#34d399;
        animation: pulse 2.2s var(--ease) infinite;}
    @keyframes pulse {
        0%   {transform: scale(1);   opacity: 0.5;}
        70%  {transform: scale(2.6); opacity: 0;}
        100% {transform: scale(2.6); opacity: 0;}
    }
    @keyframes flowerSway {
        0%, 100% {transform: rotate(-8deg);}
        50%      {transform: rotate(8deg);}
    }

    /* ── Nội dung hiện dần so le sau khi màn chào tan ── */
    @keyframes tanReveal {
        from {opacity: 0; transform: translateY(10px);}
        to   {opacity: 1; transform: none;}
    }
    .tan-rv {opacity: 0; animation: tanReveal 0.42s var(--ease) forwards;}

    /* ── Chuyển màn mượt khi đổi công cụ (sidebar) — TRƯỚC đây đổi công cụ là
       cắt cứng, không có hiệu ứng gì, khác hẳn màn chào mượt mà lúc mở web.
       Lớp .tan-page-in được JS thêm vào mỗi khi menu thực sự đổi (xem khối
       script ở cuối app.py) — chỉ 1 lần mờ+trượt nhẹ, không so le từng khối
       như màn chào vì lặp lại mỗi cú click sẽ thấy chậm/rườm. */
    @keyframes tanPageIn {
        from {opacity: 0; transform: translateY(7px);}
        to   {opacity: 1; transform: none;}
    }
    .tan-page-in {animation: tanPageIn 0.26s var(--ease);}
    @media (prefers-reduced-motion: reduce) {
        .tan-page-in {animation: none !important;}
    }

    /* ── Top bar: breadcrumb + trạng thái ── */
    .tan-topbar {
        display: flex; align-items: center; gap: 8px; flex-wrap: wrap;
        font-size: 0.82rem; color: var(--tx); padding: 0.15rem 0 0.15rem;
    }
    /* Thanh top bar: nền liền khối + viền dưới, tràn ra sát 2 mép khung nội dung */
    /* Khung nội dung có padding 96px trên / 80px hai bên — kéo âm đúng bằng đó
       để thanh chạy sát mép, cộng width bù lại phần margin âm hai bên. */
    .st-key-tan_topbar {
        background: var(--surf); border-bottom: 1px solid var(--line);
        margin: -4.6rem -1.6rem 1.1rem !important; width: calc(100% + 3.2rem) !important;
        max-width: none !important; padding: 0.5rem 1.6rem !important;
    }
    /* Nút "Xuất báo cáo" trên thanh là nút phụ (nút chính là Xử lý hàng ngày) */
    .st-key-tb_report_dl button, .st-key-tb_report_off button {
        background: var(--surf) !important; color: var(--tx) !important;
        border: 1px solid var(--line) !important; box-shadow: none !important;
        font-weight: 640;
    }
    /* Thanh progress mảnh trong thẻ số liệu */
    .kpi-bar {height: 3px; border-radius: 3px; background: var(--chip); margin-top: .7rem; overflow: hidden;}
    .kpi-bar i {display: block; height: 100%; border-radius: 3px; background: var(--acc);}
    .kpi-bar i.err {background: var(--err);}
    .tan-crumb {color: var(--tx3);}
    .tan-sep {color: var(--tx3); opacity: .6;}
    .tan-topbar-r {margin-left: auto; display: flex; align-items: center; gap: 7px;}
    .tan-h1 {font-size: 1.35rem; font-weight: 790; letter-spacing: -.03em; color: var(--tx);}
    .tan-h1sub {font-size: 0.78rem; color: var(--tx3); margin: 2px 0 0.7rem;}

    /* ── Thẻ số liệu nhỏ (bento) ── */
    .kpi-lab {font-size: 0.76rem; color: var(--tx2); font-weight: 640;
        display: flex; align-items: center; gap: 8px;}
    .kpi-ic {width: 26px; height: 26px; border-radius: 8px; display: inline-flex;
        align-items: center; justify-content: center; font-size: 0.78rem; background: var(--chip);}
    .kpi-ic.err {background: rgba(239,68,68,.14);}
    .kpi-ic.ok {background: rgba(16,185,129,.14);}
    .kpi-ic.acc {background: rgba(99,91,240,.14);}
    .kpi-val {font-size: 1.85rem; font-weight: 800; letter-spacing: -.035em;
        line-height: 1; margin-top: 1.5rem; color: var(--tx); font-variant-numeric: tabular-nums;}
    .kpi-val.err {color: var(--err);}
    .kpi-sub {font-size: 0.7rem; color: var(--tx3); font-weight: 620; margin-top: .45rem;}

    /* ── Panel (thẻ có tiêu đề + danh sách dòng) ── */
    .pan-h {display: flex; align-items: center; gap: 8px; padding-bottom: .55rem;
        margin-bottom: .5rem; border-bottom: 1px solid var(--line2);}
    .pan-t {font-size: 0.85rem; font-weight: 730; letter-spacing: -.015em; color: var(--tx);}
    .pan-row {display: flex; align-items: center; gap: 8px; flex-wrap: wrap; font-size: 0.82rem;}
    .pan-task {display: flex; align-items: center; gap: 9px; padding: .62rem .1rem;
        font-size: 0.82rem; border-bottom: 1px solid var(--line2);}
    .pan-task:last-of-type {border-bottom: 0;}
    .pan-time {margin-left: auto; font-size: 0.7rem; color: var(--tx3); font-weight: 650;
        font-variant-numeric: tabular-nums;}
    .pan-rt {color: var(--tx); font-weight: 560;}
    .pan-mut {font-size: 0.7rem; color: var(--tx3); font-weight: 620;}
    .pan-line {display: flex; align-items: center; justify-content: space-between;
        gap: 10px; padding: .48rem 0; font-size: 0.8rem; color: var(--tx2);
        border-bottom: 1px solid var(--line2);}
    .pan-line:last-child {border-bottom: 0;}
    .pan-v {font-weight: 750; color: var(--tx); font-variant-numeric: tabular-nums;}
    .pan-empty {font-size: 0.78rem; color: var(--tx3); line-height: 1.55; padding: .3rem 0 .7rem;}
    .pan-foot {margin-top: auto; padding-top: .35rem;}
    /* Hộp cảnh báo trong panel (đúng kiểu ô nhắc việc của bản dựng) */
    .pan-al {display: flex; gap: 9px; padding: .6rem .65rem; border-radius: 11px;
        margin-bottom: .45rem; border: 1px solid;}
    .pan-al.r {background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.26);}
    .pan-al.a {background: rgba(245,158,11,.10); border-color: rgba(245,158,11,.26);}
    .pan-al-ic {font-size: 0.85rem; line-height: 1.2;}
    .pan-al-t {font-size: 0.78rem; font-weight: 680; color: var(--tx); line-height: 1.35;}
    .pan-al-m {font-size: 0.68rem; color: var(--tx3); margin-top: 2px; line-height: 1.35;}

    /* ── Nhãn mục ── */
    .section-label {
        font-size: 0.63rem; font-weight: 800; color: var(--tx3);
        text-transform: uppercase; letter-spacing: 0.09em; margin-bottom: 0.6rem;
    }

    /* ── Thẻ nội dung (st.container(border=True)) ──
       Bản Streamlit này KHÔNG còn testid "stVerticalBlockBorderWrapper";
       khối có viền là stVerticalBlock mang thêm thuộc tính overflow (các khối
       thường không có). Giữ cả selector cũ cho bản Streamlit đời trước. */
    div[data-testid="stVerticalBlock"][overflow],
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--surf); border: 1px solid var(--line) !important;
        border-radius: 14px !important; box-shadow: var(--sh);
        transition: box-shadow 0.15s var(--ease);
    }
    div[data-testid="stVerticalBlock"][overflow]:hover,
    div[data-testid="stVerticalBlockBorderWrapper"]:hover {box-shadow: var(--sh-lg);}

    /* ── Thẻ HERO (ảnh mèo làm nền, phủ tối dần để chữ luôn nổi) ── */
    .tan-hero {
        position: relative; overflow: hidden; color: #fff;
        border-radius: var(--r-lg); padding: 1.15rem 1.35rem;
        min-height: 200px; display: flex; flex-direction: column; justify-content: center;
        background-image: linear-gradient(100deg, rgba(22,18,48,.95) 0%, rgba(34,26,66,.74) 46%, rgba(48,36,84,.26) 100%),
                           url("__LIGHT_BG_DATA_URI__");
        background-size: cover, cover; background-position: center, center right;
        box-shadow: 0 12px 32px rgba(30,25,70,.28);
    }
    html[data-theme="dark"] .tan-hero {
        background-image: linear-gradient(100deg, rgba(16,14,32,.95) 0%, rgba(26,22,50,.74) 46%, rgba(40,32,72,.24) 100%),
                           url("__DARK_BG_DATA_URI__");
        box-shadow: 0 14px 36px rgba(0,0,0,.45);
    }
    .tan-hero-lab {font-size: 0.74rem; font-weight: 640; opacity: .88;}
    .tan-hero-val {font-size: 2.55rem; font-weight: 820; letter-spacing: -.04em; line-height: 1; margin-top: .35rem;}
    .tan-hero-sub {font-size: 0.75rem; opacity: .85; font-weight: 600; margin-top: .45rem;}
    .tan-hero-split {display: flex; gap: 1.6rem; margin-top: .9rem; padding-top: .75rem;
        border-top: 1px solid rgba(255,255,255,.22); flex-wrap: wrap;}
    .tan-hero-k {font-size: 0.68rem; opacity: .82; font-weight: 620;}
    .tan-hero-v {font-size: 1.1rem; font-weight: 790; margin-top: 1px; letter-spacing: -.02em;}

    /* ── Chip trạng thái ── */
    .tan-chip {display:inline-block; font-size:0.64rem; font-weight:740; padding:2px 8px;
        border-radius: var(--r-pill); background: var(--chip); color: var(--tx2);}
    .tan-chip.ok {background: rgba(16,185,129,.14); color: var(--ok);}
    .tan-chip.warn {background: rgba(245,158,11,.15); color: var(--warn);}
    .tan-chip.err {background: rgba(239,68,68,.13); color: var(--err);}

    /* ── Ô nhập ── */
    [data-testid="stTextInputRootElement"],
    [data-testid="stTextAreaRootElement"],
    [data-testid="stNumberInputContainer"] {
        background: var(--surf2) !important; border-radius: var(--r-md) !important;
    }
    .stTextInput input, .stNumberInput input, .stTextArea textarea {
        background: var(--surf2) !important; border: 1px solid var(--line) !important;
        border-radius: var(--r-md) !important; color: var(--tx) !important;
        box-shadow: none !important; transition: border-color 0.12s var(--ease), box-shadow 0.12s var(--ease);
    }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        border-color: var(--acc) !important;
        box-shadow: 0 0 0 3px rgba(99,91,240,.16) !important;
    }
    div[data-testid="stNumberInput"] button {border: none; background: transparent; color: var(--tx2) !important;}
    [data-testid="stNumberInputStepDown"], [data-testid="stNumberInputStepUp"],
    button[aria-label="Open"], button[aria-label="Show password"], button[aria-label="Hide password"] {
        color: var(--tx2) !important;
    }
    div[data-testid="stSelectbox"] [role="group"] {
        background: var(--surf2) !important; border-radius: var(--r-md) !important;
    }
    div[data-testid="stSelectbox"] input[role="combobox"] {color: var(--tx) !important;}
    div:has(> [role="listbox"]) {background: var(--surf) !important;}
    [role="option"] {color: var(--tx) !important;}
    [role="option"][aria-selected="true"], [role="option"]:hover {background: var(--chip) !important;}
    [data-testid="stWidgetLabel"] p {color: var(--tx2) !important; font-weight: 600; font-size: 0.8rem;}

    /* ── st.date_input: khung ô nhập của BaseWeb hardcode nền sáng
       (rgb(237,240,245)), không kế thừa theme — vẫn trắng dù đã bật tối. ── */
    [data-testid="stDateInput"] [data-baseweb="input"] {
        background: var(--surf2) !important; border: 1px solid var(--line) !important;
        border-radius: var(--r-md) !important;
    }
    [data-testid="stDateInputField"] {
        background: transparent !important; color: var(--tx) !important;
        caret-color: var(--tx) !important;
    }
    [data-testid="stDateInputField"]::placeholder {color: var(--tx3) !important;}
    /* Lịch chọn ngày hiện trong portal NGOÀI .stApp — không kế thừa biến màu
       của .stApp nên khai báo riêng ở đây (giống khối html[data-theme] khác).
       Khung popover bọc ngoài lịch cũng hardcode nền sáng — sửa luôn. */
    [data-baseweb="popover"] {background: var(--surf) !important;}
    div[data-baseweb="calendar"] {background: var(--surf) !important; color: var(--tx) !important;}
    /* Thanh "‹ Tháng Năm ›" và hàng tên thứ (Su Mo Tu...) là 2 div thường,
       không có data-baseweb riêng — cũng bị hardcode nền sáng như trên. */
    div[data-baseweb="calendar"] div:has(> button[aria-label="Previous month."]),
    div[data-baseweb="calendar"] div[role="presentation"] {
        background: var(--surf) !important; color: var(--tx) !important;
    }
    div[data-baseweb="calendar"] [role="presentation"] {color: var(--tx2) !important;}
    div[data-baseweb="calendar"] button {color: var(--tx) !important; background: transparent !important;}
    div[data-baseweb="calendar"] button[aria-disabled="true"] {color: var(--tx3) !important;}
    /* Ô trống đầu/cuối lưới (ngày thuộc tháng trước/sau) là gridcell RỖNG
       (không có div con chứa số) — tự vẽ nền bằng ::after màu sáng hardcode.
       Chỉ nhắm đúng ô rỗng (:empty) — ô có số (kể cả ô đang chọn, cũng vẽ
       vòng tròn tô màu bằng ::after) phải giữ nguyên, không bị vạ lây. */
    div[data-baseweb="calendar"] [role="gridcell"]:empty::after {background: transparent !important;}

    /* ── st.code(): nền pre/code hardcode sáng, không kế thừa theme ── */
    [data-testid="stCode"] pre {background: var(--surf2) !important; border: 1px solid var(--line) !important;}
    [data-testid="stCode"] code {color: var(--tx) !important;}
    /* Nút "Copy to clipboard" nổi góc trên code block: khung bọc nút cũng
       hardcode nền sáng (nút tự thân trong suốt nên trước đó không thấy). */
    div:has(> [data-testid="stElementToolbarButton"]) {background: var(--surf2) !important;}
    [data-testid="stElementToolbarButton"] svg {color: var(--tx2) !important;}
    div[data-baseweb="calendar"] div[aria-selected="true"] button {
        background: var(--acc) !important; color: #fff !important;
    }

    /* ── Tải file ── */
    div[data-testid="stFileUploader"] {
        background: var(--surf2); border: 1px dashed var(--line); border-radius: var(--r-md);
        padding: 0.5rem; box-shadow: none;
        transition: border-color 0.12s var(--ease);
    }
    div[data-testid="stFileUploader"]:hover {border-color: var(--acc);}
    div[data-testid="stFileUploader"] section {background: transparent; border: none;}
    div[data-testid="stFileUploader"] button {
        background: var(--surf) !important; color: var(--tx) !important;
        border: 1px solid var(--line) !important; box-shadow: none !important;
        border-radius: var(--r-sm) !important; font-weight: 640;
    }

    /* ── Nút ── */
    div[data-testid="stMainBlockContainer"] .stButton button, .stDownloadButton button {
        border-radius: var(--r-md); font-weight: 660; border: 1px solid var(--line);
        background: var(--surf); color: var(--tx);
        box-shadow: none;
        transition: background 0.12s var(--ease), border-color 0.12s var(--ease), transform 0.1s var(--ease);
    }
    div[data-testid="stMainBlockContainer"] .stButton button:hover, .stDownloadButton button:hover {
        background: var(--chip); border-color: var(--line);
    }
    div[data-testid="stMainBlockContainer"] .stButton button:active, .stDownloadButton button:active {
        transform: scale(0.985);
    }
    div[data-testid="stMainBlockContainer"] .stButton button[kind="primary"], .stDownloadButton button {
        background: linear-gradient(135deg, var(--acc), var(--acc2));
        color: #ffffff; border: none;
        box-shadow: 0 6px 18px rgba(99,91,240,.32);
    }
    div[data-testid="stMainBlockContainer"] .stButton button[kind="primary"]:hover, .stDownloadButton button:hover {
        filter: brightness(1.06); background: linear-gradient(135deg, var(--acc), var(--acc2));
    }

    /* ── Metric ── */
    div[data-testid="stMetric"] {
        background: var(--surf); border: 1px solid var(--line); border-radius: 13px;
        padding: 0.85rem 0.95rem 0.75rem; box-shadow: var(--sh);
        transition: box-shadow 0.15s var(--ease);
    }
    div[data-testid="stMetric"]:hover {box-shadow: var(--sh-lg);}
    div[data-testid="stMetricValue"], div[data-testid="stMetricValue"] * {
        font-weight: 780 !important; color: var(--tx) !important; letter-spacing: -.03em;
        font-variant-numeric: tabular-nums;
    }
    div[data-testid="stMetricLabel"], div[data-testid="stMetricLabel"] * {
        font-size: 0.76rem !important; color: var(--tx2) !important; font-weight: 620 !important;
    }

    /* ── Cảnh báo ── */
    div[data-testid="stAlert"] {border-radius: var(--r-md); border: 1px solid transparent;}

    /* ── Form / expander / bảng ── */
    div[data-testid="stForm"] {
        background: var(--surf); border: 1px solid var(--line);
        border-radius: 14px; padding: 1rem 1.1rem; box-shadow: var(--sh);
    }
    div[data-testid="stExpander"] {background: var(--surf); border-radius: var(--r-md);}
    /* Streamlit tô nền xám sáng cho summary khi expander ĐANG MỞ — ở chế độ tối
       thành chữ trắng trên nền sáng, không đọc được. Ép trong suốt cả 2 trạng thái. */
    div[data-testid="stExpander"] summary,
    div[data-testid="stExpander"] details[open] > summary {
        color: var(--tx) !important; background: transparent !important;
    }
    div[data-testid="stDataFrame"] {
        border-radius: var(--r-md); overflow: hidden; border: 1px solid var(--line);
    }

    /* ── Chữ nền tối: ép màu cho phần Streamlit tự đặt màu tĩnh ── */
    html[data-theme="dark"] .stApp p, html[data-theme="dark"] .stApp span,
    html[data-theme="dark"] .stApp label, html[data-theme="dark"] .stApp li,
    html[data-theme="dark"] .stApp h1, html[data-theme="dark"] .stApp h2,
    html[data-theme="dark"] .stApp h3, html[data-theme="dark"] .stApp h4,
    html[data-theme="dark"] [data-testid="stMarkdownContainer"] {color: var(--tx);}
    html[data-theme="dark"] [data-testid="stCaptionContainer"],
    html[data-theme="dark"] [data-testid="stCaptionContainer"] * {color: var(--tx3) !important;}
    html[data-theme="dark"] [data-testid="stCheckbox"] label,
    html[data-theme="dark"] [data-testid="stRadio"] label {color: var(--tx) !important;}
    /* Thẻ hero luôn nền tối nên chữ bên trong luôn trắng ở cả 2 chế độ */
    .tan-hero, .tan-hero * {color: #fff !important;}

    /* ── Bố cục ── */
    div[data-testid="stMainBlockContainer"] {max-width: 1320px; padding: 4.6rem 1.6rem 2.5rem;}
    section[data-testid="stSidebar"] {width: 252px !important; min-width: 252px !important;}
    /* Khe giữa các thẻ = 13px như bản dựng (trừ cụm nút giao diện ở sidebar) */
    div[data-testid="stMain"] div[data-testid="stHorizontalBlock"] {gap: 0.82rem;}
    /* Padding trong thẻ có viền: gọn lại cho khớp bản dựng */
    div[data-testid="stVerticalBlock"][overflow] {
        padding: 0.85rem 1rem !important; display: flex; flex-direction: column;
    }
    /* Link "Xem tất cả" ở đầu panel: nút nhưng nhìn như link, đúng bản dựng */
    .st-key-dash_seeall button {
        background: transparent !important; border: none !important; box-shadow: none !important;
        color: var(--acc) !important; font-size: 0.72rem !important; font-weight: 700 !important;
        padding: 0 !important; justify-content: flex-end !important; min-height: 0 !important;
    }
    .st-key-dash_seeall button:hover {background: transparent !important; text-decoration: underline;}
    /* Phần tử cuối trong 2 panel này ghim xuống đáy thẻ như bản dựng */
    .st-key-dash_tasks > div[data-testid="stElementContainer"]:last-child,
    .st-key-dash_ho > div[data-testid="stElementContainer"]:last-child {margin-top: auto;}
    div[data-testid="stElementContainer"]:has(iframe[height="1"]) {display: none;}
    div[data-testid="stMainBlockContainer"] hr {border-color: var(--line);}

    .stButton button:focus-visible, .stDownloadButton button:focus-visible {
        outline: 3px solid rgba(99,91,240,.4); outline-offset: 2px;
    }

    @media (max-width: 768px) {
        div[data-testid="stMainBlockContainer"] {padding-left: 1rem; padding-right: 1rem;}
        .st-key-tan_topbar {margin: -2rem -1rem 0.9rem; width: calc(100% + 2rem); padding: 0.45rem 1rem;}
        .tan-hero-val {font-size: 2rem;}
        .tan-hero-split {gap: 1.1rem;}
        div[data-testid="stMetric"] {padding: 0.6rem 0.7rem 0.55rem;}
    }

    @media (prefers-reduced-motion: reduce) {
        .sb-mascot, .sb-dot::after, #bg-sakura-layer .petal,
        #boot-splash .bs-pt, #boot-splash .bs-bg, #boot-splash .bs-moon,
        #boot-splash .bs-blink {
            animation: none !important;
        }
        .tan-rv, #boot-splash .bs-card, #boot-splash .bs-corner,
        #boot-splash .bs-strip, #boot-splash .bs-enter, #boot-splash .bs-hint2 {
            animation: none !important; opacity: 1 !important; transform: none !important;
        }
        #boot-splash .bs-card { transform: translateY(-50%) !important; }
        .stApp *, #boot-splash {transition-duration: 0.01ms !important;}
    }
    `;
    doc.head.appendChild(css);

    // ── Hoa anh đào rơi liên tục ở nền — thuần transform/opacity, vô hại hiệu năng ──
    if (!doc.getElementById('bg-sakura-layer')) {
        var css2 = doc.createElement('style');
        css2.id = 'bg-sakura-style';
        css2.textContent = `
          #bg-sakura-layer { position: fixed; inset: 0; pointer-events: none; overflow: hidden; z-index: 0; }
          #bg-sakura-layer .petal {
              position: absolute; top: -20px; opacity: 0.55; will-change: transform;
              animation-name: bgSakuraFall; animation-timing-function: linear; animation-iteration-count: infinite;
          }
          @keyframes bgSakuraFall {
              0%   { transform: translate(0,0) rotate(0deg); }
              100% { transform: translate(var(--drift), 112vh) rotate(360deg); }
          }
        `;
        doc.head.appendChild(css2);
        var layer = doc.createElement('div');
        layer.id = 'bg-sakura-layer';
        var colors = ['#f6a8c9', '#f293bc', '#f9c1d9'];
        for (var i = 0; i < 10; i++) {
            var p = doc.createElement('div');
            p.className = 'petal';
            var size = 8 + Math.random()*6;
            var dur = 11 + Math.random()*8;
            p.style.left = (Math.random()*100) + 'vw';
            p.style.width = size + 'px';
            p.style.height = size + 'px';
            p.style.borderRadius = '0 60% 0 60%';
            p.style.background = 'radial-gradient(circle at 30% 30%, #fff, ' + colors[i % 3] + ' 70%)';
            p.style.setProperty('--drift', (Math.random()*160 - 80) + 'px');
            p.style.animationDuration = dur + 's';
            p.style.animationDelay = (-Math.random()*dur) + 's';
            layer.appendChild(p);
        }
        doc.body.insertBefore(layer, doc.body.firstChild);
    }

    // ── Màn chào ca trực (bản phối B+C) ──────────────────────────────────
    // Ảnh nền tràn màn hình đổi theo MÙA và DỊP LỄ, ám sắc trời đổi theo GIỜ,
    // hạt rơi theo mùa; lời chào nằm trong thẻ kính bên trái, đồng hồ góc phải,
    // thanh số liệu ở đáy. Sau 2 giây thì DỪNG chờ người dùng bấm Enter (hoặc
    // chạm màn hình) mới vào app — giống màn khoá Windows. Chỉ chào lần đầu
    // mỗi phiên tab; refresh trong cùng tab không phải chào lại.
    var _bsSeen = false;
    try { _bsSeen = window.parent.sessionStorage.getItem('tanBootSplashSeen') === '1'; } catch (e) {}
    if (!_bsSeen && !doc.getElementById('boot-splash')) {
        try { window.parent.sessionStorage.setItem('tanBootSplashSeen', '1'); } catch (e) {}
        var css3 = doc.createElement('style');
        css3.id = 'boot-splash-style';
        css3.textContent = `
          #boot-splash {
            position: fixed; inset: 0; z-index: 999999; overflow: hidden;
            background: #07090f; cursor: pointer; user-select: none;
            transition: opacity 0.5s ease;
            font-family: -apple-system, "Segoe UI", Roboto, "Helvetica Neue", sans-serif;
          }
          #boot-splash.bs-hide { opacity: 0; pointer-events: none; }

          /* Lớp 1 — ảnh nền theo mùa, phóng rất chậm cho đỡ tĩnh */
          #boot-splash .bs-bg {
            position: absolute; inset: 0; background-size: cover; background-position: center 55%;
            background-image: url('__GREET_PHOTO__');
            animation: bsKen 34s ease-in-out infinite alternate;
          }
          @keyframes bsKen {
            from { transform: scale(1.02) translate(0.6%, 0); }
            to   { transform: scale(1.09) translate(-1.2%, -1.2%); }
          }
          /* Lớp 2 — ám sắc trời theo giờ (soft-light nên ảnh giữ nguyên chi tiết) */
          #boot-splash .bs-hour { position: absolute; inset: 0; mix-blend-mode: soft-light; }
          #boot-splash[data-h="dawn"]  .bs-hour { background: linear-gradient(180deg,#2a3d6b,#ff9d5c 68%,#ffd9a8); }
          #boot-splash[data-h="day"]   .bs-hour { background: linear-gradient(180deg,#4e97dc,#cfe6f7); }
          #boot-splash[data-h="dusk"]  .bs-hour { background: linear-gradient(180deg,#3a1c58,#e0724a 62%,#ffb066); }
          #boot-splash[data-h="night"] .bs-hour { background: linear-gradient(180deg,#070c1e,#1c2352 68%,#2c2f60); }
          /* Lớp 3 — ám sắc mùa */
          #boot-splash .bs-season { position: absolute; inset: 0; mix-blend-mode: overlay; opacity: .5; }
          #boot-splash[data-s="spring"]   .bs-season { background: linear-gradient(120deg,#ffb3d9,#c9a8ff); }
          #boot-splash[data-s="summer"]   .bs-season { background: linear-gradient(120deg,#7fd4ff,#b9f0e0); }
          #boot-splash[data-s="autumn"]   .bs-season { background: linear-gradient(120deg,#e09a4a,#c96a2b); }
          #boot-splash[data-s="winter"]   .bs-season { background: linear-gradient(120deg,#8fb8e0,#cfe2f5); }
          #boot-splash[data-s="tet"]      .bs-season { background: linear-gradient(120deg,#e03a3a,#ffc93a); opacity: .34; }
          #boot-splash[data-s="trungthu"] .bs-season { background: linear-gradient(120deg,#ffb347,#ff7a45); opacity: .40; }
          #boot-splash[data-s="noel"]     .bs-season { background: linear-gradient(120deg,#4a8fd4,#e05a5a); opacity: .42; }
          /* Lớp 4 — scrim giữ chữ luôn đọc được trên mọi ảnh */
          #boot-splash .bs-scrim {
            position: absolute; inset: 0;
            background: linear-gradient(96deg,rgba(6,8,14,.76) 0%,rgba(6,8,14,.40) 40%,rgba(6,8,14,.06) 66%,rgba(6,8,14,.34) 100%),
                        linear-gradient(0deg,rgba(6,8,14,.88) 0%,rgba(6,8,14,.18) 34%,transparent 52%);
          }
          /* Lớp 5 — trăng rằm, chỉ dịp Trung Thu */
          #boot-splash .bs-moon {
            position: absolute; z-index: 3; right: 31%; top: 15%; display: none;
            width: 132px; height: 132px; border-radius: 50%;
            background: radial-gradient(circle at 60% 36%, #fffdf0 58%, #f2e2b8);
            animation: bsMoon 6s ease-in-out infinite;
          }
          #boot-splash[data-s="trungthu"] .bs-moon { display: block; }
          @keyframes bsMoon {
            0%,100% { box-shadow: 0 0 110px 38px rgba(255,225,150,.36); }
            50%     { box-shadow: 0 0 140px 54px rgba(255,225,150,.52); }
          }
          /* Lớp 6 — hạt rơi theo mùa */
          #boot-splash .bs-pt { position: absolute; z-index: 2; will-change: transform; pointer-events: none; }
          @keyframes bsFall { to { transform: translate(var(--dx), 118vh) rotate(700deg); } }
          @keyframes bsRise { to { transform: translate(var(--dx), -118vh) rotate(24deg); } }

          /* Thẻ kính chứa lời chào */
          #boot-splash .bs-card {
            position: absolute; left: 80px; top: calc(50% - 44px); transform: translateY(-50%); z-index: 6;
            width: 498px; max-width: calc(100vw - 48px); padding: 32px 34px 28px; border-radius: 24px;
            background: rgba(255,255,255,.10); backdrop-filter: blur(26px) saturate(150%);
            border: 1px solid rgba(255,255,255,.20);
            box-shadow: 0 30px 90px rgba(0,0,0,.5), inset 0 1px 0 rgba(255,255,255,.24);
            animation: bsCard .75s cubic-bezier(.34,1.32,.64,1) both;
          }
          @keyframes bsCard {
            from { opacity: 0; transform: translateY(-50%) translateX(-30px); }
            to   { opacity: 1; transform: translateY(-50%); }
          }
          #boot-splash .bs-top { display: flex; align-items: center; gap: 12px; }
          #boot-splash .bs-logo {
            width: 42px; height: 42px; border-radius: 14px; flex-shrink: 0;
            background: linear-gradient(140deg,#ff8fb1,#a78bfa);
            display: flex; align-items: center; justify-content: center;
            box-shadow: 0 10px 24px rgba(190,110,220,.45);
          }
          #boot-splash .bs-logo svg { width: 27px; height: 27px; }
          #boot-splash .bs-bn { font-size: .9rem; font-weight: 760; color: #fff; }
          #boot-splash .bs-bs { font-size: .7rem; color: rgba(255,255,255,.58); }
          #boot-splash .bs-kick {
            margin-left: auto; font-size: .68rem; font-weight: 750; letter-spacing: .06em; color: #fff;
            padding: 5px 11px; border-radius: 999px; white-space: nowrap;
            background: rgba(255,255,255,.14); border: 1px solid rgba(255,255,255,.2);
          }
          #boot-splash .bs-hi {
            margin-top: 22px; font-size: 2.05rem; font-weight: 840; letter-spacing: -.04em;
            color: #fff; line-height: 1.1;
          }
          #boot-splash .bs-hi em {
            font-style: normal; background: linear-gradient(100deg,#ffc9de,#c4b5fd 55%,#93c5fd);
            -webkit-background-clip: text; background-clip: text; color: transparent;
          }
          #boot-splash .bs-sub { margin-top: 9px; font-size: .815rem; font-weight: 600; color: rgba(255,255,255,.7); }
          #boot-splash .bs-enter {
            margin-top: 26px; display: flex; align-items: center; gap: 9px;
            font-size: .82rem; font-weight: 600; color: rgba(255,255,255,.78);
            opacity: 0; animation: bsFade .45s ease 2s forwards;
          }
          #boot-splash .bs-enter kbd {
            font: inherit; font-weight: 800; color: #fff; background: rgba(255,255,255,.2);
            border: 1px solid rgba(255,255,255,.3); border-radius: 8px; padding: 4px 12px;
          }
          #boot-splash .bs-hint2 {
            margin-top: 9px; font-size: .72rem; color: rgba(255,255,255,.6);
            opacity: 0; animation: bsFade .45s ease 2.3s forwards;
          }
          @keyframes bsFade { to { opacity: 1; } }
          #boot-splash .bs-blink { animation: bsBlink 1.6s ease-in-out infinite; }
          @keyframes bsBlink { 0%,100% { opacity: .4; } 50% { opacity: 1; } }

          /* Đồng hồ + ngày, góc phải trên */
          #boot-splash .bs-corner {
            position: absolute; right: 74px; top: 60px; z-index: 6; text-align: right;
            animation: bsRight .8s cubic-bezier(.34,1.3,.64,1) .18s both;
          }
          @keyframes bsRight { from { opacity: 0; transform: translateX(26px); } to { opacity: 1; transform: none; } }
          #boot-splash .bs-clock {
            font-size: 3.9rem; font-weight: 250; color: #fff; letter-spacing: -.04em; line-height: 1;
            font-variant-numeric: tabular-nums; text-shadow: 0 6px 40px rgba(0,0,0,.5);
          }
          #boot-splash .bs-date { margin-top: 7px; font-size: .85rem; font-weight: 600; color: rgba(255,255,255,.7); }

          /* Thanh số liệu ở đáy */
          #boot-splash .bs-strip {
            position: absolute; left: 0; right: 0; bottom: 0; z-index: 6; height: 84px;
            display: flex; align-items: center; padding: 0 80px;
            background: linear-gradient(0deg,rgba(6,8,14,.72),rgba(6,8,14,.14));
            border-top: 1px solid rgba(255,255,255,.13); backdrop-filter: blur(14px);
            animation: bsUp .6s ease .45s both;
          }
          @keyframes bsUp { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: none; } }
          #boot-splash .bs-cell { flex: 1; display: flex; flex-direction: column; gap: 3px; position: relative; }
          #boot-splash .bs-cell + .bs-cell { padding-left: 32px; }
          #boot-splash .bs-cell + .bs-cell::before {
            content: ""; position: absolute; left: 0; top: 5px; bottom: 5px; width: 1px;
            background: rgba(255,255,255,.14);
          }
          #boot-splash .bs-k { font-size: .7rem; font-weight: 640; letter-spacing: .05em; color: rgba(255,255,255,.55); }
          #boot-splash .bs-v {
            font-size: 1.36rem; font-weight: 800; color: #fff;
            font-variant-numeric: tabular-nums; letter-spacing: -.02em;
          }
          #boot-splash .bs-v.warn { color: #ffd07a; }
          #boot-splash .bs-v.mut { color: rgba(255,255,255,.4); font-weight: 600; }

          /* Màn hẹp (điện thoại): bỏ đồng hồ, thẻ kính tràn ngang, thanh số liệu gọn lại */
          @media (max-width: 820px) {
            #boot-splash .bs-corner { display: none; }
            #boot-splash .bs-card { left: 24px; right: 24px; width: auto; padding: 24px 22px 22px; }
            #boot-splash .bs-hi { font-size: 1.6rem; }
            #boot-splash .bs-strip { padding: 0 22px; height: 74px; }
            #boot-splash .bs-cell:nth-child(n+4) { display: none; }
            #boot-splash .bs-cell + .bs-cell { padding-left: 18px; }
            #boot-splash .bs-k {
                font-size: .61rem; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
            }
            #boot-splash .bs-v { font-size: 1.15rem; }
          }
        `;
        doc.head.appendChild(css3);

        var el = doc.createElement('div');
        el.id = 'boot-splash';
        el.setAttribute('data-h', '__GREET_HOUR__');
        el.setAttribute('data-s', '__GREET_SEASON__');
        el.innerHTML =
            '<div class="bs-bg"></div><div class="bs-hour"></div><div class="bs-season"></div>' +
            '<div class="bs-scrim"></div><div class="bs-moon"></div>' +
            '<div class="bs-card">' +
              '<div class="bs-top">' +
                '<div class="bs-logo"><svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg"><g>' +
                  '<path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#fff"/>' +
                  '<path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffe8f1" transform="rotate(72 50 50)"/>' +
                  '<path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#fff" transform="rotate(144 50 50)"/>' +
                  '<path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffe8f1" transform="rotate(216 50 50)"/>' +
                  '<path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#fff" transform="rotate(288 50 50)"/>' +
                  '<circle cx="50" cy="50" r="7" fill="#ffd6e6"/>' +
                '</g></svg></div>' +
                '<div><div class="bs-bn">Tân Hotel</div><div class="bs-bs">Front Office toolkit</div></div>' +
                '<span class="bs-kick">__GREET_SEASON_LABEL__</span>' +
              '</div>' +
              '<div class="bs-hi">Chào <em>__GREET_SHIFT__</em>, Tân __GREET_EMOJI__</div>' +
              '<div class="bs-sub">__GREET_SUB__</div>' +
              '<div class="bs-enter"><span class="bs-blink">&#9654;</span> Nhấn <kbd>Enter</kbd> để vào</div>' +
              '<div class="bs-hint2">hoặc chạm/bấm chuột vào màn hình</div>' +
            '</div>' +
            '<div class="bs-corner">' +
              '<div class="bs-clock">__GREET_CLOCK__</div>' +
              '<div class="bs-date">__GREET_DATE__</div>' +
            '</div>' +
            '<div class="bs-strip">__GREET_CELLS__</div>';
        doc.body.appendChild(el);

        // ── Hạt rơi theo mùa — thuần transform/opacity, chạy trên GPU ──
        var BS_PT = {
          spring:   {n: 26, col: ['#f6a8c9','#f293bc','#f9c1d9'], r: '0 60% 0 60%', sz: [10,18],  dur: [5,9],   a: 'bsFall'},
          summer:   {n: 32, col: ['#fff4c0','#ffffff','#bff0ff'], r: '50%',         sz: [3,8],    dur: [6,11],  a: 'bsRise'},
          autumn:   {n: 24, col: ['#e0913f','#c96a2b','#e8b45a'], r: '0 70% 0 70%', sz: [11,19],  dur: [6,10],  a: 'bsFall'},
          winter:   {n: 70, col: ['rgba(200,225,250,.85)'],       r: '2px',         sz: [1.6,2.8],dur: [1.0,1.9],a: 'bsFall', tall: 9},
          tet:      {n: 26, col: ['#ffd24a','#ffc107','#ffe27a'], r: '0 60% 0 60%', sz: [10,17],  dur: [5,9],   a: 'bsFall'},
          trungthu: {n: 13, col: ['#ff9f43','#ffb86b','#ff7f50'], r: '34% 34% 42% 42%', sz: [16,26], dur: [11,18], a: 'bsRise', glow: true},
          noel:     {n: 80, col: ['#ffffff','#eaf4ff'],           r: '50%',         sz: [3,7],    dur: [6,12],  a: 'bsFall'}
        }['__GREET_SEASON__'];
        if (BS_PT) {
            for (var j = 0; j < BS_PT.n; j++) {
                var pt = doc.createElement('div');
                pt.className = 'bs-pt';
                var sz = BS_PT.sz[0] + Math.random() * (BS_PT.sz[1] - BS_PT.sz[0]);
                var du = BS_PT.dur[0] + Math.random() * (BS_PT.dur[1] - BS_PT.dur[0]);
                var col = BS_PT.col[j % BS_PT.col.length];
                var fill = (BS_PT.r === '50%')
                    ? 'radial-gradient(circle,' + col + ',transparent 72%)'
                    : 'radial-gradient(circle at 30% 30%,#fff,' + col + ' 72%)';
                if (BS_PT.glow) fill = 'radial-gradient(circle at 50% 38%,#fff2c0,' + col + ' 72%)';
                pt.style.cssText = (BS_PT.a === 'bsRise' ? 'top:112vh;' : 'top:-42px;') +
                    'left:' + (Math.random() * 100) + 'vw;' +
                    'width:' + sz + 'px;height:' + (BS_PT.tall ? sz * BS_PT.tall : sz * (BS_PT.glow ? 1.3 : 1)) + 'px;' +
                    'background:' + fill + ';border-radius:' + BS_PT.r + ';' +
                    (BS_PT.glow ? 'box-shadow:0 0 18px 4px rgba(255,170,80,.55);' : '') +
                    '--dx:' + (Math.random() * 180 - 90) + 'px;' +
                    'animation:' + BS_PT.a + ' ' + du + 's linear ' + (-Math.random() * du) + 's infinite';
                el.appendChild(pt);
            }
        }

        // ── Nội dung app hiện dần so le sau khi màn chào tan ──
        function bsRevealApp() {
            var seq = [];
            var sb = doc.querySelector('section[data-testid="stSidebar"]');
            if (sb) seq.push(sb);
            var blk = doc.querySelector('div[data-testid="stMainBlockContainer"] > div[data-testid="stVerticalBlock"]');
            if (blk && blk.children.length) {
                Array.prototype.forEach.call(blk.children, function(c){ seq.push(c); });
            } else {
                var m = doc.querySelector('[data-testid="stMain"]');
                if (m) seq.push(m);
            }
            seq.forEach(function(node, i){
                node.style.animationDelay = (i * 0.06) + 's';
                node.classList.add('tan-rv');
            });
            // Gỡ lớp hiệu ứng sau khi chạy xong để không ảnh hưởng các lần rerun sau
            window.parent.setTimeout(function(){
                seq.forEach(function(node){
                    node.classList.remove('tan-rv');
                    node.style.animationDelay = '';
                });
            }, 2200);
        }

        // ── Cổng Enter: chỉ mở sau 2 giây, bấm sớm hơn không có tác dụng ──
        var bsArmed = false, bsDone = false;
        window.parent.setTimeout(function(){ bsArmed = true; }, 2000);

        function bsDismiss() {
            if (bsDone || !bsArmed) return;
            bsDone = true;
            doc.removeEventListener('keydown', bsKey, true);
            document.removeEventListener('keydown', bsKey, true);
            el.classList.add('bs-hide');
            bsRevealApp();
            window.parent.setTimeout(function(){
                if (el.parentNode) el.parentNode.removeChild(el);
            }, 600);
        }
        function bsKey(ev) {
            if (ev.key === 'Enter' || ev.key === ' ' || ev.key === 'Spacebar' || ev.key === 'Escape') {
                ev.preventDefault();
                bsDismiss();
            }
        }
        // Bắt phím ở CẢ tài liệu cha lẫn iframe này — tuỳ nơi con trỏ đang focus
        doc.addEventListener('keydown', bsKey, true);
        document.addEventListener('keydown', bsKey, true);
        el.addEventListener('click', bsDismiss);
        el.addEventListener('touchstart', bsDismiss, {passive: true});
        try { window.parent.focus(); } catch (e) {}
    }
})();
</script>
"""
    _boot_script = _boot_script.replace('__DARK_BG_DATA_URI__', _dark_bg_data_uri())
    _boot_script = _boot_script.replace('__LIGHT_BG_DATA_URI__', _light_bg_data_uri())

    # ── Nội dung màn chào: lấy từ SỐ LIỆU THẬT đã lưu trong ngày. Chưa chạy
    # công cụ nào thì ô số liệu hiện dấu "—", KHÔNG bịa con số. ──
    _g_now = now_vn()
    _g_theme = _compute_effective_theme()
    _g_season, _g_hourkey = _season_key(), _hour_key()
    _g_thu = ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu', 'Thứ Bảy', 'Chủ Nhật'][_g_now.weekday()]
    if THEME_DAY_START_HOUR <= _g_now.hour < 14:
        _g_shift, _g_hours, _g_emo = 'ca sáng', '06:00 – 14:00', '🌅'
    elif 14 <= _g_now.hour < 22:
        _g_shift, _g_hours, _g_emo = 'ca chiều', '14:00 – 22:00', '☀️' if _g_hourkey == 'day' else '🌇'
    else:
        _g_shift, _g_hours, _g_emo = 'ca đêm', '22:00 – 06:00', '🌙'

    _g_tasks = (_load_progress().get('tasks') or {})
    _g_sum = ((_g_tasks.get('daily') or {}).get('summary') or {})
    _g_todo = sum(1 for _k in ('daily', 'regcard', 'recon_person', 'recon_room')
                  if not (_g_tasks.get(_k) or {}).get('done'))
    _g_todo_txt = (f'còn {_g_todo} việc chưa xong' if _g_todo
                   else 'đã xong các việc trong ngày')

    def _g_cell(label, value, cls=''):
        _mut = ' mut' if value is None else (f' {cls}' if cls else '')
        return (f'<div class="bs-cell"><div class="bs-k">{label}</div>'
                f'<div class="bs-v{_mut}">{"—" if value is None else value}</div></div>')

    _g_cells = (_g_cell('KHÁCH LƯU TRÚ', _g_sum.get('total'))
                + _g_cell('QUỐC TẾ', _g_sum.get('intl'))
                + _g_cell('VIỆT NAM', _g_sum.get('vn'))
                + _g_cell('CHECK-IN HÔM NAY', _g_sum.get('checkin_n'))
                + _g_cell('VIỆC CHƯA XONG', _g_todo, 'warn' if _g_todo else ''))

    for _ph, _val in (
        ('__GREET_PHOTO__', _season_bg_data_uri(_g_season, _g_theme)),
        ('__GREET_HOUR__', _g_hourkey),
        ('__GREET_SEASON_LABEL__', SEASON_LABEL[_g_season]),
        ('__GREET_SEASON__', _g_season),
        ('__GREET_SHIFT__', _g_shift),
        ('__GREET_EMOJI__', _g_emo),
        ('__GREET_SUB__', f'{_g_thu}, {_g_now.strftime("%d/%m/%Y")} · {_g_hours} · {_g_todo_txt}'),
        ('__GREET_CLOCK__', _g_now.strftime('%H:%M')),
        ('__GREET_DATE__', f'{_g_thu}, {_g_now.strftime("%d/%m/%Y")}'),
        ('__GREET_CELLS__', _g_cells),
        ('__GREET_THEME__', _g_theme),
    ):
        _boot_script = _boot_script.replace(_ph, _val)
    st.iframe(_boot_script, height=1)

# ── Chế độ giao diện: áp dụng data-theme lên <html> mỗi lần rerun (khác khối
# CSS phía trên chỉ tiêm 1 lần/phiên) — vì hiệu lực có thể đổi giữa các lần
# rerun khi ở chế độ "Tự động" (qua giờ) hoặc khi người dùng đổi lựa chọn.
if 'theme_mode' not in st.session_state:
    st.session_state.theme_mode = 'auto'
_effective_theme = _compute_effective_theme()
# Badge số bên phải mục nav (số khách đã xử lý / số ghi chú sổ giao ca): số thay
# đổi mỗi lần rerun nên tiêm qua CSS ::after ở đây, không nhét vào khối CSS tĩnh.
_bd = st.session_state.get('daily_results') or {}
_bd_n = _bd.get('total') or (st.session_state.get('progress', {})
                             .get('tasks', {}).get('daily', {}).get('summary', {}) or {}).get('total')
try:
    _ho_n = len(db_load_entries(today_vn())) if db_available() else len(
        (st.session_state.get('handover') or {}).get('entries', []))
except Exception:
    _ho_n = 0
_badge_css = ''
if _bd_n:
    _badge_css += f'.st-key-nav_daily button::after{{content:"{_bd_n}";}}'
if _ho_n:
    _badge_css += f'.st-key-nav_handover button::after{{content:"{_ho_n}";}}'
st.iframe(f"""
<script>
(function(){{
  var doc = window.parent.document;
  doc.documentElement.setAttribute('data-theme', '{_effective_theme}');
  var s = doc.getElementById('tan-badge-style');
  if (!s) {{ s = doc.createElement('style'); s.id = 'tan-badge-style'; doc.head.appendChild(s); }}
  s.textContent = {_badge_css!r};
}})();
</script>
""", height=1)



# Menu selection (session state)
if "menu" not in st.session_state:
    st.session_state.menu = "dashboard"

MENU_LABELS = {
    'dashboard': 'Tổng quan ca trực', 'daily': 'Xử lý hàng ngày', 'regcard': 'Regcard + ARR',
    'handover': 'Sổ giao ca', 'recon': 'Đối chiếu (cổng mật khẩu)',
    'recon_person': 'Đối chiếu người nước ngoài', 'recon_room': 'Đối chiếu hệ thống phòng',
}

# Nạp tiến độ đã lưu trên đĩa cho HÔM NAY — chạy 1 lần khi phiên bắt đầu, và
# tự làm mới nếu phiên mở vắt qua nửa đêm (ngày mới → tiến độ trống lại).
if st.session_state.get('progress_date') != today_vn().isoformat():
    st.session_state.progress_date = today_vn().isoformat()
    st.session_state.progress = _load_progress()
    st.session_state.handover = {'entries': list(st.session_state.progress.get('handover_entries', []))}

def set_theme_mode(mode):
    st.session_state.theme_mode = mode

def go_menu(name):
    st.session_state.menu = name

# ── Sidebar điều hướng ────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('''
    <div class="sb-brand">
        <div class="sb-brand-title"><span class="sb-mascot">
            <svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg">
                <g>
                    <path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffc2dd"/>
                    <path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffb3d1" transform="rotate(72 50 50)"/>
                    <path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffc2dd" transform="rotate(144 50 50)"/>
                    <path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffb3d1" transform="rotate(216 50 50)"/>
                    <path d="M50 50 C38 36 40 14 46 9 C48 7 52 7 54 9 C60 14 62 36 50 50 Z" fill="#ffc2dd" transform="rotate(288 50 50)"/>
                    <circle cx="50" cy="50" r="7" fill="#fff6ee"/>
                    <circle cx="47" cy="47" r="1.3" fill="#ffcf6b"/>
                    <circle cx="53" cy="47" r="1.3" fill="#ffcf6b"/>
                    <circle cx="50" cy="52.5" r="1.3" fill="#ffcf6b"/>
                </g>
            </svg>
        </span> Tân Hotel</div>
        <div class="sb-brand-sub">Front Office toolkit</div>
    </div>
    ''', unsafe_allow_html=True)

    st.markdown('<div class="sb-section">Công cụ</div>', unsafe_allow_html=True)
    st.button("Tổng quan ca trực", key="nav_dashboard", use_container_width=True,
              icon=":material/dashboard:",
              type="primary" if st.session_state.menu == "dashboard" else "secondary",
              on_click=go_menu, args=("dashboard",))
    st.button("Xử lý hàng ngày", key="nav_daily", use_container_width=True,
              icon=":material/checklist:",
              type="primary" if st.session_state.menu == "daily" else "secondary",
              on_click=go_menu, args=("daily",))
    st.button("Regcard + ARR", key="nav_regcard", use_container_width=True,
              icon=":material/print:",
              type="primary" if st.session_state.menu == "regcard" else "secondary",
              on_click=go_menu, args=("regcard",))
    st.button("Sổ giao ca", key="nav_handover", use_container_width=True,
              icon=":material/handshake:",
              type="primary" if st.session_state.menu == "handover" else "secondary",
              on_click=go_menu, args=("handover",))

    st.markdown('<div class="sb-section">Đối chiếu</div>', unsafe_allow_html=True)
    if st.session_state.get("recon_ok"):
        st.button("Người nước ngoài", key="nav_recon_person", use_container_width=True,
                  icon=":material/public:",
                  type="primary" if st.session_state.menu == "recon_person" else "secondary",
                  on_click=go_menu, args=("recon_person",))
        st.button("Hệ thống phòng", key="nav_recon_room", use_container_width=True,
                  icon=":material/door_front:",
                  type="primary" if st.session_state.menu == "recon_room" else "secondary",
                  on_click=go_menu, args=("recon_room",))
    else:
        st.button("Đối chiếu lưu trú", key="nav_recon", use_container_width=True,
                  icon=":material/lock:",
                  type="primary" if st.session_state.menu == "recon" else "secondary",
                  on_click=go_menu, args=("recon",))

    st.markdown('<div class="sb-section">Giao diện</div>', unsafe_allow_html=True)
    # Nút gạt 3 trạng thái (segmented) thay cho ô chọn — nhìn gọn và bấm 1 lần
    _tcols = st.columns(3, gap="small")
    for _tci, (_tk, _tlbl) in enumerate([('auto', 'Tự động'), ('light', 'Sáng'), ('dark', 'Tối')]):
        _tcols[_tci].button(_tlbl, key=f"theme_btn_{_tk}", use_container_width=True,
                            type="primary" if st.session_state.get('theme_mode') == _tk else "secondary",
                            on_click=set_theme_mode, args=(_tk,))

    _sb_shift = ('Ca sáng · 06:00–14:00' if 6 <= now_vn().hour < 14
                 else 'Ca chiều · 14:00–22:00' if 14 <= now_vn().hour < 22 else 'Ca đêm · 22:00–06:00')
    _sb_store = 'Supabase đã kết nối' if db_available() else 'Lưu tạm trên máy chủ'
    st.markdown(
        f'<div class="sb-user"><div class="sb-av">T</div>'
        f'<div><div class="sb-un">Tân</div><div class="sb-ur">{_sb_shift}</div></div></div>'
        f'<div class="sb-status"><span class="sb-dot"></span>{_sb_store}</div>',
        unsafe_allow_html=True)

# ── Thanh top bar: breadcrumb + trạng thái + hành động (thay banner "Welcome") ──
_page_name = MENU_LABELS.get(st.session_state.menu, 'Tổng quan ca trực')
_tb_now = now_vn()
_tb_chip = ('<span class="tan-chip ok">● Supabase đã kết nối</span>' if db_available()
            else '<span class="tan-chip">● Lưu tạm trên máy chủ</span>')
_tb_d = st.session_state.get('daily_results')
_tb_rc = st.session_state.get('rc_results')
_tb_rp = st.session_state.get('recon_results')
_tb_rr = st.session_state.get('reconr_results')
# Báo cáo ngày chỉ dựng khi đang ở Tổng quan VÀ đã có ít nhất 1 công cụ chạy
# xong — tránh dựng workbook thừa ở mọi trang, mọi lần rerun.
_tb_report = bool(_tb_d or _tb_rc or _tb_rp or _tb_rr) and st.session_state.menu == "dashboard"

with st.container(key="tan_topbar"):
    _tbl, _tbm, _tbr = st.columns([4.6, 1.25, 1.35], vertical_alignment="center")
    _tbl.markdown(
        f'<div class="tan-topbar"><span class="tan-crumb">Tân Hotel</span>'
        f'<span class="tan-sep">/</span><b>{_page_name}</b>'
        f'<span class="tan-topbar-r">{_tb_chip}'
        f'<span class="tan-chip">📅 {_tb_now.strftime("%d/%m/%Y")}</span></span></div>',
        unsafe_allow_html=True)
    if _tb_report:
        _tb_rp_date = (_tb_d or {}).get('date_str') or today_vn().strftime('%d_%m')
        _tb_wb = build_daily_report(_tb_rp_date.replace('_', '/'),
                                    _tb_d if _tb_d and _tb_d.get('has_xlsx') else None,
                                    (_tb_rc or {}).get('arr_stats'), _tb_rp, _tb_rr)
        _tbm.download_button("⇩ Xuất báo cáo", wb_to_bytes(_tb_wb),
                             file_name=f"bao_cao_ngay_{_tb_rp_date}.xlsx",
                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             use_container_width=True, key="tb_report_dl")
    else:
        _tbm.button("⇩ Xuất báo cáo", key="tb_report_off", use_container_width=True, disabled=True,
                    help="Mở Tổng quan ca trực và chạy ít nhất một công cụ để xuất báo cáo ngày")
    _tbr.button("⚡ Xử lý hàng ngày", key="tb_go_daily", use_container_width=True,
                type="primary", on_click=go_menu, args=("daily",))

# ── Tạo file ARR từ file Arrival (Book) Smile ──────────────────────────────
def build_arr(book_bytes):
    """Tạo file ARR ĐÚNG định dạng của ARR Converter gốc (tool HTML riêng, không
    phải file mẫu in cũ):
    - 6 cột: Conf# / Arrival / Departure / Company / Notice / [số phòng] — đọc
      cột nguồn theo TÊN (Conf#, Folio#, Type, Arrival, Departure, Company,
      Notice), không theo vị trí cố định như bản cũ.
    - Font Patrick Hand toàn bộ; Conf# cỡ 50 đậm nền cam nhạt; số phòng cỡ 50;
      các ô còn lại cỡ 20. Dòng dữ liệu cao 120, dòng header cao 142.5.
    - Số phòng = số dòng Folio# hợp lệ trùng Conf# (bỏ dòng Type='**' - dummy).
    - Dòng phụ chèn ngay sau booking tương ứng, gộp A:F, nền màu theo loại:
      CÀ THẺ (cam) · THU TIỀN (xanh lá) · XEM LẠI BU (vàng) · FOC LATE C/O (xanh
      dương, tự đọc giờ trong Notice nếu có, vd "FOC LATE C/O 18:00").
    - Nhận diện nghiệp vụ đầy đủ (OTA, từ khóa CÀ THẺ/THU TIỀN/FOC/XEM LẠI BU)
      y hệt bộ từ khóa của ARR Converter gốc.
    """
    import re as _re2
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    df = pd.read_excel(io.BytesIO(book_bytes), header=None)
    hdr = None
    for i in range(min(5, len(df))):
        if any(str(v).strip() == 'Conf#' for v in df.iloc[i] if pd.notna(v)):
            hdr = i; break
    if hdr is None:
        raise ValueError("Không tìm thấy dòng header chứa \"Conf#\" trong file. Kiểm tra lại file Arrival Smile.")

    headers = [str(v).strip() if pd.notna(v) else '' for v in df.iloc[hdr]]
    col = {}
    for i, h in enumerate(headers):
        if h and h not in col:
            col[h] = i
    data = df.iloc[hdr + 1:].reset_index(drop=True)

    def C(name):
        return col.get(name, -1)

    conf_c, folio_c, type_c = C('Conf#'), C('Folio#'), C('Type')
    arr_c, dep_c, comp_c, notice_c = C('Arrival'), C('Departure'), C('Company'), C('Notice')
    if -1 in (conf_c, folio_c, arr_c, comp_c):
        raise ValueError("File thiếu cột bắt buộc (Conf#, Folio#, Arrival, Company). Kiểm tra lại file.")

    # Sửa lỗi Excel đảo dd/mm↔mm/dd với ngày ≤12 từ Smile export (dùng chung hàm _fix_date)
    def _arr_fmt_date(v):
        fixed = _fix_date(v)
        if fixed is not None:
            return fixed.strftime('%d/%m/%y')
        return str(v).strip() if isinstance(v, str) and v.strip() else ''

    # ── Đếm số phòng (= số dòng Folio# hợp lệ) theo từng Conf#, bỏ dòng Type='**' (dummy) ──
    room_counts = {}
    dummy_count = 0
    for _, row in data.iterrows():
        conf = row.iloc[conf_c] if conf_c >= 0 else None
        folio = row.iloc[folio_c] if folio_c >= 0 else None
        if pd.notna(conf) and pd.notna(folio):
            typ = row.iloc[type_c] if type_c >= 0 else None
            if str(typ).strip() == '**':
                dummy_count += 1
                continue
            room_counts[conf] = room_counts.get(conf, 0) + 1

    # ── Danh sách booking theo ĐÚNG thứ tự xuất hiện, mỗi Conf# 1 dòng ──
    seen = set()
    ordered = []
    for _, row in data.iterrows():
        conf = row.iloc[conf_c] if conf_c >= 0 else None
        folio = row.iloc[folio_c] if folio_c >= 0 else None
        typ = row.iloc[type_c] if type_c >= 0 else None
        arr = row.iloc[arr_c] if arr_c >= 0 else None
        comp = row.iloc[comp_c] if comp_c >= 0 else None
        dep = row.iloc[dep_c] if dep_c >= 0 else None
        notice = row.iloc[notice_c] if notice_c >= 0 else None
        if pd.isna(conf) or pd.isna(folio) or pd.isna(arr) or pd.isna(comp):
            continue
        if str(typ).strip() == '**':
            continue
        if conf in seen:
            continue
        if not room_counts.get(conf):
            continue
        seen.add(conf)
        ordered.append({
            'type': 'bk', 'conf': conf,
            'arrival': _arr_fmt_date(arr),
            'departure': _arr_fmt_date(dep) if pd.notna(dep) else '',
            'company': str(comp).strip(),
            'notice': str(notice).strip() if pd.notna(notice) else '',
            'rooms': room_counts[conf],
        })
    if not ordered:
        raise ValueError("File không có dữ liệu booking hợp lệ nào.")

    # ── Bộ nhận diện nghiệp vụ — y hệt ARR Converter gốc ──
    ARR_OTA = ['EXPEDIA','BOOKING','AGODA','TRIP.COM','CTRIP','AIRBNB','TRAVELOKA',
        'KLOOK','KAYAK','PRICELINE','HOTELS.COM','ORBITZ','TRIVAGO','MAKEMYTRIP',
        'LASTMINUTE','HOSTELWORLD','WOTIF','HOTWIRE','VRBO','HOMEAWAY','IVIVU',
        'MYTOUR','LUXSTAY','VNTRIP','GOTADI','TRAVELPORT','SKYSCANNER','BESTPRICE',
        'LATEROOMS','EASYJET','RYANAIR','JETSTAR','HOTELBEDS','TOURICO','GETAROOM']
    ARR_CA_THE = ['TACC','CC UPON','CHARGE CC','CHARGE CARD','CREDIT CARD','DEBIT CARD',
        'CC AUTH','AUTH CC','AUTHORIZE CC','AUTHORIZE CARD','CC ON ARRIVAL',
        'BILL TO CC','SWIPE CC','SWIPE CARD','PRE-AUTH','PREAUTH','PREPAID CC',
        'CHARGE ON CC','CARD ON ARRIVAL','CC AT CI','CC AT CHECK','PAY BY CARD',
        'CARD PAYMENT','TC UPON','TC ON ARRIVAL','TAKE CC','TAKE CARD']
    ARR_THU_TIEN = ['PAY UPON','PAY ON ARRIVAL','CASH ON ARRIVAL','CASH UPON','COLLECT CASH',
        'COLLECT PAYMENT','COLLECT ON ARRIVAL','PAYMENT ON ARRIVAL','CASH PAYMENT',
        'CASH AT CHECK','CASH AT CI','DUE ON ARRIVAL','PAYABLE ON ARRIVAL',
        'PAY AT CI','PAY AT CHECK','CASH DUE','OUTSTANDING','BALANCE DUE',
        'PAYMENT DUE','COLLECT AT CI','COLLECT AT CHECK',
        'RC UPON','UPON C/I','UPON CI','UPON CHECK-IN','UPON CHECKIN',
        'GOA UPON','ROH UPON','COLLECT UPON']
    ARR_XEM_LAI = ['CASH UPON','CASH ON ARRIVAL','CASH AT CI','CASH PAYMENT',
        'PAY UPON','PAY AT CI','COLLECT CASH','CASH DUE']
    ARR_FOC_LCO = ['FOC LATE CHECK','FOC LATE CHECKOUT','FOC LATE C/O','FOC LCO',
        'LCO FOC','LATE CHECK OUT FOC','LATE CHECKOUT FOC','LATE C/O FOC',
        'COMP LATE CHECK','COMP LATE CHECKOUT','COMP LCO',
        'COMPLIMENTARY LATE CHECK','COMPLIMENTARY LCO',
        'GRATIS LATE CHECK','FREE LATE CHECK']

    def _pay_type(notice, company):
        n = str(notice or '').upper()
        co = str(company or '').upper()
        is_ota = any(o in co for o in ARR_OTA)
        is_ca_the = any(k in n for k in ARR_CA_THE)
        is_thu_tien = any(k in n for k in ARR_THU_TIEN)
        is_foc_lco = any(k in n for k in ARR_FOC_LCO)
        has_foc = 'FOC' in n or 'COMP' in n or 'COMPLIMENTARY' in n
        has_lco = 'LATE CHECK' in n or ' LCO' in n or 'LATE C/O' in n or 'LATE CHECKOUT' in n
        if is_foc_lco or (has_foc and has_lco):
            return 'foc_lco'
        if is_ota and any(k in n for k in ARR_XEM_LAI):
            return 'xem_lai_bu'
        if is_ca_the:
            return 'ca_the'
        if is_thu_tien:
            return 'thu_tien'
        return 'none'

    result = []
    for i, bk in enumerate(ordered):
        result.append(bk)
        if i >= len(ordered) - 1:
            continue
        pt = _pay_type(bk['notice'], bk['company'])
        if pt == 'ca_the':
            result.append({'type': 'sep', 'conf': 'CÀ THẺ'})
        elif pt == 'thu_tien':
            result.append({'type': 'sep', 'conf': 'THU TIỀN'})
        elif pt == 'xem_lai_bu':
            result.append({'type': 'sep', 'conf': 'XEM LẠI BU'})
        elif pt == 'foc_lco':
            m = _re2.search(r'\b(\d{1,2}[:Hh]\d{2})\b', bk['notice'])
            time_str = (' ' + m.group(1).upper()) if m else ''
            result.append({'type': 'sep', 'conf': 'FOC LATE C/O' + time_str})

    # ── Xuất Excel đúng định dạng ARR Converter gốc ──
    wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
    for i, w in enumerate([39.4, 15.1, 16.0, 21.9, 50.0, 10.3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style='thin')
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    conf_fill = PatternFill('solid', fgColor='FDEADA')
    fill_colors = {'CÀ THẺ': 'FDEADA', 'THU TIỀN': 'D4F4E8', 'XEM LẠI BU': 'FFF8DC'}

    ws.row_dimensions[1].height = 142.5
    for i, h in enumerate(['Conf#', 'Arrival', 'Departure', 'Company', 'Notice', None], 1):
        cell = ws.cell(1, i)
        cell.value = h
        cell.font = Font(name='Patrick Hand', size=50 if h == 'Conf#' else 20)
        cell.alignment = center_wrap
        cell.border = border_all

    r = 2
    for item in result:
        ws.row_dimensions[r].height = 120.0
        if item['type'] == 'sep':
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            label = item.get('conf') or 'CÀ THẺ'
            color = fill_colors.get(label) or ('D6EAF8' if label.startswith('FOC LATE C/O') else 'FDEADA')
            cell = ws.cell(r, 1)
            cell.value = label
            cell.font = Font(name='Patrick Hand', size=50)
            cell.alignment = center_wrap
            cell.fill = PatternFill('solid', fgColor=color)
            cell.border = border_all
        else:
            vals = [item['conf'], item['arrival'], item['departure'], item['company'], item['notice'], item['rooms']]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci)
                cell.value = v
                cell.font = Font(name='Patrick Hand', size=50 if ci in (1, 6) else 20, bold=(ci == 1))
                cell.alignment = center_wrap
                cell.border = border_all
                if ci == 1:
                    cell.fill = conf_fill
        r += 1

    bookings = [x for x in result if x['type'] == 'bk']
    stats = {
        'bookings': len(bookings),
        'rooms': sum(b['rooms'] for b in bookings),
        'ota': sum(1 for b in bookings if any(o in b['company'].upper() for o in ARR_OTA)),
        'dummy': dummy_count,
        'ca_the': sum(1 for x in result if x['type'] == 'sep' and x['conf'] == 'CÀ THẺ'),
        'thu_tien': sum(1 for x in result if x['type'] == 'sep' and x['conf'] == 'THU TIỀN'),
        'xem_lai_bu': sum(1 for x in result if x['type'] == 'sep' and x['conf'] == 'XEM LẠI BU'),
        'foc_lco': sum(1 for x in result if x['type'] == 'sep' and x['conf'].startswith('FOC LATE C/O')),
    }
    return wb, stats


# ── Dashboard ca trực ─────────────────────────────────────────────────────
if st.session_state.menu == "dashboard":
    _now = now_vn()
    _thu = ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu', 'Thứ Bảy', 'Chủ Nhật'][_now.weekday()]
    _shift = ('Ca sáng' if 6 <= _now.hour < 14 else 'Ca chiều' if 14 <= _now.hour < 22 else 'Ca đêm')

    _d = st.session_state.get('daily_results')
    _rc = st.session_state.get('rc_results')
    _re_p = st.session_state.get('recon_results')
    _re_r = st.session_state.get('reconr_results')
    # Tiến độ đã lưu trên đĩa cho HÔM NAY — còn nguyên dù tải lại trang hoặc
    # công cụ đó được chạy ở 1 phiên/tab khác trước đó trong ngày.
    _ptasks = st.session_state.get('progress', {}).get('tasks', {})
    _p_daily = _ptasks.get('daily', {})
    _p_regcard = _ptasks.get('regcard', {})
    _p_rp = _ptasks.get('recon_person', {})
    _p_rr = _ptasks.get('recon_room', {})

    st.markdown(f'<div class="tan-h1">Tổng quan ca trực</div>'
                f'<div class="tan-h1sub">{_thu}, {_now.strftime("%d/%m/%Y")} · {_shift} · '
                f'cập nhật {_now.strftime("%H:%M")}</div>', unsafe_allow_html=True)

    # Số liệu khách: ưu tiên kết quả phiên hiện tại, dự phòng bản đã lưu trong ngày
    if _d and _d.get('has_xlsx'):
        _iss_d = _d.get('issues')
        _n_red = int((_iss_d['Mức độ'] == '🔴').sum()) if _iss_d is not None and len(_iss_d) else 0
        _n_yel = int((_iss_d['Mức độ'] == '🟡').sum()) if _iss_d is not None and len(_iss_d) else 0
        _hero = {'total': _d['total'], 'intl': _d['intl'], 'vn': _d['vn'], 'red': _n_red, 'yellow': _n_yel,
                 'rooms': _d.get('rooms_cnt'), 'cin': _d.get('checkin_n'), 'cout': _d.get('checkout_n'),
                 'visa_watch': len(_d.get('visa_watch') or []),
                 'unknown_nats': len(_d.get('unknown_nats') or []),
                 'invalid_ids': len(_d.get('kbtt_invalid_ids') or []), 'stale': None}
    elif _p_daily.get('summary'):
        _ps = _p_daily['summary']
        _hero = {'total': _ps.get('total'), 'intl': _ps.get('intl'), 'vn': _ps.get('vn'),
                 'red': _ps.get('red_issues') or 0, 'yellow': _ps.get('yellow_issues') or 0,
                 'rooms': _ps.get('rooms_cnt'), 'cin': _ps.get('checkin_n'), 'cout': _ps.get('checkout_n'),
                 'visa_watch': _ps.get('visa_watch_count') or 0,
                 'unknown_nats': 0, 'invalid_ids': 0, 'stale': _p_daily.get('time')}
    else:
        _hero = None

    # ── HÀNG 1: thẻ HERO (ảnh mèo) + 2 thẻ số liệu ──
    _r1a, _r1b, _r1c = st.columns([2, 1, 1], gap="small")
    with _r1a:
        if _hero:
            _yday = _yesterday_total()
            _cmp = ''
            if _yday:
                _delta = (_hero['total'] or 0) - _yday
                _cmp = (f'▲ {_delta} khách so với hôm qua · ' if _delta > 0 else
                        f'▼ {abs(_delta)} khách so với hôm qua · ' if _delta < 0 else
                        'bằng hôm qua · ')
            _room_txt = f"{_hero['rooms']} phòng có khách · " if _hero.get('rooms') else ''
            _stale_txt = (f"số liệu lúc {_hero['stale']}" if _hero['stale'] else "cập nhật trong phiên này")
            st.markdown(
                '<div class="tan-hero">'
                '<div class="tan-hero-lab">🛏️ Tổng khách lưu trú hôm nay</div>'
                f'<div class="tan-hero-val">{_hero["total"]}</div>'
                f'<div class="tan-hero-sub">{_cmp}{_room_txt}{_stale_txt}</div>'
                '<div class="tan-hero-split">'
                f'<div><div class="tan-hero-k">🌍 Quốc tế</div><div class="tan-hero-v">{_hero["intl"]}</div></div>'
                f'<div><div class="tan-hero-k">🇻🇳 Việt Nam</div><div class="tan-hero-v">{_hero["vn"]}</div></div>'
                f'<div><div class="tan-hero-k">🔑 Check-in</div><div class="tan-hero-v">'
                f'{_hero["cin"] if _hero.get("cin") is not None else "—"}</div></div>'
                f'<div><div class="tan-hero-k">🚪 Check-out</div><div class="tan-hero-v">'
                f'{_hero["cout"] if _hero.get("cout") is not None else "—"}</div></div>'
                '</div></div>', unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="tan-hero">'
                '<div class="tan-hero-lab">👋 Chào ca trực</div>'
                '<div class="tan-hero-val" style="font-size:1.45rem">Chưa có số liệu hôm nay</div>'
                '<div class="tan-hero-sub">Bắt đầu bằng <b>Xử lý hàng ngày</b> hoặc '
                '<b>Regcard + ARR</b> ở sidebar — số liệu sẽ tự lên đây.</div>'
                '<div class="tan-hero-split">'
                '<div><div class="tan-hero-k">🌍 Quốc tế</div><div class="tan-hero-v">—</div></div>'
                '<div><div class="tan-hero-k">🇻🇳 Việt Nam</div><div class="tan-hero-v">—</div></div>'
                '<div><div class="tan-hero-k">🔑 Check-in</div><div class="tan-hero-v">—</div></div>'
                '<div><div class="tan-hero-k">🚪 Check-out</div><div class="tan-hero-v">—</div></div>'
                '</div></div>', unsafe_allow_html=True)

    with _r1b:
        with st.container(border=True, height=200):
            _red_n = _hero['red'] if _hero else None
            _cls = 'err' if _red_n else 'ok'
            _red_pct = round(100 * _red_n / _hero['total']) if (_hero and _hero.get('total') and _red_n) else 0
            st.markdown(
                f'<div class="kpi-lab"><span class="kpi-ic {_cls}">⚠️</span>Lỗi dữ liệu</div>'
                f'<div class="kpi-val {"err" if _red_n else ""}">{_red_n if _hero else "—"}</div>'
                f'<div class="kpi-sub">{"cần sửa trước khi nộp" if _red_n else ("dữ liệu sạch" if _hero else "chưa có dữ liệu")}</div>'
                f'<div class="kpi-bar"><i class="err" style="width:{min(_red_pct, 100)}%"></i></div>',
                unsafe_allow_html=True)

    with _r1c:
        with st.container(border=True, height=200):
            _rate = st.session_state.get('rate_input') or (_d or {}).get('rate')
            st.markdown(
                '<div class="kpi-lab"><span class="kpi-ic acc">💱</span>Tỷ giá VCB</div>'
                f'<div class="kpi-val" style="font-size:1.5rem">{f"{_rate:,.0f}" if _rate else "—"}</div>'
                f'<div class="kpi-sub">{"USD → VNĐ (chuyển khoản)" if _rate else "lấy tỷ giá ở Xử lý hàng ngày"}</div>'
                f'<div class="kpi-bar"><i style="width:{100 if _rate else 0}%"></i></div>',
                unsafe_allow_html=True)

    st.write("")

    # ── HÀNG 2: tiến độ công việc + cảnh báo + bàn giao gần nhất ──
    _TICK_ON = ('<span style="display:inline-block;width:15px;height:15px;border-radius:5px;'
                'background:linear-gradient(135deg,#34d399,#059669);color:#fff;font-size:9px;'
                'line-height:15px;text-align:center;vertical-align:-2px;">✓</span>')
    _TICK_OFF = ('<span style="display:inline-block;width:15px;height:15px;border-radius:5px;'
                 'border:1.5px solid var(--line);vertical-align:-2px;"></span>')
    _r2a, _r2b, _r2c = st.columns([2, 1, 1], gap="small")

    with _r2a:
        with st.container(border=True, height=420, key='dash_tasks'):
            if db_available():
                _ho_rows = db_load_entries(today_vn())
                _handover_n = len(_ho_rows)
            else:
                _ho_rows = None
                _handover_n = len((st.session_state.get('handover') or {}).get('entries', []))
            _tasks = [
                ("Xử lý hàng ngày (KBTT · VNM · ĐK14)", _d is not None or _p_daily.get('done'),
                 "daily", None if _d else _p_daily.get('time')),
                ("Regcard + file ARR", _rc is not None or _p_regcard.get('done'),
                 "regcard", None if _rc else _p_regcard.get('time')),
                ("Đối chiếu người nước ngoài", _re_p is not None or _p_rp.get('done'),
                 "recon_person" if st.session_state.get("recon_ok") else "recon",
                 None if _re_p else _p_rp.get('time')),
                ("Đối chiếu hệ thống phòng", _re_r is not None or _p_rr.get('done'),
                 "recon_room" if st.session_state.get("recon_ok") else "recon",
                 None if _re_r else _p_rr.get('time')),
                (f"Sổ giao ca ({_handover_n} ghi chú)", _handover_n > 0, "handover", None),
            ]
            _done_n = sum(1 for _t in _tasks if _t[1])
            _pend = next((_t[2] for _t in _tasks if not _t[1]), None)
            _hc1, _hc2 = st.columns([2.5, 1], vertical_alignment="center")
            _hc1.markdown(f'<div class="pan-h" style="border:0;padding-bottom:0;margin-bottom:0">'
                          f'<span class="pan-t">Tiến độ công việc trong ca</span>'
                          f'<span class="tan-chip">{_done_n}/{len(_tasks)}</span></div>',
                          unsafe_allow_html=True)
            _hc2.button("Xem tất cả →", key="dash_seeall", use_container_width=True,
                        disabled=_pend is None,
                        help="Mở công cụ chưa chạy đầu tiên trong ca",
                        on_click=go_menu, args=(_pend or "daily",))
            st.markdown('<div style="border-bottom:1px solid var(--line2);margin:0 0 .35rem"></div>',
                        unsafe_allow_html=True)
            _row_html = []
            for _label, _done, _target, _stale_time in _tasks:
                _chip = ('<span class="tan-chip ok">Xong</span>' if _done
                         else '<span class="tan-chip">Chưa chạy</span>')
                _tm = _stale_time or ('✓' if _done else '—')
                _row_html.append(
                    f'<div class="pan-task">{_TICK_ON if _done else _TICK_OFF}'
                    f'<span class="pan-rt">{_label}</span>{_chip}'
                    f'<span class="pan-time">{_tm}</span></div>')
            st.markdown(''.join(_row_html), unsafe_allow_html=True)
            # 2 dòng tóm tắt cuối panel — số booking/phòng đến (khi Regcard + ARR đã chạy)
            _sum_as = (_rc or {}).get('arr_stats') or _p_regcard.get('summary')
            st.markdown(
                f'<div class="pan-foot"><div class="pan-line"><span>Phòng đến trong ngày</span>'
                f'<span class="pan-v">{(_sum_as or {}).get("rooms", "—")}</span></div>'
                f'<div class="pan-line"><span>Booking đã xử lý</span>'
                f'<span class="pan-v">{(_sum_as or {}).get("bookings", "—")}</span></div></div>',
                unsafe_allow_html=True)

    with _r2b:
        with st.container(border=True, height=420, key='dash_alerts'):
            _alerts = []
            if _hero:
                if _hero.get('red'):
                    _alerts.append(('r', '🔴', f"{_hero['red']} vấn đề cần sửa",
                                    'phải xử lý trước khi nộp hồ sơ công an'))
                if _hero.get('visa_watch'):
                    _alerts.append(('a', '🛂', f"{_hero['visa_watch']} khách sắp hết hạn tạm trú",
                                    'kiểm tra ở Xử lý hàng ngày'))
                if _hero.get('yellow'):
                    _alerts.append(('a', '🟡', f"{_hero['yellow']} vấn đề nên kiểm tra",
                                    'không chặn nộp hồ sơ'))
                if _hero.get('invalid_ids'):
                    _alerts.append(('a', '🪪', f"{_hero['invalid_ids']} khách chưa có hộ chiếu thật",
                                    'đã để trống số giấy tờ trong KBTT'))
                if _hero.get('unknown_nats'):
                    _alerts.append(('a', '🌐', f"{_hero['unknown_nats']} quốc tịch chưa có mã",
                                    'đã giữ nguyên chữ gốc, cần kiểm tra'))
            st.markdown(f'<div class="pan-h"><span class="pan-t">⚠️ Cảnh báo</span>'
                        f'<span class="tan-chip">{len(_alerts)}</span></div>', unsafe_allow_html=True)
            if _alerts:
                st.markdown(''.join(
                    f'<div class="pan-al {_k}"><div class="pan-al-ic">{_ic}</div>'
                    f'<div><div class="pan-al-t">{_t}</div><div class="pan-al-m">{_m}</div></div></div>'
                    for _k, _ic, _t, _m in _alerts), unsafe_allow_html=True)
            elif _hero:
                st.markdown('<div class="pan-empty">✅ Không có cảnh báo nào — dữ liệu hôm nay sạch.</div>',
                            unsafe_allow_html=True)
            else:
                st.markdown('<div class="pan-empty">Cảnh báo sẽ hiện sau khi chạy '
                            '<b>Xử lý hàng ngày</b>.</div>', unsafe_allow_html=True)

    with _r2c:
        with st.container(border=True, height=420, key='dash_ho'):
            st.markdown('<div class="pan-h"><span class="pan-t">Bàn giao gần nhất</span></div>',
                        unsafe_allow_html=True)
            _recent = []
            if _ho_rows is not None and len(_ho_rows):
                for _, _r in _ho_rows.head(5).iterrows():
                    _room = f" · P.{_r['room']}" if _r['room'] else ""
                    _recent.append((_r['entry_time'], f"{_r['category']}{_room}"))
            elif not db_available():
                for _e in reversed((st.session_state.get('handover') or {}).get('entries', [])[-5:]):
                    _room = f" · P.{_e['room']}" if _e.get('room') else ""
                    _recent.append((_e['time'], f"{_e['cat']}{_room}"))
            if _recent:
                st.markdown(''.join(
                    f'<div class="pan-line"><span>🏷 {_t}</span><span class="pan-mut">{_tm}</span></div>'
                    for _tm, _t in _recent), unsafe_allow_html=True)
            else:
                st.markdown('<div class="pan-empty">Chưa có ghi chú nào trong ca này.</div>',
                            unsafe_allow_html=True)
            st.button("Mở sổ giao ca →", key="dash_go_ho", use_container_width=True,
                      on_click=go_menu, args=("handover",))

    st.write("")

# ── Daily processing screen ───────────────────────────────────────────────
if st.session_state.menu == "daily":
    st.write("")
    st.markdown('<div class="section-label">⚙️ Cài đặt</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col1, col2 = st.columns(2)
        with col1:
            if 'rate_input' not in st.session_state:
                st.session_state.rate_input = 29535.15
            def _apply_vcb_rate():
                try:
                    _rates, _ts = fetch_vcb_rates()
                    st.session_state.rate_input = _rates['USD']
                    _eur = f" · EUR {_rates['EUR']:,.2f}" if _rates.get('EUR') else ''
                    st.session_state['vcb_note'] = (
                        f"✅ Tỷ giá chuyển khoản VCB lúc {_ts}: USD {_rates['USD']:,.2f}{_eur} "
                        f"— đã điền USD vào ô tỷ giá.")
                except Exception as _e:
                    st.session_state['vcb_note'] = (
                        f"⚠️ Không lấy được tỷ giá VCB ({_e}) — nhập tay như bình thường.")
            rate = st.number_input("💱 Tỷ giá USD/EUR → VNĐ", step=0.01, format="%.2f", key="rate_input")
            st.button("🔄 Lấy tỷ giá VCB (chuyển khoản)", on_click=_apply_vcb_rate, key="btn_vcb")
            if st.session_state.get('vcb_note'):
                st.caption(st.session_state['vcb_note'])
        with col2:
            today = today_vn()
            date_str = st.text_input("📅 Ngày (dùng cho tên file)", value=f"{today.day}_{today.month:02d}")

    st.write("")
    st.markdown('<div class="section-label">🚔 Cài đặt ĐK14</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col_n, col_c = st.columns(2)
        with col_n:
            dk_notifier = st.text_input(
                "Người thông báo lưu trú", value="Đỗ Duy Tân", key="dk_notifier",
                help="Điền vào cột (11) khi file nguồn không có sẵn tên. Đổi được tuỳ ý.")
        with col_c:
            dk_checkin = st.date_input(
                "Ngày check-in mặc định", value=today, key="dk_checkin", format="DD/MM/YYYY",
                help="Dùng cho ô 'Từ ngày' ở đầu sổ, và cho khách thiếu ngày đến trong file nguồn.")

    st.write("")
    st.markdown('<div class="section-label">📂 Tải file lên</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col_x, col_s = st.columns(2)
        with col_x:
            xlsx_file = st.file_uploader("File XLSX — Dữ liệu khách (bắt buộc)", type=['xlsx'], key="daily_xlsx")
        with col_s:
            xls_file = st.file_uploader("File IH — Nguồn ĐK14 (tùy chọn)", type=['xls', 'xlsx'], key="daily_xls")

        visa_file = st.file_uploader(
            "File Visa — dữ liệu thô date visa (tùy chọn, để tự điền cột 'Thời hạn tạm trú tại VN' trong KBTT)",
            type=['xlsx'], key="daily_visa")

    st.write("")

    if st.button("⚡ Bắt đầu xử lý", type="primary", disabled=(xlsx_file is None and xls_file is None), use_container_width=True):
        with st.spinner("Đang xử lý..."):
            try:
                progress = st.progress(0, text="Bắt đầu...")
                zip_buf = io.BytesIO()
                files_made = []
                has_xlsx = xlsx_file is not None
                has_dk14 = False
                conv = 0; gks_cnt = 0; gbl_cnt = 0
                df = None; df_intl = None; df_vn = None
                visa_map = {}; visa_unmatched = []; visa_source = None
                kbtt_invalid_ids = []; vnm_ward_unmatched = []; dk14_skipped = []; cnt_report = []
                # Đọc file visa (nếu có) → map tên → date visa
                if visa_file is not None:
                    try:
                        visa_map = parse_visa_file(visa_file.read())
                    except Exception as _ve:
                        st.warning(f"⚠️ Không đọc được file visa: {_ve}")

                out_files = {}   # tên file → bytes: dùng cho ZIP + nút tải từng file riêng
                _issues = None; _visa_watch = []
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    # ── Xử lý file XLSX (nếu có) ──
                    if has_xlsx:
                        progress.progress(10, text="Quy đổi tỷ giá...")
                        xlsx_bytes = xlsx_file.read()
                        wb, conv = process_xlsx(xlsx_bytes, rate)

                        # Chia giá phòng connecting SAU khi đã quy đổi tỷ giá — ĐƠN GIÁ
                        # nguồn có thể vẫn ở dạng ngoại tệ nhỏ (VD "278.35" EUR) lúc này
                        # còn là số NGUYÊN chưa quy đổi; chia trước rồi mới quy đổi từng
                        # nửa riêng lẻ sẽ làm tròn 2 lần → sai số hàng nghìn đồng so với
                        # quy đổi trước rồi mới chia (đã xảy ra thực tế: 278.35 EUR bị cắt
                        # về 278, chia 139/139, mỗi bên quy đổi ra ~4.150.550đ — lệch hẳn
                        # so với quy đổi cả 278.35 rồi mới chia đúng 8.313.043 → chia đôi).
                        progress.progress(15, text="Chia giá phòng connecting...")
                        wb_bytes_conv = wb_to_bytes(wb)
                        wb_bytes_conv, cnt_report = split_connecting_room_prices(wb_bytes_conv)
                        wb = load_workbook(io.BytesIO(wb_bytes_conv))
                        # Ép các cột dạng mã đọc bằng chuỗi — nếu không, pandas tự suy
                        # diễn cột "trông giống số" thành float, làm mất số 0 đứng đầu
                        # (vd số điện thoại "0912345678" → 912345678.0). An toàn kể cả
                        # khi cột không tồn tại trong file (bỏ qua, không lỗi).
                        _id_cols = {c: str for c in
                                   ('SỐ GIẤY TỜ', 'SỐ ĐIỆN THOẠI', 'SỐ PHÒNG', 'MÃ CHECKIN')}
                        df = pd.read_excel(io.BytesIO(xlsx_bytes), dtype=_id_cols)
                        df_intl = df[df['LOẠI KHÁCH']=='Quốc tế'].reset_index(drop=True)
                        df_vn   = df[df['LOẠI KHÁCH']=='Việt Nam'].reset_index(drop=True)

                        progress.progress(25, text="Kiểm tra chất lượng dữ liệu...")
                        _issues = validate_guests(df)
                        _visa_watch = check_visa_expiry(df_intl, visa_map=visa_map)

                        progress.progress(35, text="Tách file Quốc tế / Việt Nam...")
                        wb_intl = split_wb(wb, 'Quốc tế')
                        wb_vn   = split_wb(wb, 'Việt Nam')

                        progress.progress(50, text="Điền mẫu KBTT...")
                        wb_kbtt, visa_unmatched, visa_source, kbtt_invalid_ids = build_kbtt(df_intl, visa_map=visa_map)

                        progress.progress(65, text="Điền mẫu Thông báo lưu trú VNM...")
                        wb_vnm, gks_cnt, gbl_cnt, vnm_ward_unmatched = build_vnm(df_vn)

                        out_files[f'converted_{date_str}.xlsx']      = wb_to_bytes(wb)
                        out_files[f'KhachQuocTe_{date_str}.xlsx']    = wb_to_bytes(wb_intl)
                        out_files[f'KhachVietNam_{date_str}.xlsx']   = wb_to_bytes(wb_vn)
                        out_files[f'ho_so_KBTT_NNN_{date_str}.xlsx'] = wb_to_bytes(wb_kbtt)
                        out_files[f'thong_bao_luu_tru_VNM_{date_str}.xlsx'] = wb_to_bytes(wb_vnm)
                        files_made += ["📄 converted (file chung)", "🌍 KhachQuocTe", "🇻🇳 KhachVietNam",
                                       "📝 KBTT NNN", "📑 Thông báo lưu trú VNM"]

                    # ── Xử lý file ĐK14 (độc lập, chỉ cần file nguồn IH) ──
                    if xls_file:
                        progress.progress(85, text="Điền mẫu ĐK14...")
                        xls_bytes = xls_file.read()
                        (dk14_bytes, dk_count, dk14_skipped,
                         dk14_issues, dk14_map) = build_dk14(xls_bytes, dk_notifier, dk_checkin)
                        out_files[f'dk14_{date_str}.xlsx'] = dk14_bytes
                        has_dk14 = True
                        files_made.append("🚔 ĐK14")

                    for _ofn, _ofb in out_files.items():
                        zf.writestr(_ofn, _ofb)

                progress.progress(100, text="Hoàn tất!")
                progress.empty()

                # Lưu kết quả vào session — kết quả & nút tải không biến mất sau rerun
                _daily = {'files_made': files_made, 'zip': zip_buf.getvalue(),
                          'files': out_files, 'rate': rate,
                          'date_str': date_str, 'has_xlsx': has_xlsx, 'has_dk14': has_dk14,
                          'dk14_count': dk_count if has_dk14 else None, 'dk14_skipped': dk14_skipped,
                          'dk14_issues': dk14_issues if has_dk14 else [],
                          'dk14_map': dk14_map if has_dk14 else None,
                          'cnt_report': cnt_report}
                if has_xlsx:
                    unknown_nats = []
                    for q in df_intl.get('QUỐC TỊCH', pd.Series([], dtype=str)).dropna().unique():
                        mapped = lookup_nat_kbtt(q)
                        if not _re.match(r'^[A-Z]{2,3} - ', str(mapped)):
                            unknown_nats.append(str(q))
                    # Thống kê phục vụ báo cáo ngày cho quản lý
                    _nat_top = (df.get('QUỐC TỊCH', pd.Series(dtype=str)).dropna().astype(str)
                                .str.strip().value_counts().head(10))
                    _arr_s = pd.to_datetime(df.get('NGÀY ĐẾN'), dayfirst=True, errors='coerce')
                    _dep_col_d = 'NGÀY ÐI' if 'NGÀY ÐI' in df.columns else ('NGÀY ĐI' if 'NGÀY ĐI' in df.columns else None)
                    _avg_nights = None
                    if _dep_col_d:
                        _dep_s = pd.to_datetime(df[_dep_col_d], dayfirst=True, errors='coerce')
                        _nvals = (_dep_s - _arr_s).dt.days.dropna()
                        _nvals = _nvals[_nvals > 0]
                        if len(_nvals):
                            _avg_nights = round(float(_nvals.mean()), 1)
                    _rooms_cnt = int(df.get('SỐ PHÒNG', pd.Series(dtype=str)).dropna().astype(str)
                                     .str.strip().nunique())
                    # Khách check-in / check-out HÔM NAY — đếm theo ngày đến/đi
                    # trùng ngày hiện tại (giờ Việt Nam), phục vụ thẻ tổng quan.
                    _today_ts = pd.Timestamp(today_vn())
                    _checkin_n = int((_arr_s.dt.normalize() == _today_ts).sum()) if _arr_s is not None else 0
                    _checkout_n = 0
                    if _dep_col_d:
                        _checkout_n = int((pd.to_datetime(df[_dep_col_d], dayfirst=True, errors='coerce')
                                           .dt.normalize() == _today_ts).sum())
                    _daily.update({'issues': _issues, 'visa_watch': _visa_watch,
                                   'nat_top': [(str(k), int(v)) for k, v in _nat_top.items()],
                                   'avg_nights': _avg_nights, 'rooms_cnt': _rooms_cnt or None,
                                   'checkin_n': _checkin_n, 'checkout_n': _checkout_n})
                    _daily.update({'total': len(df), 'intl': len(df_intl), 'vn': len(df_vn),
                                   'gks': gks_cnt, 'gbl': gbl_cnt, 'conv': conv,
                                   'unknown_nats': unknown_nats,
                                   'visa_used': visa_source is not None,
                                   'visa_source': visa_source,
                                   'visa_matched': len(df_intl) - len(visa_unmatched) if visa_source else 0,
                                   'visa_unmatched': visa_unmatched,
                                   'visa_skipped_vn': visa_map.get('skipped_vn', 0) if isinstance(visa_map, dict) else 0,
                                   'kbtt_invalid_ids': kbtt_invalid_ids,
                                   'vnm_ward_unmatched': vnm_ward_unmatched})
                st.session_state['daily_results'] = _daily

                def _mark_daily_done(state, _daily=_daily, has_xlsx=has_xlsx, has_dk14=has_dk14):
                    task = {'done': True, 'time': now_vn().strftime('%H:%M:%S'),
                            'has_xlsx': has_xlsx, 'has_dk14': has_dk14}
                    if has_xlsx:
                        iss = _daily.get('issues')
                        task['summary'] = {
                            'total': _daily.get('total'), 'intl': _daily.get('intl'), 'vn': _daily.get('vn'),
                            'gks': _daily.get('gks'), 'gbl': _daily.get('gbl'), 'conv': _daily.get('conv'),
                            'red_issues': int((iss['Mức độ'] == '🔴').sum()) if iss is not None and len(iss) else 0,
                            'yellow_issues': int((iss['Mức độ'] == '🟡').sum()) if iss is not None and len(iss) else 0,
                            'visa_watch_count': len(_daily.get('visa_watch') or []),
                            'rooms_cnt': _daily.get('rooms_cnt'),
                            'checkin_n': _daily.get('checkin_n'), 'checkout_n': _daily.get('checkout_n'),
                        }
                    state.setdefault('tasks', {})['daily'] = task
                _progress_update(_mark_daily_done)
            except Exception as e:
                st.session_state.pop('daily_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    _dr = st.session_state.get('daily_results')
    if _dr:
        st.success("✅ Xử lý hoàn tất!")
        if _dr['has_xlsx']:
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Tổng khách", _dr['total'])
            c2.metric("Quốc tế", _dr['intl'])
            c3.metric("Việt Nam", _dr['vn'])
            c4.metric("GKS + GBL", f"{_dr['gks']} + {_dr['gbl']}")
            st.info(f"💱 Đã quy đổi tỷ giá cho **{_dr['conv']}** ô (tỷ giá {_dr.get('rate', 0):,.2f})")
            _cnt_rep = _dr.get('cnt_report') or []
            if _cnt_rep:
                with st.expander(f"🚪 Đã tự động chia giá {len(_cnt_rep)} cặp phòng connecting", expanded=False):
                    st.caption("Phòng connecting chung mã checkin nhưng chỉ 1 phòng có giá — đã chia đều "
                               "(lệch 1 đồng thì phòng số nhỏ hơn nhận phần làm tròn xuống).")
                    st.dataframe(pd.DataFrame([
                        {'Mã checkin': r['checkin'], 'Phòng A': r['room_a'], 'Giá A (mới)': r['split_a'],
                         'Phòng B': r['room_b'], 'Giá B (mới)': r['split_b'],
                         'Tổng gốc': r['total']} for r in _cnt_rep]),
                        use_container_width=True, hide_index=True)
            if _dr['unknown_nats']:
                st.warning("⚠️ Quốc tịch chưa có mã (giữ nguyên tên, cần kiểm tra): " + ", ".join(_dr['unknown_nats']))
            if _dr.get('visa_used'):
                st.info(f"🛂 Đã điền date visa cho **{_dr['visa_matched']}/{_dr['intl']}** khách quốc tế "
                        f"(khớp theo hộ chiếu/tên từ file Visa rời bạn đã upload).")
                if _dr.get('visa_skipped_vn'):
                    st.caption(f"ℹ️ Đã tự động bỏ qua {_dr['visa_skipped_vn']} khách Việt Nam trong file visa (không cần thời hạn tạm trú).")
                if _dr.get('visa_unmatched'):
                    st.warning("⚠️ Không tìm thấy date visa cho (cột tạm trú để trống): "
                               + ", ".join(_dr['visa_unmatched']))
            if _dr.get('kbtt_invalid_ids'):
                st.warning("⚠️ Số giấy tờ chỉ là mã tạm (chưa có hộ chiếu thật) — đã để trống trong hồ sơ KBTT, "
                           "cần bổ sung số hộ chiếu thật trước khi nộp cho: " + ", ".join(_dr['kbtt_invalid_ids']))
            if _dr.get('vnm_ward_unmatched'):
                st.warning("⚠️ Không tự tra được mã Phường/Xã (đã giữ nguyên chữ gốc trong file VNM, "
                           "cần chọn lại theo danh mục): "
                           + ", ".join(f"{n} ('{p}')" for n, p in _dr['vnm_ward_unmatched']))

            # ── Kiểm tra chất lượng dữ liệu trước khi nộp hồ sơ ──
            _iss = _dr.get('issues')
            if _iss is not None and len(_iss):
                _n_red = int((_iss['Mức độ'] == '🔴').sum())
                _n_yel = int((_iss['Mức độ'] == '🟡').sum())
                st.write("")
                st.markdown('<div class="section-label">✅ Kiểm tra dữ liệu trước khi nộp hồ sơ</div>',
                           unsafe_allow_html=True)
                if _n_red:
                    st.error(f"🔴 **{_n_red}** vấn đề cần sửa trước khi nộp hồ sơ công an "
                             f"+ 🟡 **{_n_yel}** vấn đề nên kiểm tra lại.")
                else:
                    st.warning(f"🟡 **{_n_yel}** vấn đề nên kiểm tra lại (không chặn nộp hồ sơ).")
                st.dataframe(_iss, use_container_width=True, hide_index=True)
            elif _iss is not None:
                st.success("✅ Kiểm tra dữ liệu: không phát hiện vấn đề nào.")

            # ── Cảnh báo visa/tạm trú sắp hết hạn ──
            _vw = _dr.get('visa_watch') or []
            if _vw:
                st.write("")
                st.markdown('<div class="section-label">🛂 Cảnh báo hạn tạm trú / visa</div>', unsafe_allow_html=True)
                _vw_days = st.slider("Cảnh báo khách còn lưu trú mà visa hết hạn trong vòng (ngày)",
                                     1, 30, 3, key="visa_warn_days")
                _cutoff = today_vn() + datetime.timedelta(days=_vw_days)
                _soon = []
                for _v in _vw:
                    _vd = datetime.date.fromisoformat(_v['visa'])
                    _dep = datetime.date.fromisoformat(_v['dep']) if _v.get('dep') else None
                    # chỉ cảnh báo khách còn ở (chưa checkout trước ngày visa hết hạn)
                    if _vd <= _cutoff and (_dep is None or _dep >= today_vn()):
                        _soon.append({'Phòng': _v['room'], 'Họ tên': _v['name'], 'Quốc tịch': _v['nat'],
                                     'Hết hạn tạm trú': _vd.strftime('%d/%m/%Y'),
                                     'Còn': (_vd - today_vn()).days,
                                     'Ngày đi dự kiến': (_dep.strftime('%d/%m/%Y') if _dep else '—')})
                if _soon:
                    _df_soon = pd.DataFrame(_soon).sort_values('Còn')
                    st.error(f"🔴 **{len(_soon)}** khách cần gia hạn/rời đi trước khi tạm trú hết hạn:")
                    st.dataframe(_df_soon, use_container_width=True, hide_index=True)
                else:
                    st.success(f"✅ Không có khách nào hết hạn tạm trú trong {_vw_days} ngày tới.")
        elif _dr['has_dk14']:
            st.info("ℹ️ Chỉ tạo file ĐK14 (không có file XLSX dữ liệu khách).")

        if _dr.get('has_dk14'):
            _dk_iss = _dr.get('dk14_issues') or []
            _n_skip = len(_dr.get('dk14_skipped') or [])
            _n_fix = sum(1 for t, *_ in _dk_iss if t == 'fix')
            _n_warn = sum(1 for t, *_ in _dk_iss if t == 'warn')
            _m1, _m2, _m3, _m4 = st.columns(4)
            _m1.metric("Tổng khách vào sổ", _dr['dk14_count'])
            _m2.metric("Đã tự sửa", _n_fix)
            _m3.metric("Cần xem lại", _n_warn)
            _m4.metric("Bỏ qua", _n_skip)

            if _dr.get('dk14_map'):
                _dcols, _dheads = _dr['dk14_map']
                _LBL = [('name', 'B — Họ và tên'), ('dob', 'C/D — Ngày sinh (tách Nam/Nữ)'),
                        ('gender', '(dùng để tách ngày sinh Nam/Nữ)'), ('nationality', 'E — Quốc tịch'),
                        ('country', 'E — Quốc tịch (dự phòng)'), ('id', 'F — Số giấy tờ'),
                        ('cuTru', 'G — Loại cư trú'), ('tinh', 'G — Tỉnh'), ('quan', 'G — Quận'),
                        ('phuong', 'G — Phường'), ('addressDetail', 'G — Địa chỉ chi tiết'),
                        ('arrival', 'H — Thời gian đến'), ('departure', 'I — Thời gian đi'),
                        ('room', 'J — Số phòng'), ('notifier', 'K — Người thông báo')]
                with st.expander("🔎 Cột đã tự nhận diện trong file nguồn", expanded=False):
                    for _k, _lb in _LBL:
                        _i = _dcols.get(_k)
                        if _i is None:
                            continue
                        _h = str(_dheads[_i]).strip() if _i < len(_dheads) and _dheads[_i] is not None else '(?)'
                        st.markdown(f"- cột **{chr(65 + _i)}** “{_h}” → ĐK14 **{_lb}**")
                    _miss = [k for k in ('name', 'dob', 'gender', 'id', 'arrival', 'departure', 'room')
                             if k not in _dcols]
                    if _miss:
                        st.warning("⚠️ Không tìm thấy cột: " + ", ".join(_miss))

            if _dk_iss:
                _ICON = {'fix': '🔧', 'warn': '⚠️', 'skip': '⏭️'}
                with st.expander(f"📋 Nhật ký {len(_dk_iss)} dòng đã sửa / cần xem lại", expanded=bool(_n_warn)):
                    st.dataframe(pd.DataFrame(
                        [{'': _ICON.get(t, ''), 'Dòng': r, 'Tên': n, 'Nội dung': msg}
                         for t, r, n, msg in _dk_iss]),
                        use_container_width=True, hide_index=True)
            else:
                st.success("✅ ĐK14: không phát hiện vấn đề nào trong dữ liệu nguồn.")

        st.markdown("**File đã tạo:** " + " · ".join(_dr['files_made']))

        st.download_button(
            label="⬇️ Tải về tất cả file (ZIP)",
            data=_dr['zip'],
            file_name=f"hotel_{_dr['date_str']}.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )

        # ── Tải riêng từng file (không cần giải nén ZIP) ──
        if _dr.get('files'):
            with st.expander("⬇️ Tải riêng từng file"):
                for _fn, _fb in _dr['files'].items():
                    st.download_button(f"⬇️ {_fn}", _fb, file_name=_fn,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"dl_single_{_fn}")

# ── Regcard screen ────────────────────────────────────────────────────────
if st.session_state.menu == "regcard":
    st.write("")
    st.markdown('<div class="section-label">🖨️ Tạo Registration Card + file ARR</div>', unsafe_allow_html=True)
    st.caption("Điền dữ liệu từ file Arrival Smile lên mẫu Regcard PDF gốc, đồng thời tạo file ARR "
               "(đúng định dạng ARR Converter — nhóm Conf# · đếm phòng · Cà Thẻ / Thu Tiền / Xem Lại BU / FOC Late C/O) — ra cùng lúc 2 file.")

    with st.container(border=True):
        rc_file = st.file_uploader("File Excel dữ liệu booking (.xlsx)", type=['xlsx'], key="rc_xlsx")

        only_main = st.checkbox("Chỉ tạo cho khách chính (có mã Conf#)", value=True,
                                help="Bỏ chọn để tạo regcard cho tất cả khách, kể cả khách đi cùng phòng")

    st.write("")

    if st.button("🖨️ Tạo Regcard PDF + file ARR", type="primary", disabled=rc_file is None, use_container_width=True):
        with st.spinner("Đang tạo Regcard PDF và file ARR..."):
            try:
                rc_bytes = rc_file.read()
                pdf_data, count = build_regcards(rc_bytes, only_main=only_main)

                # Tạo file ARR từ cùng file đầu vào (lỗi ARR không làm hỏng PDF)
                arr_bytes, arr_stats, arr_err = None, None, None
                try:
                    wb_arr, arr_stats = build_arr(rc_bytes)
                    arr_bytes = wb_to_bytes(wb_arr)
                except Exception as _e:
                    arr_err = str(_e)

                st.session_state['rc_results'] = {
                    'pdf': pdf_data, 'count': count,
                    'arr': arr_bytes, 'arr_stats': arr_stats, 'arr_err': arr_err,
                    'date': today_vn().strftime('%d_%m'),
                    'arr_date': today_vn().strftime('%d.%m.%Y'),
                }

                def _mark_regcard_done(state, count=count, arr_stats=arr_stats):
                    task = {'done': True, 'time': now_vn().strftime('%H:%M:%S'), 'regcards': count}
                    if arr_stats:
                        task['summary'] = {k: arr_stats.get(k) for k in
                                           ('bookings', 'rooms', 'ota', 'ca_the', 'thu_tien', 'xem_lai_bu', 'foc_lco')}
                    state.setdefault('tasks', {})['regcard'] = task
                _progress_update(_mark_regcard_done)
            except Exception as e:
                st.session_state.pop('rc_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    # Kết quả lưu trong session — 2 nút tải không biến mất sau khi bấm 1 nút
    _res = st.session_state.get('rc_results')
    if _res:
        if _res['count'] == 0:
            st.warning("⚠️ Không tìm thấy khách nào để tạo regcard. Kiểm tra lại file.")
        else:
            st.success(f"✅ Đã tạo {_res['count']} regcard"
                       + (" + file ARR!" if _res['arr'] else "!"))
            if _res['arr_stats']:
                a1, a2, a3, a4, a5 = st.columns(5)
                a1.metric("🖨️ Regcard", _res['count'])
                a2.metric("📦 Booking", _res['arr_stats']['bookings'])
                a3.metric("💳 Cà Thẻ", _res['arr_stats']['ca_the'])
                a4.metric("💵 Thu Tiền", _res['arr_stats']['thu_tien'])
                a5.metric("⚠️ Xem Lại BU", _res['arr_stats']['xem_lai_bu'])
                if _res['arr_stats'].get('foc_lco'):
                    st.caption(f"🛎️ FOC Late C/O: {_res['arr_stats']['foc_lco']} booking")
            else:
                st.metric("Số regcard", _res['count'])

            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    label=f"⬇️ Tải {_res['count']} Regcard (PDF)",
                    data=_res['pdf'],
                    file_name=f"regcards_{_res['date']}.pdf",
                    mime="application/pdf",
                    use_container_width=True, type="primary", key="dl_rc_pdf")
            with d2:
                if _res['arr']:
                    st.download_button(
                        label="⬇️ Tải file ARR (Excel)",
                        data=_res['arr'],
                        file_name=f"Arr {_res['arr_date']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary", key="dl_rc_arr")
            if _res['arr_err']:
                st.warning(f"⚠️ Không tạo được file ARR: {_res['arr_err']}")


# ── Sổ giao ca ─────────────────────────────────────────────────────────────
if st.session_state.menu == "handover":
    st.write("")
    st.markdown('<div class="section-label">🤝 Sổ giao ca</div>', unsafe_allow_html=True)

    _db_on = db_available()

    if _db_on:
        st.caption("Ghi chú trong ca (khách nợ, yêu cầu đặc biệt, sự cố, đồ thất lạc...) để ca sau nắm được. "
                   "Lưu trên **đám mây (Supabase)** — không mất khi app deploy lại, xem lại được mọi ngày trong quá khứ.")

        with st.container(border=True):
            hc1, hc2, hc3 = st.columns([1, 1, 1])
            with hc1:
                h_shift = st.selectbox("Ca trực", ["Ca sáng", "Ca chiều", "Ca đêm"], key="h_shift")
            with hc2:
                h_staff = st.text_input("Tên lễ tân trực", key="h_staff", placeholder="VD: Tân")
            with hc3:
                _avail_dates = db_load_dates()
                _sel_date = st.date_input("📅 Xem ngày", value=today_vn(),
                                          max_value=today_vn(), key="h_view_date",
                                          help="Chọn lại ngày trong quá khứ để xem lịch sử sổ giao ca")
        _is_today = _sel_date == today_vn()

        _df_e = db_load_entries(_sel_date)
        _summary = compute_day_summary(_df_e)

        st.write("")
        if not _is_today:
            st.info(f"📜 Đang xem lại lịch sử ngày **{_sel_date.strftime('%d/%m/%Y')}** (chỉ xem, "
                    "không thêm/xóa được — quay lại hôm nay để ghi chú mới).")

        st.markdown('<div class="section-label">📊 Tổng hợp tự động</div>', unsafe_allow_html=True)
        if _summary['total'] == 0:
            st.caption(f"Chưa có ghi chú nào ngày {_sel_date.strftime('%d/%m/%Y')}.")
        else:
            sc1, sc2 = st.columns([1, 2])
            with sc1:
                st.metric("Tổng ghi chú", _summary['total'])
            with sc2:
                _cat_txt = " · ".join(f"{k}: {v}" for k, v in _summary['by_category'].items())
                st.markdown(f"**Theo phân loại:** {_cat_txt}")
            if _summary['rooms']:
                st.caption("🚪 Phòng được nhắc tới: " + ", ".join(_summary['rooms']))

        if _is_today:
            st.write("")
            with st.container(border=True):
                with st.form("handover_add_form", clear_on_submit=True):
                    fc1, fc2 = st.columns([1, 1])
                    with fc1:
                        h_cat = st.selectbox("Phân loại", ["Khách nợ", "Yêu cầu đặc biệt", "Sự cố",
                                                             "Đồ thất lạc", "Bảo trì", "Khác"], key="h_cat")
                    with fc2:
                        h_room = st.text_input("Số phòng (nếu có)", key="h_room")
                    h_note = st.text_area("Nội dung bàn giao", key="h_note", height=80)
                    h_submit = st.form_submit_button("➕ Thêm vào sổ giao ca", type="primary", use_container_width=True)
                    if h_submit:
                        if not h_note.strip():
                            st.warning("⚠️ Vui lòng nhập nội dung bàn giao.")
                        else:
                            db_add_entry(today_vn(), now_vn().strftime('%H:%M'),
                                        h_cat, h_room.strip(), h_note.strip())
                            st.rerun()

        st.write("")
        if not _df_e.empty:
            st.markdown(f"**{len(_df_e)} ghi chú ngày {_sel_date.strftime('%d/%m/%Y')}**")
            for _, _row in _df_e.iterrows():
                with st.container(border=True):
                    ic1, ic2 = st.columns([10, 1])
                    with ic1:
                        _room_txt = f" · 🚪 Phòng {_row['room']}" if _row['room'] else ""
                        st.markdown(f"🕐 **{_row['entry_time']}** · 🏷️ {_row['category']}{_room_txt}")
                        st.write(_row['note'])
                    with ic2:
                        if _is_today and st.button("🗑️", key=f"h_del_{_row['id']}", help="Xóa ghi chú này"):
                            db_delete_entry(_row['id'])
                            st.rerun()

            st.write("")
            _entries_for_xlsx = [{'time': r['entry_time'], 'cat': r['category'],
                                  'room': r['room'] or '', 'note': r['note']}
                                 for _, r in _df_e.iloc[::-1].iterrows()]
            _wb_ho = build_handover_xlsx(
                {'date': _sel_date.strftime('%d/%m/%Y'),
                 'shift': h_shift if _is_today else '', 'staff': h_staff if _is_today else ''},
                _entries_for_xlsx)
            st.download_button(f"⬇️ Tải sổ giao ca ngày {_sel_date.strftime('%d_%m')} (Excel)", wb_to_bytes(_wb_ho),
                               file_name=f"giao_ca_{_sel_date.strftime('%d_%m_%Y')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, type="primary", key="dl_handover")
        elif _is_today:
            st.info("Chưa có ghi chú nào trong ca này. Thêm ghi chú ở form phía trên.")

    else:
        # ── Chưa cấu hình Supabase — dùng lại lưu tạm trên đĩa server (chỉ sống
        # trong ngày, mất khi deploy lại). Xem secrets.toml.example để kết nối
        # lưu trữ đám mây bền vững + xem lại lịch sử nhiều ngày.
        st.caption("Ghi chú trong ca (khách nợ, yêu cầu đặc biệt, sự cố, đồ thất lạc...) để ca sau nắm được. "
                   "Tự động lưu trên server theo ngày — mở lại/tải lại trang trong ngày vẫn còn nguyên. "
                   "Bấm **Tải file Excel** cuối trang khi cần in hoặc lưu trữ lâu dài.")
        st.info("☁️ **Chưa kết nối lưu trữ đám mây** — ghi chú chỉ lưu tạm trên server, sẽ mất khi app deploy lại "
                "và không xem lại được các ngày trước. Xem hướng dẫn kết nối Supabase trong `secrets.toml.example` "
                "để lưu trữ bền vững + xem lại lịch sử nhiều ngày.")
        _db_err = st.session_state.get('_db_last_error')
        if _db_err:
            with st.expander("🔍 Xem lý do kết nối thất bại (để tự sửa secrets)"):
                st.code(_db_err, language=None)
        if st.button("🔄 Thử kết nối lại", key="db_retry",
                    help="Bấm sau khi đã sửa Secrets — không cần Reboot cả app"):
            _db_schema_ready.clear()
            st.rerun()

        with st.container(border=True):
            hc1, hc2 = st.columns(2)
            with hc1:
                h_shift = st.selectbox("Ca trực", ["Ca sáng", "Ca chiều", "Ca đêm"], key="h_shift")
            with hc2:
                h_staff = st.text_input("Tên lễ tân trực", key="h_staff", placeholder="VD: Tân")

            st.write("")
            with st.form("handover_add_form", clear_on_submit=True):
                fc1, fc2 = st.columns([1, 1])
                with fc1:
                    h_cat = st.selectbox("Phân loại", ["Khách nợ", "Yêu cầu đặc biệt", "Sự cố",
                                                         "Đồ thất lạc", "Bảo trì", "Khác"], key="h_cat")
                with fc2:
                    h_room = st.text_input("Số phòng (nếu có)", key="h_room")
                h_note = st.text_area("Nội dung bàn giao", key="h_note", height=80)
                h_submit = st.form_submit_button("➕ Thêm vào sổ giao ca", type="primary", use_container_width=True)
                if h_submit:
                    if not h_note.strip():
                        st.warning("⚠️ Vui lòng nhập nội dung bàn giao.")
                    else:
                        _new_entry = {'time': now_vn().strftime('%H:%M'),
                                      'cat': h_cat, 'room': h_room.strip(), 'note': h_note.strip()}
                        _progress_update(lambda state: state.setdefault('handover_entries', []).append(_new_entry))
                        st.session_state.handover['entries'].append(_new_entry)
                        st.rerun()

        _entries = st.session_state.handover['entries']
        st.write("")
        if _entries:
            st.markdown(f"**{len(_entries)} ghi chú trong ca này**")
            for _ei, _e in enumerate(reversed(_entries)):
                _real_i = len(_entries) - 1 - _ei
                with st.container(border=True):
                    ic1, ic2 = st.columns([10, 1])
                    with ic1:
                        _room_txt = f" · 🚪 Phòng {_e['room']}" if _e['room'] else ""
                        st.markdown(f"🕐 **{_e['time']}** · 🏷️ {_e['cat']}{_room_txt}")
                        st.write(_e['note'])
                    with ic2:
                        if st.button("🗑️", key=f"h_del_{_real_i}", help="Xóa ghi chú này"):
                            _target = _e
                            def _m(state, _target=_target):
                                entries = state.setdefault('handover_entries', [])
                                if _target in entries:
                                    entries.remove(_target)
                            _progress_update(_m)
                            st.session_state.handover['entries'] = list(
                                st.session_state.progress.get('handover_entries', []))
                            st.rerun()

            st.write("")
            _wb_ho = build_handover_xlsx(
                {'date': today_vn().strftime('%d/%m/%Y'), 'shift': h_shift, 'staff': h_staff},
                _entries)
            st.download_button("⬇️ Tải sổ giao ca (Excel)", wb_to_bytes(_wb_ho),
                               file_name=f"giao_ca_{today_vn().strftime('%d_%m')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, type="primary", key="dl_handover")
        else:
            st.info("Chưa có ghi chú nào trong ca này. Thêm ghi chú ở form phía trên.")


# ── Đối chiếu: sub-menu 2 lựa chọn (có cổng mật khẩu riêng) ────────────────
try:
    RECON_PASS = st.secrets.get("recon_pass", "368736")
except Exception:
    RECON_PASS = "368736"
if "recon_ok" not in st.session_state:
    st.session_state.recon_ok = False

def _check_recon():
    if st.session_state.get("recon_pass_input", "") == RECON_PASS:
        st.session_state.recon_ok = True
        st.session_state.recon_pass_err = False
    else:
        st.session_state.recon_pass_err = True

if st.session_state.menu == "recon":
    st.write("")

    # Cổng mật khẩu cho tính năng đối chiếu
    if not st.session_state.recon_ok:
        st.markdown('<div class="section-label">🔒 Nhập mật khẩu để truy cập</div>', unsafe_allow_html=True)
        st.caption("Tính năng Đối chiếu lưu trú được bảo vệ bằng mật khẩu riêng.")
        with st.form("recon_pass_form", clear_on_submit=False):
            st.text_input("Mật khẩu", key="recon_pass_input", type="password", placeholder="Nhập mật khẩu")
            ok = st.form_submit_button("Mở khóa →", type="primary", use_container_width=True)
            if ok:
                _check_recon()
                if st.session_state.recon_ok:
                    st.session_state.menu = "recon_person"
                    st.rerun()
        if st.session_state.get("recon_pass_err"):
            st.error("❌ Mật khẩu không đúng!")
        st.stop()

    st.info("Đã mở khóa — chọn công cụ Đối chiếu ở sidebar bên trái.")

# ── Kiểm tra lưu trú người nước ngoài ─────────────────────────────────────
if st.session_state.menu == "recon_person":
    st.write("")
    st.markdown('<div class="section-label">🌏 Kiểm tra lưu trú người nước ngoài</div>', unsafe_allow_html=True)
    st.caption("So khớp khách Smile vs Trang quản lý người nước ngoài theo số hộ chiếu — tìm khách chưa đăng ký / đăng ký trùng.")

    rc1, rc2 = st.columns(2)
    with rc1:
        smile_file = st.file_uploader("File Smile — Inhouse (.xlsx)", type=['xlsx'], key="recon_smile")
    with rc2:
        luutru_file = st.file_uploader("File Trang lưu trú người nước ngoài (.xlsx)", type=['xlsx'], key="recon_luutru")

    today_str = st.text_input("📅 Ngày xuất file (hôm nay)", value=today_vn().strftime('%d/%m/%Y'),
                              help="Dùng để loại bỏ: khách arrival hôm nay (Smile) và khách ngày đi dự kiến hôm nay (Lưu trú)")

    st.write("")

    if st.button("🔍 Bắt đầu kiểm tra", type="primary",
                 disabled=(smile_file is None or luutru_file is None), use_container_width=True):
        with st.spinner("Đang đối chiếu..."):
            try:
                today = pd.to_datetime(today_str, format='%d/%m/%Y')
                _rp = reconcile(smile_file.read(), luutru_file.read(), today)
                st.session_state['recon_results'] = _rp

                def _mark_recon_person_done(state, _rp=_rp):
                    state.setdefault('tasks', {})['recon_person'] = {
                        'done': True, 'time': now_vn().strftime('%H:%M:%S'),
                        'summary': {'chua_dang_ky': len(_rp.get('chua_dk', [])),
                                   'thua': len(_rp.get('thua', [])), 'trung': len(_rp.get('dup', []))}}
                _progress_update(_mark_recon_person_done)
            except Exception as e:
                st.session_state.pop('recon_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    r = st.session_state.get('recon_results')
    if r:
        st.success("✅ Kiểm tra hoàn tất!")

        c1, c2 = st.columns(2)
        c1.metric("Smile (sau lọc)", r['smile_filtered'], f"từ {r['smile_total']} (bỏ VNM + arrival/departure hôm nay)")
        c2.metric("Lưu trú (sau lọc)", r['luutru_filtered'], f"từ {r['luutru_total']} (bỏ đi dự kiến hôm nay)")

        st.divider()
        st.markdown("### 👤 Kết quả đối chiếu người (theo số hộ chiếu)")

        n_chua = len(r['chua_dk']); n_thua = len(r['thua']); n_dup = len(r['dup'])

        # Bảng tổng hợp chênh lệch
        m1, m2, m3 = st.columns(3)
        m1.metric("🔴 Chưa đăng ký", n_chua)
        m2.metric("🟡 Có/lưu trú, thiếu/Smile", n_thua)
        m3.metric("🟠 Đăng ký trùng", n_dup)

        if n_chua == 0 and n_thua == 0 and n_dup == 0:
            st.success("✅ Khớp hoàn toàn! Không có ai thiếu/thừa/trùng.")

        if n_chua > 0:
            st.error(f"🔴 {n_chua} khách trên Smile nhưng CHƯA đăng ký lưu trú:")
            st.dataframe(r['chua_dk'], use_container_width=True, hide_index=True)
        else:
            st.success("✅ Không có khách nào chưa đăng ký lưu trú.")

        if n_thua > 0:
            st.warning(f"🟡 Chênh lệch **{n_thua} người**: có trên Trang quản lý người nước ngoài nhưng KHÔNG có trên Smile (có thể đã checkout nhưng chưa xóa khỏi lưu trú):")
            st.dataframe(r['thua'], use_container_width=True, hide_index=True)
            # Nút tải danh sách chênh lệch
            _csv = r['thua'].to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Tải danh sách chênh lệch (CSV)", _csv,
                               file_name="chenh_lech_luu_tru.csv", mime="text/csv")

        if n_dup > 0:
            st.warning(f"🟠 {n_dup} dòng ĐĂNG KÝ TRÙNG trên lưu trú:")
            st.dataframe(r['dup'], use_container_width=True, hide_index=True)

# ── Kiểm tra hệ thống quản lý lưu trú phòng ────────────────────────────────
def reconcile_rooms(smile_bytes, room_bytes, today):
    """Đối chiếu phòng inhouse từ file khách lưu trú Smile (trừ khách Arrival hôm nay)
    với file Excel chỉ chứa danh sách số phòng."""
    from collections import Counter

    # ── Đọc file khách lưu trú Smile ──
    df1 = pd.read_excel(io.BytesIO(smile_bytes), header=0)
    if 'Rm#' not in df1.columns:
        raise ValueError("File Smile không có cột 'Rm#'. Vui lòng dùng file khách lưu trú xuất từ Smile.")
    smile = df1.dropna(subset=['Rm#']).copy()
    smile['room'] = smile['Rm#'].apply(_norm_room)
    smile['Arrival'] = pd.to_datetime(smile['Arrival'], errors='coerce') if 'Arrival' in smile else pd.NaT
    _ln = smile['Last Name'].astype(str).str.strip() if 'Last Name' in smile else ''
    _fn = smile['First Name'].astype(str).str.strip() if 'First Name' in smile else ''
    smile['name'] = (_ln + ' ' + _fn).str.strip() if 'Last Name' in smile else ''
    smile_total = len(smile)
    # Cột Departure (dò tên linh hoạt)
    _dep_col = next((c for c in df1.columns if 'depart' in str(c).lower()), None)
    smile['Departure'] = pd.to_datetime(df1[_dep_col], errors='coerce') if _dep_col else pd.NaT
    # Trừ khách Arrival = hôm nay và khách Departure = hôm nay (trả phòng)
    if 'Arrival' in smile:
        smile_f = smile[(smile['Arrival'].dt.date != today.date()) &
                        (smile['Departure'].dt.date != today.date())].copy()
    else:
        smile_f = smile.copy()

    # ── Đọc file chỉ chứa số phòng: lấy tất cả ô có dữ liệu ──
    raw = pd.read_excel(io.BytesIO(room_bytes), header=None, dtype=str)
    rooms_sys = []
    for _, row_vals in raw.iterrows():
        for v in row_vals:
            if pd.isna(v): continue
            r = _norm_room(v)
            if not r: continue
            # bỏ ô tiêu đề nếu lỡ có (vd "Số phòng", "Room", "STT")
            if _norm_nat(r) in ('sophong', 'phong', 'room', 'rm', 'stt'): continue
            rooms_sys.append(r)
    if not rooms_sys:
        raise ValueError("File số phòng không có dữ liệu. Vui lòng kiểm tra lại file.")

    sys_rooms = set(rooms_sys)
    _cnt = Counter(rooms_sys)
    sys_dup = sorted((r for r, c in _cnt.items() if c > 1), key=lambda x: (len(x), x))

    import re as _re3
    def _is_virtual(r):
        return bool(_re3.fullmatch(r'9\d{3}', r))  # phòng ảo 9000-9999 (posting master)

    smile_rooms = set(r for r in smile_f['room'] if r and not _is_virtual(r))
    sys_rooms = set(r for r in sys_rooms if not _is_virtual(r))

    def _sortkey(x): return (len(x), x)
    room_chua = sorted(smile_rooms - sys_rooms, key=_sortkey)  # inhouse nhưng CHƯA có trong file phòng
    room_thua = sorted(sys_rooms - smile_rooms, key=_sortkey)  # có trong file phòng nhưng KHÔNG còn inhouse

    # Chi tiết khách trong các phòng chưa đăng ký (tiện đăng ký bổ sung)
    if room_chua:
        detail = smile_f[smile_f['room'].isin(room_chua)].copy()
        detail['_arr'] = detail['Arrival'].dt.strftime('%d/%m/%Y') if 'Arrival' in detail else ''
        _nat = detail['NAT'] if 'NAT' in detail else ''
        detail = pd.DataFrame({
            'Số phòng': detail['room'].values,
            'Họ tên': detail['name'].values,
            'Quốc tịch': _nat.values if hasattr(_nat, 'values') else _nat,
            'Ngày đến': detail['_arr'].values if hasattr(detail['_arr'], 'values') else '',
        }).sort_values('Số phòng', key=lambda s: s.map(lambda x: (len(x), x)))
    else:
        detail = pd.DataFrame(columns=['Số phòng', 'Họ tên', 'Quốc tịch', 'Ngày đến'])

    return {
        'smile_total': smile_total, 'smile_filtered': len(smile_f),
        'sys_total': len(rooms_sys), 'sys_unique': len(sys_rooms),
        'smile_rooms': len(smile_rooms),
        'room_chua': room_chua, 'room_thua': room_thua, 'sys_dup': sys_dup,
        'room_match': len(smile_rooms & sys_rooms),
        'detail_chua': detail,
    }


if st.session_state.menu == "recon_room":
    st.write("")
    st.markdown('<div class="section-label">🚪 Kiểm tra hệ thống quản lý lưu trú phòng</div>', unsafe_allow_html=True)
    st.caption("So khớp số phòng inhouse từ file khách lưu trú Smile với file danh sách số phòng — tìm phòng chưa đăng ký / thừa / trùng.")

    rr1, rr2 = st.columns(2)
    with rr1:
        smile_file_r = st.file_uploader("File khách lưu trú Smile (.xlsx)", type=['xlsx'], key="reconr_smile")
    with rr2:
        room_file = st.file_uploader("File số phòng (.xlsx — chỉ chứa danh sách số phòng)", type=['xlsx'], key="reconr_room")

    today_str_r = st.text_input("📅 Ngày xuất file (hôm nay)", value=today_vn().strftime('%d/%m/%Y'),
                                key="reconr_today",
                                help="Khách có Arrival = ngày này trên Smile sẽ được loại bỏ khỏi đối chiếu")

    st.write("")

    if st.button("🔍 Bắt đầu kiểm tra", type="primary", key="reconr_run",
                 disabled=(smile_file_r is None or room_file is None), use_container_width=True):
        with st.spinner("Đang đối chiếu phòng..."):
            try:
                today_r = pd.to_datetime(today_str_r, format='%d/%m/%Y')
                _rr = reconcile_rooms(smile_file_r.read(), room_file.read(), today_r)
                st.session_state['reconr_results'] = _rr

                def _mark_recon_room_done(state, _rr=_rr):
                    state.setdefault('tasks', {})['recon_room'] = {
                        'done': True, 'time': now_vn().strftime('%H:%M:%S'),
                        'summary': {'chua_dang_ky': len(_rr.get('room_chua', [])),
                                   'thua': len(_rr.get('room_thua', [])), 'trung': len(_rr.get('sys_dup', []))}}
                _progress_update(_mark_recon_room_done)
            except Exception as e:
                st.session_state.pop('reconr_results', None)
                st.error(f"❌ Lỗi: {e}")
                st.exception(e)

    rr = st.session_state.get('reconr_results')
    if rr:
        st.success("✅ Kiểm tra hoàn tất!")

        c1, c2, c3 = st.columns(3)
        c1.metric("Phòng inhouse (Smile)", rr['smile_rooms'],
                  f"{rr['smile_filtered']} khách (từ {rr['smile_total']}, đã trừ arrival hôm nay)")
        c2.metric("Phòng trong file", rr['sys_unique'],
                  (f"{rr['sys_total']} dòng" if rr['sys_total'] != rr['sys_unique'] else None))
        c3.metric("🟢 Phòng khớp", rr['room_match'])

        st.divider()
        st.markdown("### 🚪 Kết quả đối chiếu phòng")

        n_chua = len(rr['room_chua']); n_thua = len(rr['room_thua']); n_dup = len(rr['sys_dup'])

        m1, m2, m3 = st.columns(3)
        m1.metric("🔴 Chưa đăng ký", n_chua)
        m2.metric("🟡 Thừa trong file", n_thua)
        m3.metric("🟠 Trùng trong file", n_dup)

        if n_chua == 0 and n_thua == 0 and n_dup == 0:
            st.success("✅ Khớp hoàn toàn! Không có phòng thiếu/thừa/trùng.")
            st.balloons()

        if n_chua > 0:
            st.error(f"🔴 {n_chua} phòng có khách inhouse nhưng CHƯA có trong file số phòng: "
                     + ", ".join(rr['room_chua']))
            st.markdown("**Chi tiết khách trong các phòng chưa đăng ký:**")
            st.dataframe(rr['detail_chua'], use_container_width=True, hide_index=True)
            _csv_r = rr['detail_chua'].to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Tải danh sách phòng chưa đăng ký (CSV)", _csv_r,
                               file_name="phong_chua_dang_ky.csv", mime="text/csv")
        else:
            st.success("✅ Tất cả phòng inhouse đều đã có trong file số phòng.")

        if n_thua > 0:
            st.warning(f"🟡 {n_thua} phòng có trong file nhưng KHÔNG còn khách inhouse "
                       f"(có thể đã checkout nhưng chưa gỡ): "
                       + ", ".join(rr['room_thua']))

        if n_dup > 0:
            st.warning(f"🟠 {n_dup} phòng bị TRÙNG (xuất hiện nhiều lần) trong file số phòng: "
                       + ", ".join(rr['sys_dup']))

# ── Hiệu ứng mượt khi ĐỔI CÔNG CỤ (sidebar) ─────────────────────────────────
# Đặt Ở CUỐI file (sau mọi khối `if st.session_state.menu == ...`) — không
# phải ngẫu nhiên: đây là lệnh cuối cùng được gửi cho mỗi lượt rerun, nên khi
# JS bên dưới chạy, toàn bộ nội dung trang MỚI (đúng công cụ vừa chuyển sang)
# chắc chắn đã render xong. Đặt sớm hơn (như khối set data-theme ở trên) sẽ
# có nguy cơ JS thao tác nhầm lên nội dung TRANG CŨ vẫn còn trên DOM.
# So sánh menu hiện tại với menu LƯU TRÊN CHÍNH TRANG CHA (không phải session
# JS nội bộ iframe — iframe bị tạo mới mỗi lượt rerun, không nhớ được gì) để
# chỉ phát hiệu ứng khi THỰC SỰ đổi công cụ, không phát khi rerun vì lý do
# khác (gõ chữ, tick ô, mở expander...) — tránh nhấp nháy phiền mỗi thao tác.
st.iframe(f"""
<script>
(function(){{
    var doc = window.parent.document;
    var cur = {st.session_state.menu!r};
    var prev = doc.documentElement.getAttribute('data-tan-menu');
    if (prev !== null && prev !== cur) {{
        var main = doc.querySelector('[data-testid="stMainBlockContainer"]');
        if (main) {{
            main.classList.remove('tan-page-in');
            void main.offsetWidth;
            main.classList.add('tan-page-in');
        }}
    }}
    doc.documentElement.setAttribute('data-tan-menu', cur);
}})();
</script>
""", height=1)


